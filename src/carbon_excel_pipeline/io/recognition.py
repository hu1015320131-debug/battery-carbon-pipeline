"""WP6-1 shared workbook, sheet, header and semantic recognition core."""

from __future__ import annotations

from collections import Counter
from typing import Any

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from carbon_excel_pipeline.io.recognition_models import (
    HeaderCandidateResult,
    MappingStatus,
    MatchType,
    RecognitionStatus,
    RecognitionWarning,
    SheetRecognitionResult,
    WorkbookRecognitionResult,
)
from carbon_excel_pipeline.io.semantic_registry import SemanticFieldRegistry


def _has_data_evidence(
    worksheet: Worksheet,
    *,
    header_row: int,
    mapped_columns: list[int],
    evidence_rows: int,
) -> int:
    if not mapped_columns or header_row >= worksheet.max_row:
        return 0
    end_row = min(worksheet.max_row, header_row + evidence_rows)
    evidence_count = 0
    for row_number in range(header_row + 1, end_row + 1):
        values = [worksheet.cell(row_number, index).value for index in mapped_columns]
        if any(value is not None and str(value).strip() for value in values):
            evidence_count += 1
    return evidence_count


def _candidate_for_row(
    worksheet: Worksheet,
    *,
    row_number: int,
    registry: SemanticFieldRegistry,
) -> HeaderCandidateResult:
    raw_values = next(
        worksheet.iter_rows(
            min_row=row_number,
            max_row=row_number,
            values_only=True,
        )
    )
    mappings = [
        registry.map_header(value, column_index=index)
        for index, value in enumerate(raw_values, start=1)
        if value is not None and str(value).strip()
    ]
    mapped = [item for item in mappings if item.semantic_field is not None]
    counts = Counter(item.semantic_field for item in mapped)
    duplicates = sorted(name for name, count in counts.items() if count > 1)
    if duplicates:
        for item in mapped:
            if item.semantic_field in duplicates:
                item.mapping_status = MappingStatus.AMBIGUOUS
                item.match_type = MatchType.AMBIGUOUS
                item.warning_code = "DUPLICATE_SEMANTIC_MAPPING"
    unique_count = len(counts)
    data_evidence = _has_data_evidence(
        worksheet,
        header_row=row_number,
        mapped_columns=[item.column_index for item in mapped],
        evidence_rows=registry.data_evidence_rows,
    )
    nonempty_count = len(mappings)
    score = (
        unique_count * 100
        + len(mapped) * 10
        + min(data_evidence, 1) * 5
        + min(nonempty_count, 20)
    )
    return HeaderCandidateResult(
        row_number=row_number,
        score=score,
        semantic_match_count=len(mapped),
        unique_semantic_count=unique_count,
        nonempty_cell_count=nonempty_count,
        data_evidence_count=data_evidence,
        mappings=mappings,
        duplicate_semantics=duplicates,
    )


def recognize_sheet(
    worksheet: Worksheet,
    *,
    sheet_index: int,
    registry: SemanticFieldRegistry,
) -> SheetRecognitionResult:
    scan_rows = min(registry.scan_rows, max(worksheet.max_row, 1))
    candidates = [
        _candidate_for_row(worksheet, row_number=row, registry=registry)
        for row in range(1, scan_rows + 1)
    ]
    ranked = sorted(candidates, key=lambda item: (-item.score, item.row_number))
    best = ranked[0]
    confident = (
        best.unique_semantic_count >= registry.minimum_unique_semantics
        and best.data_evidence_count > 0
    )
    tied = [
        item
        for item in ranked[1:]
        if confident
        and item.score == best.score
        and item.unique_semantic_count == best.unique_semantic_count
        and item.data_evidence_count > 0
    ]
    warnings: list[RecognitionWarning] = []
    merged_ranges = [str(item) for item in worksheet.merged_cells.ranges]
    if merged_ranges:
        warnings.append(
            RecognitionWarning(
                code="MERGED_CELL_DATA_CONTEXT_DETECTED",
                message_cn="检测到合并单元格；本阶段仅记录上下文，不执行向下填充。",
                context={"merged_ranges": merged_ranges},
            )
        )
    if not confident:
        status = RecognitionStatus.UNRECOGNIZED
        selected_header_row = None
        warnings.append(
            RecognitionWarning(
                code="INSUFFICIENT_HEADER_EVIDENCE",
                message_cn="没有足够的字段语义和数据证据确认业务表头。",
                context={
                    "best_candidate_row": best.row_number,
                    "unique_semantic_count": best.unique_semantic_count,
                    "data_evidence_count": best.data_evidence_count,
                },
            )
        )
    elif tied:
        status = RecognitionStatus.AMBIGUOUS
        selected_header_row = None
        warnings.append(
            RecognitionWarning(
                code="AMBIGUOUS_HEADER_CANDIDATES",
                message_cn="存在得分相同且无法可靠区分的表头候选。",
                context={
                    "candidate_rows": [best.row_number, *[item.row_number for item in tied]],
                    "score": best.score,
                },
            )
        )
    elif best.duplicate_semantics:
        status = RecognitionStatus.AMBIGUOUS
        selected_header_row = None
        warnings.append(
            RecognitionWarning(
                code="DUPLICATE_SEMANTIC_MAPPING",
                message_cn="同一语义字段匹配到多个原始列，不能自动覆盖或选择。",
                context={"semantic_fields": best.duplicate_semantics},
            )
        )
    else:
        selected_header_row = best.row_number
        unmapped = [
            item.raw_header
            for item in best.mappings
            if item.mapping_status == MappingStatus.UNMAPPED
        ]
        if unmapped:
            warnings.append(
                RecognitionWarning(
                    code="UNMAPPED_FIELDS_PRESENT",
                    message_cn="表头中存在尚未注册的字段，已完整保留。",
                    context={"raw_headers": unmapped},
                )
            )
        missing_units = [
            item.raw_header for item in best.mappings if item.warning_code
        ]
        if missing_units:
            warnings.append(
                RecognitionWarning(
                    code="UNIT_MISSING_OR_UNKNOWN",
                    message_cn="字段语义已识别，但单位缺失或不在受控单位规则中。",
                    context={"raw_headers": missing_units},
                )
            )
        status = (
            RecognitionStatus.RECOGNIZED_WITH_WARNING
            if warnings
            else RecognitionStatus.RECOGNIZED
        )

    formula_count = sum(
        1 for row in worksheet.iter_rows() for cell in row if cell.data_type == "f"
    )
    return SheetRecognitionResult(
        sheet_name=worksheet.title,
        sheet_index=sheet_index,
        sheet_state=worksheet.sheet_state,
        physical_row_count=worksheet.max_row,
        column_count=worksheet.max_column,
        dimension=worksheet.calculate_dimension(),
        formula_count=formula_count,
        merged_ranges=merged_ranges,
        status=status,
        selected_header_row=selected_header_row,
        recognition_score=best.score,
        recognized_field_count=best.unique_semantic_count,
        mappings=best.mappings,
        header_candidates=ranked,
        warnings=warnings,
    )


def recognize_workbook(
    workbook: Workbook,
    *,
    config: dict[str, Any],
    workbook_name: str,
    input_fingerprint: str,
) -> WorkbookRecognitionResult:
    registry = SemanticFieldRegistry(config)
    sheets = [
        recognize_sheet(worksheet, sheet_index=index, registry=registry)
        for index, worksheet in enumerate(workbook.worksheets)
    ]
    usable = [
        item
        for item in sheets
        if item.status
        in {RecognitionStatus.RECOGNIZED, RecognitionStatus.RECOGNIZED_WITH_WARNING}
    ]
    ambiguous = [item for item in sheets if item.status == RecognitionStatus.AMBIGUOUS]
    ranked = sorted(
        usable or ambiguous or sheets,
        key=lambda item: (-item.recognition_score, item.sheet_index),
    )
    best = ranked[0] if ranked else None
    if best is None:
        status = RecognitionStatus.UNRECOGNIZED
    elif best.status == RecognitionStatus.AMBIGUOUS:
        status = RecognitionStatus.AMBIGUOUS
    elif usable:
        status = best.status
    else:
        status = RecognitionStatus.UNRECOGNIZED
    warnings: list[RecognitionWarning] = []
    if not usable and not ambiguous:
        warnings.append(
            RecognitionWarning(
                code="WORKBOOK_UNRECOGNIZED",
                message_cn="工作簿扫描完成，但没有工作表具备足够的业务结构证据。",
                context={"sheet_count": len(sheets)},
            )
        )
    return WorkbookRecognitionResult(
        workbook_name=workbook_name,
        input_fingerprint=input_fingerprint,
        status=status,
        sheets=sheets,
        best_candidate_sheet=best.sheet_name if best else None,
        best_candidate_header_row=best.selected_header_row if best else None,
        warnings=warnings,
    )


def legacy_header_result(
    result: SheetRecognitionResult,
    *,
    registry: SemanticFieldRegistry,
) -> dict[str, Any]:
    """Keep the frozen Day 2 consumer shape while exposing WP6-1 separately."""
    detected = result.status in {
        RecognitionStatus.RECOGNIZED,
        RecognitionStatus.RECOGNIZED_WITH_WARNING,
    }
    mapping_preview: list[dict[str, Any]] = []
    for definition in registry.fields:
        if not definition.legacy_output or not definition.legacy_target_field:
            continue
        matched = [
            item
            for item in result.mappings
            if item.semantic_field == definition.semantic_field
        ]
        mapping_preview.append(
            {
                "target_field": definition.legacy_target_field,
                "required": definition.required,
                "status": (
                    "CONFLICT"
                    if len(matched) > 1
                    or any(item.mapping_status == MappingStatus.AMBIGUOUS for item in matched)
                    else "AUTO_MATCHED"
                    if len(matched) == 1
                    else "NOT_FOUND"
                ),
                "matched_columns": [
                    {
                        "column_index": item.column_index,
                        "column_letter": item.column_letter,
                        "source_header": item.raw_header,
                    }
                    for item in matched
                ],
            }
        )
    blocking_errors: list[dict[str, Any]] = []
    if result.status == RecognitionStatus.UNRECOGNIZED:
        blocking_errors.append(
            {
                "stage": "HEADER_DETECTION",
                "error_code": "HEADER_REQUIRED_FIELDS_NOT_FOUND",
                "message_cn": "未找到具备足够语义字段和数据证据的表头。",
                "source_location": f"工作表 {result.sheet_name}，扫描窗口内",
                "original_value": {
                    "best_match_count": result.recognized_field_count,
                    "recognition_status": result.status,
                },
                "rule": "表头候选必须同时具备多个受控语义字段和后续数据证据。",
                "impact": "该工作表不会进入旧版后续处理，但工作簿扫描继续。",
                "fix_suggestion": "检查表头位置和字段名称，或补充受控字段别名后重试。",
            }
        )
    if result.status == RecognitionStatus.AMBIGUOUS:
        blocking_errors.append(
            {
                "stage": "HEADER_DETECTION",
                "error_code": "HEADER_RECOGNITION_AMBIGUOUS",
                "message_cn": "存在表头候选歧义或重复语义列，无法安全自动选择。",
                "source_location": f"工作表 {result.sheet_name}",
                "original_value": {
                    "candidate_rows": [
                        item.row_number
                        for item in result.header_candidates
                        if item.score == result.recognition_score
                    ],
                    "duplicate_mappings": result.duplicate_mappings,
                },
                "rule": "同一语义只能对应一个源列，得分相同的候选不能静默选择。",
                "impact": "该工作表不会进入旧版后续处理，但候选信息完整保留。",
                "fix_suggestion": "确认唯一表头行和唯一字段列后重试。",
            }
        )
    return {
        "sheet_name": result.sheet_name,
        "scan_rows": registry.scan_rows,
        "detected": detected,
        "header_row": result.selected_header_row if detected else None,
        "best_candidate_row": result.header_candidates[0].row_number,
        "match_count": result.recognized_field_count,
        "minimum_match_count": registry.minimum_unique_semantics,
        "mapping_preview": mapping_preview,
        "blocking_errors": blocking_errors,
        "candidate_scores": [
            {
                "row_number": item.row_number,
                "match_count": item.unique_semantic_count,
                "nonempty_cell_count": item.nonempty_cell_count,
                "data_evidence_count": item.data_evidence_count,
                "score": item.score,
                "conflicts": item.duplicate_semantics,
            }
            for item in result.header_candidates
        ],
        "recognition_status": result.status,
        "recognition_warnings": [warning.to_dict() for warning in result.warnings],
    }


def recognition_markdown(result: WorkbookRecognitionResult) -> str:
    best = result.best_candidate
    lines = [
        "# WP6-1 数据结构识别报告",
        "",
        f"- 文件：`{result.workbook_name}`",
        f"- 输入 SHA-256：`{result.input_fingerprint}`",
        f"- 工作表数量：{result.sheet_count}",
        f"- 最终识别状态：`{result.status}`",
        f"- 最佳候选工作表：`{result.best_candidate_sheet or '无'}`",
        f"- 最佳候选 Header 行：{result.best_candidate_header_row or '未确认'}",
        "",
        "## 工作表识别",
        "",
        "|工作表|状态|Header 行|识别字段数|未映射字段数|Warning 数|",
        "|---|---|---:|---:|---:|---:|",
    ]
    for sheet in result.sheets:
        lines.append(
            f"|{sheet.sheet_name}|{sheet.status}|"
            f"{sheet.selected_header_row or ''}|{sheet.recognized_field_count}|"
            f"{len(sheet.unmapped_fields)}|{len(sheet.warnings)}|"
        )
    if best:
        lines.extend(["", "## 最佳候选字段映射", ""])
        lines.extend(
            [
                "|列|原始字段|识别语义|单位|状态|",
                "|---:|---|---|---|---|",
            ]
        )
        for mapping in best.mappings:
            lines.append(
                f"|{mapping.column_index}|{mapping.raw_header}|"
                f"{mapping.semantic_field or ''}|{mapping.detected_unit or ''}|"
                f"{mapping.mapping_status}|"
            )
        lines.extend(["", "## Warnings / Ambiguities", ""])
        if best.warnings:
            for warning in best.warnings:
                lines.append(f"- `{warning.code}`：{warning.message_cn}")
        else:
            lines.append("- 无")
    return "\n".join(lines) + "\n"
