"""Safe Day 2 Excel receipt, inspection and reporting."""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import zipfile
from datetime import date, datetime, time, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from carbon_excel_pipeline.errors import PipelineUserError
from carbon_excel_pipeline.io.header_detector import load_alias_config
from carbon_excel_pipeline.io.recognition import (
    legacy_header_result,
    recognition_markdown,
    recognize_workbook,
)
from carbon_excel_pipeline.io.semantic_registry import SemanticFieldRegistry


RUN_STAGE_DIRS = (
    "00_input_copy",
    "01_import",
    "02_capability",
    "02_scope_filter",
    "03_standardized",
    "04_qc",
    "05_activity",
    "06_third_party_input",
    "07_factor_results",
    "08_matching",
    "09_calculation",
    "10_output",
    "11_logs",
)
OLE_COMPOUND_MAGIC = bytes.fromhex("D0CF11E0A1B11AE1")
PROJECT_ROOT = Path(__file__).resolve().parents[3]
PUBLIC_CANDIDATE_ROOT = PROJECT_ROOT.parent / "02_open_source_release_candidate"
RUN_ID_PATTERN = re.compile(r"^RUN-[A-Za-z0-9_-]+$")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def _safe_cell_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def _user_error(
    *,
    code: str,
    message: str,
    location: str,
    value: Any,
    rule: str,
    suggestion: str,
) -> PipelineUserError:
    return PipelineUserError(
        stage="FILE_RECEIPT",
        error_code=code,
        message_cn=message,
        source_location=location,
        original_value=value,
        rule=rule,
        impact="阻断整次运行",
        fix_suggestion=suggestion,
    )


def validate_xlsx_source(source_path: Path, *, max_size_bytes: int) -> dict[str, Any]:
    source_path = source_path.expanduser().resolve()
    if not source_path.exists():
        raise _user_error(
            code="FILE_NOT_FOUND",
            message="未找到输入文件。",
            location=source_path.name,
            value=source_path.name,
            rule="输入路径必须指向存在的普通文件。",
            suggestion="重新选择存在的Excel文件。",
        )
    if not source_path.is_file():
        raise _user_error(
            code="INPUT_NOT_A_FILE",
            message="输入路径不是文件。",
            location=source_path.name,
            value=source_path.name,
            rule="输入必须是单个.xlsx文件。",
            suggestion="请选择具体Excel文件，不要选择文件夹。",
        )
    if source_path.suffix.lower() != ".xlsx":
        raise _user_error(
            code="UNSUPPORTED_FILE_EXTENSION",
            message="当前版本只支持.xlsx文件。",
            location=source_path.name,
            value=source_path.suffix,
            rule="第一版仅接收.xlsx，不接收.xls或.xlsm。",
            suggestion="请将文件另存为无宏的.xlsx后重试。",
        )
    size_bytes = source_path.stat().st_size
    if size_bytes == 0:
        raise _user_error(
            code="EMPTY_FILE",
            message="输入文件为空。",
            location=source_path.name,
            value=size_bytes,
            rule="输入文件大小必须大于0。",
            suggestion="请选择包含数据的有效.xlsx文件。",
        )
    if size_bytes > max_size_bytes:
        raise _user_error(
            code="FILE_TOO_LARGE",
            message="输入文件超过允许的大小。",
            location=source_path.name,
            value={"size_bytes": size_bytes, "limit_bytes": max_size_bytes},
            rule="第一版输入文件不得超过配置的大小上限。",
            suggestion="缩小文件或调整经过审批的文件大小配置。",
        )
    with source_path.open("rb") as handle:
        prefix = handle.read(8)
    if prefix == OLE_COMPOUND_MAGIC:
        raise _user_error(
            code="FILE_ENCRYPTED_OR_LEGACY_CONTAINER",
            message="文件可能已加密，或使用了旧式Office容器。",
            location=source_path.name,
            value="OLE_COMPOUND_CONTAINER",
            rule="第一版不支持加密、.xls或伪装成.xlsx的旧式文件。",
            suggestion="取消密码保护，并另存为标准.xlsx后重试。",
        )
    if not zipfile.is_zipfile(source_path):
        raise _user_error(
            code="FILE_CORRUPT_NOT_XLSX_PACKAGE",
            message="文件不是有效的.xlsx压缩包，可能已经损坏。",
            location=source_path.name,
            value="NOT_ZIP_PACKAGE",
            rule=".xlsx必须是可读取的Open XML压缩包。",
            suggestion="用Excel重新打开并另存文件，或选择未损坏的副本。",
        )
    try:
        with zipfile.ZipFile(source_path) as archive:
            bad_member = archive.testzip()
    except (OSError, zipfile.BadZipFile) as exc:
        raise _user_error(
            code="FILE_CORRUPT_ZIP_ERROR",
            message="Excel压缩包校验失败。",
            location=source_path.name,
            value=type(exc).__name__,
            rule=".xlsx内部压缩成员必须完整可读。",
            suggestion="重新保存Excel文件后再试。",
        ) from None
    if bad_member:
        raise _user_error(
            code="FILE_CORRUPT_MEMBER",
            message="Excel文件内部存在损坏的压缩成员。",
            location=source_path.name,
            value=bad_member,
            rule=".xlsx内部所有压缩成员必须通过CRC校验。",
            suggestion="重新保存Excel文件后再试。",
        )
    stat = source_path.stat()
    return {
        "path": source_path,
        "file_name": source_path.name,
        "size_bytes": size_bytes,
        "modified_time_utc": datetime.fromtimestamp(
            stat.st_mtime, tz=timezone.utc
        ).isoformat(),
        "modified_time_ns": stat.st_mtime_ns,
        "sha256": sha256_file(source_path),
    }


def create_isolated_run(run_root: Path, *, run_id: str | None = None) -> Path:
    run_root = run_root.expanduser().resolve()
    for forbidden_root in (PROJECT_ROOT, PUBLIC_CANDIDATE_ROOT):
        try:
            run_root.relative_to(forbidden_root.resolve())
        except ValueError:
            continue
        raise _user_error(
            code="RUN_ROOT_INSIDE_GIT_WORKSPACE",
            message="运行目录不能位于私有仓库或公开候选目录中。",
            location=run_root.name,
            value=run_root.name,
            rule="真实输入副本和运行输出必须位于仓库外的隔离临时目录。",
            suggestion="请选择WP6对应阶段证据目录下的专用runs目录。",
        )
    run_root.mkdir(parents=True, exist_ok=True)
    chosen_id = run_id or (
        f"RUN-{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%S%fZ')}-{uuid4().hex[:8].upper()}"
    )
    if not RUN_ID_PATTERN.fullmatch(chosen_id):
        raise _user_error(
            code="INVALID_RUN_ID",
            message="运行编号格式不合法。",
            location="Run_ID",
            value=chosen_id,
            rule="运行编号必须以RUN-开头且只能包含字母、数字、下划线和连字符。",
            suggestion="不指定运行编号，让程序自动生成安全编号。",
        )
    run_dir = run_root / chosen_id
    if run_dir.exists():
        raise _user_error(
            code="RUN_ID_ALREADY_EXISTS",
            message="运行编号已存在，不能覆盖旧运行。",
            location=chosen_id,
            value=chosen_id,
            rule="每次运行必须使用新的隔离目录。",
            suggestion="不指定运行编号，或改用尚未存在的新编号。",
        )
    for stage in RUN_STAGE_DIRS:
        (run_dir / stage).mkdir(parents=True, exist_ok=False)
    return run_dir


def _open_workbook(
    copy_path: Path, *, data_only: bool = False, read_only: bool = False
):
    try:
        return load_workbook(
            filename=copy_path,
            read_only=read_only,
            data_only=data_only,
            keep_links=False,
        )
    except (InvalidFileException, KeyError, OSError, ValueError, zipfile.BadZipFile) as exc:
        raise _user_error(
            code="WORKBOOK_OPEN_FAILED",
            message="Excel工作簿无法正常打开。",
            location=copy_path.name,
            value=type(exc).__name__,
            rule="工作簿必须是未加密且结构完整的.xlsx文件。",
            suggestion="请用Excel重新保存为标准.xlsx并确认未设置密码。",
        ) from None


def _sheet_preview(worksheet, *, max_rows: int) -> list[dict[str, Any]]:
    preview: list[dict[str, Any]] = []
    upper = min(max_rows, worksheet.max_row)
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=1, max_row=upper, values_only=True), start=1
    ):
        preview.append(
            {
                "source_row": row_number,
                "values": [_safe_cell_value(value) for value in values],
            }
        )
    return preview


def _has_record_value(value: Any) -> bool:
    return value is not None and (not isinstance(value, str) or bool(value.strip()))


def _materialize_recognized_records(
    *,
    formula_worksheet,
    value_worksheet,
    sheet_result,
    workbook_name: str,
    input_fingerprint: str,
) -> dict[str, Any]:
    """Freeze recognized values so WP6-2 never needs to reopen or rescan Excel."""
    header_row = sheet_result.selected_header_row
    mappings = [
        item
        for item in sheet_result.mappings
        if item.semantic_field is not None and item.mapping_status == "MAPPED"
    ]
    units = {item.semantic_field: item.detected_unit for item in mappings}
    records: list[dict[str, Any]] = []
    if header_row is not None:
        formula_rows = formula_worksheet.iter_rows(
            min_row=header_row + 1,
            max_row=formula_worksheet.max_row,
            values_only=True,
        )
        value_rows = value_worksheet.iter_rows(
            min_row=header_row + 1,
            max_row=value_worksheet.max_row,
            values_only=True,
        )
        for source_row, (formula_row, value_row) in enumerate(
            zip(formula_rows, value_rows), start=header_row + 1
        ):
            values: dict[str, Any] = {}
            formula_fields: list[str] = []
            has_semantic_value = False
            for mapping in mappings:
                index = mapping.column_index - 1
                raw_value = formula_row[index] if index < len(formula_row) else None
                cached_value = value_row[index] if index < len(value_row) else None
                chosen_value = cached_value if isinstance(raw_value, str) and raw_value.startswith("=") else raw_value
                if isinstance(raw_value, str) and raw_value.startswith("="):
                    formula_fields.append(mapping.semantic_field)
                safe_value = _safe_cell_value(chosen_value)
                values[mapping.semantic_field] = safe_value
                has_semantic_value = has_semantic_value or _has_record_value(raw_value) or _has_record_value(cached_value)
            if has_semantic_value:
                records.append(
                    {
                        "Source_Row": source_row,
                        "values": values,
                        "formula_fields": sorted(formula_fields),
                    }
                )
    return {
        "schema_version": "WP6_1_RECOGNIZED_RECORDS_V1",
        "workbook_name": workbook_name,
        "input_fingerprint": input_fingerprint,
        "sheet_name": sheet_result.sheet_name,
        "sheet_index": sheet_result.sheet_index,
        "header_row": header_row,
        "denominator_definition": (
            "Rows after the recognized header with at least one non-empty mapped semantic value."
        ),
        "units": units,
        "column_mappings": [item.to_dict() for item in mappings],
        "record_count": len(records),
        "records": records,
    }


def inspect_excel_to_run(
    source_path: Path,
    *,
    run_root: Path,
    alias_config_path: Path,
    max_size_bytes: int = 50 * 1024 * 1024,
    preview_rows: int = 20,
    run_id: str | None = None,
) -> dict[str, Any]:
    source = validate_xlsx_source(source_path, max_size_bytes=max_size_bytes)
    run_dir = create_isolated_run(run_root, run_id=run_id)
    input_copy = run_dir / "00_input_copy" / source["file_name"]
    shutil.copy2(source["path"], input_copy)
    copy_hash = sha256_file(input_copy)
    if copy_hash != source["sha256"]:
        raise _user_error(
            code="INPUT_COPY_HASH_MISMATCH",
            message="隔离副本与原始文件指纹不一致。",
            location=input_copy.name,
            value={"source_sha256": source["sha256"], "copy_sha256": copy_hash},
            rule="隔离副本必须与原始输入逐字节一致。",
            suggestion="停止使用该副本并重新运行文件接收。",
        )

    alias_config = load_alias_config(alias_config_path)
    registry = SemanticFieldRegistry(alias_config)
    workbook = _open_workbook(input_copy, data_only=False)
    value_workbook = _open_workbook(input_copy, data_only=True, read_only=True)
    sheet_inventory: list[dict[str, Any]] = []
    header_results: list[dict[str, Any]] = []
    previews: list[dict[str, Any]] = []
    recognized_records: dict[str, Any] | None = None
    extra_records: list[dict[str, Any]] = []
    try:
        recognition = recognize_workbook(
            workbook,
            config=alias_config,
            workbook_name=source["file_name"],
            input_fingerprint=source["sha256"],
        )
        recognition_by_sheet = {item.sheet_name: item for item in recognition.sheets}
        for worksheet in workbook.worksheets:
            sheet_result = recognition_by_sheet[worksheet.title]
            header = legacy_header_result(sheet_result, registry=registry)
            inventory = {
                "sheet_name": worksheet.title,
                "sheet_index": sheet_result.sheet_index,
                "sheet_state": worksheet.sheet_state,
                "physical_row_count": worksheet.max_row,
                "column_count": worksheet.max_column,
                "dimension": worksheet.calculate_dimension(),
                "header_detected": header["detected"],
                "header_row": header["header_row"],
                "recognition_status": sheet_result.status,
                "recognition_score": sheet_result.recognition_score,
                "recognized_field_count": sheet_result.recognized_field_count,
                "data_row_count": (
                    worksheet.max_row - header["header_row"]
                    if header["detected"] and header["header_row"] is not None
                    else None
                ),
                "formula_count": sheet_result.formula_count,
                "merged_cell_count": sheet_result.merged_cell_count,
                "merged_ranges": sheet_result.merged_ranges,
            }
            sheet_inventory.append(inventory)
            header_results.append(header)
            previews.append(
                {
                    "sheet_name": worksheet.title,
                    "preview_row_limit": preview_rows,
                    "rows": _sheet_preview(worksheet, max_rows=preview_rows),
                }
            )
        best = recognition.best_candidate
        extra_records: list[dict[str, Any]] = []
        if best is not None and best.selected_header_row is not None:
            recognized_records = _materialize_recognized_records(
                formula_worksheet=workbook[best.sheet_name],
                value_worksheet=value_workbook[best.sheet_name],
                sheet_result=best,
                workbook_name=recognition.workbook_name,
                input_fingerprint=recognition.input_fingerprint,
            )
            activity_fields = {
                "Reported_Activity_Value",
                "Quantity_PCS",
                "Unit_Weight",
                "Reported_Purchase_Quantity",
            }
            for sheet_result in recognition.sheets:
                if sheet_result.sheet_name == best.sheet_name:
                    continue
                if sheet_result.status not in {"RECOGNIZED", "RECOGNIZED_WITH_WARNING"}:
                    continue
                if sheet_result.selected_header_row is None:
                    continue
                mapped = {item.semantic_field for item in sheet_result.mappings if item.semantic_field}
                if not (mapped & activity_fields):
                    continue
                extra_records.append(
                    _materialize_recognized_records(
                        formula_worksheet=workbook[sheet_result.sheet_name],
                        value_worksheet=value_workbook[sheet_result.sheet_name],
                        sheet_result=sheet_result,
                        workbook_name=recognition.workbook_name,
                        input_fingerprint=recognition.input_fingerprint,
                    )
                )
        else:
            extra_records = []
    finally:
        workbook.close()
        value_workbook.close()

    eligible_sheets = [item for item in header_results if item["detected"]]
    blocking_errors = [
        error
        for item in header_results
        for error in item.get("blocking_errors", [])
        if not item["detected"]
    ]
    source_hash_after = sha256_file(source["path"])
    source_stat_after = source["path"].stat()
    source_unchanged = (
        source_hash_after == source["sha256"]
        and source_stat_after.st_size == source["size_bytes"]
        and source_stat_after.st_mtime_ns == source["modified_time_ns"]
    )
    if not source_unchanged:
        raise _user_error(
            code="SOURCE_FILE_CHANGED_DURING_INSPECTION",
            message="检查期间原始文件发生变化。",
            location=source["file_name"],
            value={"before_sha256": source["sha256"], "after_sha256": source_hash_after},
            rule="原始文件在一次运行期间必须保持不变。",
            suggestion="关闭正在编辑该文件的程序，确认版本后重新运行。",
        )

    status = "PASS" if eligible_sheets else "BLOCKED"
    receipt_report = {
        "run_id": run_dir.name,
        "status": "ACCEPTED",
        "source_file_name": source["file_name"],
        "extension": source["path"].suffix.lower(),
        "size_bytes": source["size_bytes"],
        "max_size_bytes": max_size_bytes,
        "modified_time_utc": source["modified_time_utc"],
        "source_sha256": source["sha256"],
        "copy_file_name": input_copy.name,
        "copy_sha256": copy_hash,
        "source_unchanged": source_unchanged,
        "original_file_overwritten": False,
    }
    summary = {
        "run_id": run_dir.name,
        "status": status,
        "recognition_status": recognition.status,
        "source_file_name": source["file_name"],
        "sheet_count": len(sheet_inventory),
        "best_candidate_sheet": recognition.best_candidate_sheet,
        "best_candidate_header_row": recognition.best_candidate_header_row,
        "eligible_sheet_count": len(eligible_sheets),
        "eligible_sheets": [item["sheet_name"] for item in eligible_sheets],
        "blocking_error_count": len(blocking_errors),
        "reports": {
            "file_receipt": "01_import/file_receipt_report.json",
            "sheet_inventory": "01_import/sheet_inventory.json",
            "header_detection": "01_import/header_detection.json",
            "field_mapping_preview": "01_import/field_mapping_preview.json",
            "input_preview": "01_import/input_preview.json",
            "workbook_inventory": "01_import/workbook_inventory.json",
            "sheet_recognition": "01_import/sheet_recognition.json",
            "header_candidates": "01_import/header_candidates.json",
            "semantic_field_mapping": "01_import/semantic_field_mapping.json",
            "recognition_summary": "01_import/recognition_summary.json",
            "recognized_records": "01_import/recognized_records.json",
            "recognized_records_by_sheet": "01_import/recognized_records_by_sheet.json",
            "recognition_report": "01_import/WP6-1_数据结构识别报告.md",
        },
    }
    import_dir = run_dir / "01_import"
    _write_json(import_dir / "file_receipt_report.json", receipt_report)
    _write_json(import_dir / "sheet_inventory.json", sheet_inventory)
    _write_json(import_dir / "header_detection.json", header_results)
    _write_json(
        import_dir / "field_mapping_preview.json",
        [
            {
                "sheet_name": item["sheet_name"],
                "detected": item["detected"],
                "header_row": item["header_row"],
                "match_count": item["match_count"],
                "mapping_preview": item["mapping_preview"],
                "blocking_errors": item["blocking_errors"],
            }
            for item in header_results
        ],
    )
    _write_json(import_dir / "input_preview.json", previews)
    recognition_payload = recognition.to_dict()
    sheet_payload = [item.to_dict() for item in recognition.sheets]
    _write_json(
        import_dir / "workbook_inventory.json",
        {
            "workbook_name": recognition.workbook_name,
            "input_fingerprint": recognition.input_fingerprint,
            "sheet_count": recognition.sheet_count,
            "sheets": sheet_inventory,
        },
    )
    _write_json(import_dir / "sheet_recognition.json", sheet_payload)
    _write_json(
        import_dir / "header_candidates.json",
        [
            {
                "sheet_name": item.sheet_name,
                "candidates": [candidate.to_dict() for candidate in item.header_candidates],
            }
            for item in recognition.sheets
        ],
    )
    _write_json(
        import_dir / "semantic_field_mapping.json",
        [
            {
                "sheet_name": item.sheet_name,
                "header_row": item.selected_header_row,
                "recognition_status": item.status,
                "field_mappings": [mapping.to_dict() for mapping in item.mappings],
            }
            for item in recognition.sheets
        ],
    )
    _write_json(import_dir / "recognition_summary.json", recognition_payload)
    _write_json(
        import_dir / "recognized_records.json",
        recognized_records
        or {
            "schema_version": "WP6_1_RECOGNIZED_RECORDS_V1",
            "workbook_name": recognition.workbook_name,
            "input_fingerprint": recognition.input_fingerprint,
            "sheet_name": None,
            "sheet_index": None,
            "header_row": None,
            "denominator_definition": (
                "Rows after the recognized header with at least one non-empty mapped semantic value."
            ),
            "units": {},
            "column_mappings": [],
            "record_count": 0,
            "records": [],
        },
    )
    _write_json(
        import_dir / "recognized_records_by_sheet.json",
        {
            "schema_version": "WP6_1_RECOGNIZED_RECORDS_BY_SHEET_V1",
            "workbook_name": recognition.workbook_name,
            "input_fingerprint": recognition.input_fingerprint,
            "sheets": extra_records,
        },
    )
    (import_dir / "WP6-1_数据结构识别报告.md").write_text(
        recognition_markdown(recognition), encoding="utf-8"
    )
    _write_json(
        run_dir / "11_logs" / "wp6_1_recognition_log.json",
        {
            "run_id": run_dir.name,
            "input_fingerprint": source["sha256"],
            "workbook_sheets": [item.sheet_name for item in recognition.sheets],
            "best_candidate_sheet": recognition.best_candidate_sheet,
            "selected_header": recognition.best_candidate_header_row,
            "recognition_score": (
                recognition.best_candidate.recognition_score
                if recognition.best_candidate
                else 0
            ),
            "mapped_fields": (
                recognition.best_candidate.recognized_fields
                if recognition.best_candidate
                else []
            ),
            "unmapped_fields": (
                recognition.best_candidate.unmapped_fields
                if recognition.best_candidate
                else []
            ),
            "warnings": (
                [item.to_dict() for item in recognition.best_candidate.warnings]
                if recognition.best_candidate
                else []
            ),
            "final_recognition_status": recognition.status,
        },
    )
    _write_json(import_dir / "day2_run_summary.json", summary)
    return {**summary, "run_directory": str(run_dir)}
