"""Controlled semantic header aliases and unit parsing for WP6-1."""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from typing import Any

from openpyxl.utils import get_column_letter

from carbon_excel_pipeline.io.recognition_models import (
    FieldMappingResult,
    MappingStatus,
    MatchType,
)


_UNNAMED_PATTERN = re.compile(r"^unnamed\s*:\s*\d+(?:\.\d+)?$", re.IGNORECASE)
_BRACKETED_UNIT_PATTERN = re.compile(
    r"\s*[（(]\s*(?:"
    r"t\s*/\s*(?:年|year)|"
    r"kg\s*/\s*(?:年|year)|"
    r"kg\s*co2e?\s*/\s*kg|"
    r"t\s*co2e?\s*/\s*t|"
    r"t\s*co2e?\s*/\s*(?:年|year)|"
    r"kg\s*co2e?\s*/\s*(?:年|year)|"
    r"(?:kg|t)\s*/\s*pcs|"
    r"g\s*/\s*pcs|"
    r"g\s*/\s*(?:年|year)"
    r")\s*[)）]\s*",
    re.IGNORECASE,
)


def normalize_header(value: Any) -> str:
    """Normalize layout noise without discarding the original header value."""
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("\u3000", " ").replace("\t", " ")
    text = re.sub(r"\s+", " ", text).strip().casefold()
    if _UNNAMED_PATTERN.fullmatch(text):
        return ""
    return text


def semantic_header_key(value: Any) -> str:
    normalized = normalize_header(value)
    return re.sub(r"\s+", " ", _BRACKETED_UNIT_PATTERN.sub(" ", normalized)).strip()


def parse_header_unit(value: Any) -> str | None:
    compact = re.sub(r"\s+", "", normalize_header(value))
    if re.search(r"kgco2e?/kg", compact, re.IGNORECASE):
        return "kgCO2e/kg"
    if re.search(r"tco2e?/t", compact, re.IGNORECASE):
        return "tCO2e/t"
    if re.search(r"tco2e?/(?:年|year)", compact, re.IGNORECASE):
        return "tCO2e/year"
    if re.search(r"kgco2e?/(?:年|year)", compact, re.IGNORECASE):
        return "kgCO2e/year"
    if re.search(r"(?:^|[（(])t/(?:年|year)(?:[)）]|$)", compact, re.IGNORECASE):
        return "t/year"
    if re.search(r"(?:^|[（(])kg/(?:年|year)(?:[)）]|$)", compact, re.IGNORECASE):
        return "kg/year"
    if re.search(r"kg/pcs", compact, re.IGNORECASE):
        return "kg/PCS"
    if re.search(r"t/pcs", compact, re.IGNORECASE):
        return "t/PCS"
    if re.search(r"g/pcs", compact, re.IGNORECASE):
        return "g/PCS"
    if re.search(r"g/(?:年|year)", compact, re.IGNORECASE):
        return "g/year"
    return None


@dataclass(frozen=True)
class SemanticFieldDefinition:
    semantic_field: str
    legacy_target_field: str | None
    canonical_source_header: str
    aliases: tuple[str, ...]
    unit_required: bool
    legacy_output: bool
    required: bool


class SemanticFieldRegistry:
    """Validated exact/alias lookup shared by scanner, UI, CLI and tests."""

    def __init__(self, config: dict[str, Any]):
        self.scan_rows = int(config.get("scan_rows", 10))
        self.data_evidence_rows = int(config.get("data_evidence_rows", 3))
        self.minimum_unique_semantics = int(
            config.get("minimum_unique_semantics", 2)
        )
        self.fields = tuple(self._load_definition(item) for item in config["fields"])
        self._full_alias_index: dict[str, tuple[SemanticFieldDefinition, MatchType]] = {}
        self._semantic_alias_index: dict[
            str, tuple[SemanticFieldDefinition, MatchType]
        ] = {}
        self._compact_alias_index: dict[
            str, tuple[SemanticFieldDefinition, MatchType]
        ] = {}
        for definition in self.fields:
            self._register_alias(
                definition.canonical_source_header, definition, MatchType.EXACT
            )
            for alias in definition.aliases:
                self._register_alias(alias, definition, MatchType.ALIAS)

    @staticmethod
    def _load_definition(item: dict[str, Any]) -> SemanticFieldDefinition:
        semantic_field = str(item.get("semantic_field") or item["target_field"])
        legacy_target = item.get("legacy_target_field")
        if legacy_target is None and item.get("legacy_output", True):
            legacy_target = item.get("target_field", semantic_field)
        return SemanticFieldDefinition(
            semantic_field=semantic_field,
            legacy_target_field=str(legacy_target) if legacy_target else None,
            canonical_source_header=str(item["canonical_source_header"]),
            aliases=tuple(str(value) for value in item.get("aliases", [])),
            unit_required=bool(item.get("unit_required", False)),
            legacy_output=bool(item.get("legacy_output", True)),
            required=bool(item.get("required", False)),
        )

    def _register_alias(
        self,
        raw_alias: str,
        definition: SemanticFieldDefinition,
        match_type: MatchType,
    ) -> None:
        for index, key in (
            (self._full_alias_index, normalize_header(raw_alias)),
            (self._semantic_alias_index, semantic_header_key(raw_alias)),
            (
                self._compact_alias_index,
                re.sub(r"\s+", "", semantic_header_key(raw_alias)),
            ),
        ):
            if not key:
                continue
            existing = index.get(key)
            if existing and existing[0].semantic_field != definition.semantic_field:
                raise ValueError(
                    f"Alias {raw_alias!r} maps to both "
                    f"{existing[0].semantic_field!r} and {definition.semantic_field!r}."
                )
            if existing is None or match_type == MatchType.EXACT:
                index[key] = (definition, match_type)

    def definition_for_semantic(
        self, semantic_field: str
    ) -> SemanticFieldDefinition | None:
        return next(
            (item for item in self.fields if item.semantic_field == semantic_field),
            None,
        )

    def map_header(self, value: Any, *, column_index: int) -> FieldMappingResult:
        raw_header = "" if value is None else str(value)
        normalized = normalize_header(value)
        match = self._full_alias_index.get(normalized)
        if match is None:
            match = self._semantic_alias_index.get(semantic_header_key(value))
        if match is None:
            match = self._compact_alias_index.get(
                re.sub(r"\s+", "", semantic_header_key(value))
            )
        detected_unit = parse_header_unit(value)
        if match is None:
            return FieldMappingResult(
                column_index=column_index,
                column_letter=get_column_letter(column_index),
                raw_header=raw_header,
                normalized_header=normalized,
                semantic_field=None,
                legacy_target_field=None,
                match_type=MatchType.UNMAPPED,
                detected_unit=detected_unit,
                mapping_status=MappingStatus.UNMAPPED,
            )
        definition, match_type = match
        warning_code = (
            "UNIT_MISSING_OR_UNKNOWN"
            if definition.unit_required and detected_unit is None
            else None
        )
        return FieldMappingResult(
            column_index=column_index,
            column_letter=get_column_letter(column_index),
            raw_header=raw_header,
            normalized_header=normalized,
            semantic_field=definition.semantic_field,
            legacy_target_field=definition.legacy_target_field,
            match_type=match_type,
            detected_unit=detected_unit,
            mapping_status=MappingStatus.MAPPED,
            warning_code=warning_code,
        )
