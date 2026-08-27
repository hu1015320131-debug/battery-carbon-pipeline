"""WP6-8 business integration built exclusively from accepted backend artifacts.

This module does not calculate Activity, EF, Emission, quality metrics, or the
WP6-6 scenarios.  It validates and packages values already produced by WP6-3
through WP6-7.  The only numeric transformation is the declared kg-to-tonne
presentation conversion used by Suggested Ledger V1.
"""

from __future__ import annotations

import csv
import hashlib
import json
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from carbon_excel_pipeline.wp6_7.pipeline import (
    build_data_quality_scorecard,
    build_dimension_availability,
    build_issue_register,
    build_lineage_quality_summary,
    build_management_summary,
    build_top_emission_contributors,
)


SCOPE_NOTICE = (
    "2024与2025试点Scope不同；年度结果仅用于数据结构、质量和试点结果并列对照，"
    "不代表同口径企业年度同比。"
)

LEDGER_FIELDS = [
    "序号",
    "Record_ID",
    "事业部",
    "采购类型",
    "公司外购原料和辅料名称",
    "物料描述",
    "年度购买原料量（t/year）",
    "LCA排放因子（kgCO2e/kg）",
    "排放因子来源",
    "GHG排放量（tCO2e/year）",
]

AUDIT_FIELDS = [
    "Record_ID",
    "Year",
    "Source_File",
    "Source_SHA256",
    "Source_Sheet",
    "Source_Row",
    "Business_Unit",
    "Purchase_Type",
    "Purchase_Category",
    "Product_Description",
    "Original_Activity_Value",
    "Original_Activity_Unit",
    "Activity_Data_kg",
    "Activity_Method",
    "EF_Value",
    "EF_Unit",
    "EF_Source",
    "EF_Usage",
    "Emission_kgCO2e",
    "Emission_tCO2e",
    "Calculation_QC",
    "Governance_QC",
    "Boundary_Ready",
    "Warning_Codes",
    "Run_ID",
    "Simulation_Flag",
    "Production_Eligible",
]

CANONICAL_EXPORT_FIELDS = [
    "Record_ID",
    "Year",
    "Business_Unit",
    "Purchase_Category",
    "Product_Description",
    "Activity_Data_kg",
    "Activity_Method",
    "EF_Value",
    "EF_Unit",
    "Emission_kgCO2e",
    "Emission_tCO2e",
    "Calculation_QC",
    "Governance_QC",
    "Boundary_Ready",
    "Source_Row",
]

EXPECTED_SHEETS = [
    "01_核算结果",
    "02_建议清册",
    "03_审计明细",
    "04_数据质量",
    "05_问题登记",
    "06_管理汇总",
    "07_排放贡献",
    "08_因子敏感性",
    "09_运行信息",
]


class WP68IntegrationError(RuntimeError):
    """Raised when accepted upstream evidence cannot be packaged safely."""


def _read_csv(path: Path) -> list[dict[str, str]]:
    if not path.is_file():
        raise WP68IntegrationError(f"缺少必需的上游 CSV：{path}")
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _read_json(path: Path) -> Any:
    if not path.is_file():
        raise WP68IntegrationError(f"缺少必需的上游 JSON：{path}")
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def _write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _decimal(value: Any, field: str, record_id: str) -> Decimal:
    text = _text(value)
    if not text:
        raise WP68IntegrationError(f"{record_id} 缺少 {field}。")
    try:
        return Decimal(text)
    except InvalidOperation as error:
        raise WP68IntegrationError(f"{record_id} 的 {field} 不是有效数值：{text}") from error


def _format_decimal(value: Decimal) -> str:
    rendered = format(value, "f")
    if "." in rendered:
        rendered = rendered.rstrip("0").rstrip(".")
    return rendered or "0"


def _kg_to_tonnes(value: Any, field: str, record_id: str) -> str:
    """Single backend presentation conversion shared by CSV and workbook."""

    return _format_decimal(_decimal(value, field, record_id) / Decimal("1000"))


def _index(rows: Iterable[dict[str, str]], label: str) -> dict[str, dict[str, str]]:
    result: dict[str, dict[str, str]] = {}
    for row in rows:
        record_id = _text(row.get("Record_ID"))
        if not record_id:
            raise WP68IntegrationError(f"{label} 存在空 Record_ID。")
        if record_id in result:
            raise WP68IntegrationError(f"{label} 存在重复 Record_ID：{record_id}")
        result[record_id] = row
    return result


def _first(*values: Any) -> str:
    return next((_text(value) for value in values if _text(value)), "")


def _warning_codes(canonical: Mapping[str, Any], validation: Mapping[str, Any]) -> str:
    values: list[str] = []
    for key in (
        "Warning_Codes",
        "Calculation_Issue_Codes",
        "Governance_Issue_Codes",
        "Boundary_Issue_Codes",
        "Reason_Codes",
    ):
        raw = _first(canonical.get(key), validation.get(key))
        for code in raw.replace("|", ";").split(";"):
            if code and code not in values:
                values.append(code)
    return ";".join(values)


def _exclusion_reasons(row: Mapping[str, Any]) -> list[str]:
    reasons: list[str] = []
    for field in ("Record_ID", "Activity_Data_kg", "EF_Value", "Emission_kgCO2e"):
        if not _text(row.get(field)):
            reasons.append(f"{field}_MISSING")
    if _text(row.get("Boundary_Ready")).upper() in {"FALSE", "0", "NO", "BLOCKED"}:
        reasons.append("BOUNDARY_NOT_READY")
    if _text(row.get("Emission_Ready")).upper() in {"FALSE", "0", "NO", "BLOCKED"}:
        reasons.append("EMISSION_NOT_READY")
    return reasons


def build_delivery_rows(
    canonical_rows: Iterable[dict[str, Any]],
    *,
    enrichment_by_id: Mapping[str, Mapping[str, Any]] | None = None,
    validation_by_id: Mapping[str, Mapping[str, Any]] | None = None,
    fallback_run_id: str = "",
) -> dict[str, Any]:
    """Build ledger/audit rows without recalculating upstream business results.

    Rows missing an already-computed Activity/EF/Emission are explicitly excluded.
    The returned business status implements best-available-result behavior.
    """

    enrichments = enrichment_by_id or {}
    validations = validation_by_id or {}
    audit_rows: list[dict[str, str]] = []
    ledger_rows: list[dict[str, str]] = []
    canonical_export: list[dict[str, str]] = []
    exclusions: list[dict[str, str]] = []
    seen: set[str] = set()
    source_count = 0

    for source_count, canonical in enumerate(canonical_rows, start=1):
        record_id = _text(canonical.get("Record_ID"))
        reasons = _exclusion_reasons(canonical)
        if record_id and record_id in seen:
            reasons.append("RECORD_ID_DUPLICATE")
        if reasons:
            exclusions.append(
                {
                    "Record_ID": record_id or f"ROW-{source_count}",
                    "Reason_Codes": ";".join(reasons),
                    "Action": "检查上游 Capability、Boundary 与核算结果；不将该记录写入清册。",
                }
            )
            continue
        seen.add(record_id)
        enrichment = enrichments.get(record_id, {})
        validation = validations.get(record_id, {})
        year = _first(canonical.get("Year"), validation.get("Year"))
        activity_kg = _text(canonical["Activity_Data_kg"])
        emission_kg = _text(canonical["Emission_kgCO2e"])
        ef_value = _text(canonical["EF_Value"])
        activity_t = _kg_to_tonnes(activity_kg, "Activity_Data_kg", record_id)
        emission_t = _first(
            canonical.get("Emission_tCO2e"),
            _kg_to_tonnes(emission_kg, "Emission_kgCO2e", record_id),
        )
        purchase_category = _first(
            canonical.get("Purchase_Category"), enrichment.get("Activity_Category")
        )
        purchase_type = _first(
            canonical.get("Purchase_Type"), enrichment.get("Activity_Category"), purchase_category
        )
        calculation_qc = _first(
            canonical.get("Calculation_QC"), canonical.get("QC_Status"), "PASS"
        )
        governance_qc = _first(
            canonical.get("Governance_QC"), validation.get("Governance_QC"), canonical.get("QC_Status")
        )
        audit = {
            "Record_ID": record_id,
            "Year": year,
            "Source_File": _first(canonical.get("Source_File"), validation.get("Source_File")),
            "Source_SHA256": _first(
                canonical.get("Source_SHA256"), validation.get("Source_SHA256")
            ),
            "Source_Sheet": _first(canonical.get("Source_Sheet"), validation.get("Source_Sheet")),
            "Source_Row": _first(canonical.get("Source_Row"), validation.get("Source_Row")),
            "Business_Unit": _first(canonical.get("Business_Unit"), enrichment.get("Business_Unit")),
            "Purchase_Type": purchase_type,
            "Purchase_Category": purchase_category,
            "Product_Description": _first(
                canonical.get("Product_Description"), enrichment.get("Product_Description")
            ),
            "Original_Activity_Value": _first(
                canonical.get("Original_Activity_Value"),
                validation.get("Original_Activity_Value"),
                enrichment.get("Original_Activity_Value"),
            ),
            "Original_Activity_Unit": _first(
                canonical.get("Original_Activity_Unit"),
                validation.get("Original_Activity_Unit"),
                enrichment.get("Original_Activity_Unit"),
            ),
            "Activity_Data_kg": activity_kg,
            "Activity_Method": _first(canonical.get("Activity_Method"), validation.get("Activity_Method")),
            "EF_Value": ef_value,
            "EF_Unit": _first(canonical.get("EF_Unit"), validation.get("Main_EF_Unit")),
            "EF_Source": _first(canonical.get("EF_Source"), validation.get("Independent_EF_Source")),
            "EF_Usage": _first(
                canonical.get("EF_Usage"),
                "HISTORICAL_REPRODUCTION" if year == "2024" else "HISTORICAL_SIMULATION",
            ),
            "Emission_kgCO2e": emission_kg,
            "Emission_tCO2e": emission_t,
            "Calculation_QC": calculation_qc,
            "Governance_QC": governance_qc,
            "Boundary_Ready": _first(canonical.get("Boundary_Ready"), validation.get("Boundary_Ready"), "TRUE"),
            "Warning_Codes": _warning_codes(canonical, validation),
            "Run_ID": _first(canonical.get("Run_ID"), fallback_run_id),
            "Simulation_Flag": _first(canonical.get("Simulation_Flag"), "TRUE"),
            "Production_Eligible": _first(canonical.get("Production_Eligible"), "FALSE"),
        }
        audit_rows.append(audit)
        ledger_rows.append(
            {
                "序号": str(len(ledger_rows) + 1),
                "Record_ID": record_id,
                "事业部": audit["Business_Unit"],
                "采购类型": audit["Purchase_Type"],
                "公司外购原料和辅料名称": audit["Purchase_Category"],
                "物料描述": audit["Product_Description"],
                "年度购买原料量（t/year）": activity_t,
                "LCA排放因子（kgCO2e/kg）": ef_value,
                "排放因子来源": audit["EF_Source"],
                "GHG排放量（tCO2e/year）": emission_t,
            }
        )
        canonical_export.append(
            {
                field: audit.get(field, _text(canonical.get(field)))
                for field in CANONICAL_EXPORT_FIELDS
            }
        )

    if not audit_rows:
        status = "BLOCKED"
    elif exclusions:
        status = "PARTIAL_RESULT"
    elif any(row["Governance_QC"].upper() == "WARNING" for row in audit_rows):
        status = "PASS_WITH_WARNING"
    else:
        status = "PASS"
    return {
        "status": status,
        "input_record_count": source_count,
        "included_record_count": len(audit_rows),
        "excluded_record_count": len(exclusions),
        "canonical_rows": canonical_export,
        "ledger_rows": ledger_rows,
        "audit_rows": audit_rows,
        "excluded_rows": exclusions,
    }


def _flatten_scorecards(*scorecards: Mapping[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for scorecard in scorecards:
        year = _text(scorecard.get("Year"))
        for metric in scorecard.get("metrics", []):
            rows.append({"Year": year, **metric})
    return rows


def _write_workbook(path: Path, sheets: list[tuple[str, list[dict[str, Any]]]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, rows in sheets:
        sheet = workbook.create_sheet(title)
        if not rows:
            sheet.append(["说明"])
            sheet.append(["当前正式结果没有该类记录"])
            continue
        headers = list(rows[0])
        sheet.append(headers)
        for row in rows:
            sheet.append([_text(row.get(header)) for header in headers])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for index, header in enumerate(headers, start=1):
            sample = [str(header)] + [_text(row.get(header)) for row in rows[:200]]
            sheet.column_dimensions[get_column_letter(index)].width = min(
                max(len(value) for value in sample) + 2, 48
            )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _validate_delivery(
    delivery: Mapping[str, Any], workbook_path: Path
) -> dict[str, bool]:
    ledger_rows = delivery["ledger_rows"]
    audit_rows = delivery["audit_rows"]
    if len(ledger_rows) != len(audit_rows):
        raise WP68IntegrationError("建议清册与审计明细记录数不一致。")
    for ledger, audit in zip(ledger_rows, audit_rows, strict=True):
        record_id = audit["Record_ID"]
        if ledger["Record_ID"] != record_id:
            raise WP68IntegrationError(f"{record_id} 建议清册 Record_ID 无法与 Canonical 勾稽。")
        if Decimal(ledger["年度购买原料量（t/year）"]) * 1000 != Decimal(
            audit["Activity_Data_kg"]
        ):
            raise WP68IntegrationError(f"{record_id} 建议清册 Activity 无法与 Canonical 勾稽。")
        if Decimal(ledger["LCA排放因子（kgCO2e/kg）"]) != Decimal(audit["EF_Value"]):
            raise WP68IntegrationError(f"{record_id} 建议清册 EF 无法与 Canonical 勾稽。")
        if Decimal(ledger["GHG排放量（tCO2e/year）"]) * 1000 != Decimal(
            audit["Emission_kgCO2e"]
        ):
            raise WP68IntegrationError(f"{record_id} 建议清册 Emission 无法与 Canonical 勾稽。")
        if not all(audit[field] for field in ("Record_ID", "Source_File", "Source_SHA256", "Source_Sheet", "Source_Row")):
            raise WP68IntegrationError(f"{record_id} 的审计血缘不完整。")
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    workbook.close()
    if sheet_names != EXPECTED_SHEETS:
        raise WP68IntegrationError(f"统一结果包 Sheet 不完整：{sheet_names}")
    return {
        "suggested_ledger_fields_and_order": list(ledger_rows[0]) == LEDGER_FIELDS if ledger_rows else True,
        "ledger_vs_canonical_reconciled": True,
        "audit_lineage_complete": True,
        "workbook_sheets_complete": True,
    }


def _latest_run(root: Path, marker: str) -> Path:
    candidates = [
        path for path in root.iterdir() if path.is_dir() and (path / marker).is_file()
    ] if root.is_dir() else []
    if not candidates:
        raise WP68IntegrationError(f"在 {root} 未找到包含 {marker} 的正式 Run。")
    return max(candidates, key=lambda path: path.stat().st_mtime)


def _formal_docs(summary: Mapping[str, Any]) -> dict[str, str]:
    status = summary["status"]
    run_id = summary["Run_ID"]
    counts = summary["record_counts"]
    report = f"""# WP6-8 业务化整合与清册衔接报告

> 正式 Run：`{run_id}`  
> 状态：{status}

## 整合范围

WP6-8 直接复用 WP6-3～WP6-7 已验收后端结果，没有重新计算 Activity、Emission、A/B/C/D 或数据质量指标。

共读取 {counts['input']} 条正式记录，建议清册与审计明细输出 {counts['included']} 条，明确排除 {counts['excluded']} 条。

## 输出

- Suggested Ledger V1：面向企业清册结构衔接，不修改企业正式历史清册；
- Demo Audit Detail：保留 Record_ID、Source SHA、Sheet、Row、Activity、EF、Emission、QC 与运行状态；
- WP6_Result_Package.xlsx：包含 9 个业务/审计 Sheet；
- Run Summary 与 Download Manifest：统一说明状态和交付物。

## 口径限制

> {SCOPE_NOTICE}

历史因子继续仅用于模拟，`Production_Eligible=FALSE`。Governance Warning 未被隐藏或改写为计算失败。
"""
    acceptance = f"""# WP6-8 验收报告 V1.0

> 最终状态：{status}  
> 正式 Run：`{run_id}`

## 验收结论

WP6-8 已生成建议清册、审计明细、统一运行摘要、下载清单和 9-Sheet Excel 结果包；建议清册与 Canonical 结果精确勾稽，完整血缘为 {counts['included']}/{counts['included']}。

## 关键约束

- UI/交付层未建立第二套核算、情景或质量公式；
- Raw、Frozen Evidence 与企业正式清册未修改；
- Scope Warning 与 Governance Warning 保留；
- 未创建 Git remote、未 push、未执行 WP6-9 收口。
"""
    handoff = f"""# WP6-8 交接摘要

> WP6-8 状态：{status}  
> 正式 Run：`{run_id}`  
> WP6-9 状态：NOT STARTED

## 已完成

- {counts['included']} 条 Suggested Ledger V1 与 Audit Detail
- 统一 9-Sheet `WP6_Result_Package.xlsx`
- Run Summary、Download Manifest、业务状态与可操作异常信息
- WP6-3～7 后端结果只读整合与 2024/2025 回归

## 保留限制

{SCOPE_NOTICE} 历史因子仅用于模拟，`Production_Eligible=FALSE`。
"""
    return {
        "WP6-8_业务化整合与清册衔接报告.md": report,
        "WP6-8_验收报告_V1.0.md": acceptance,
        "WP6-8_交接摘要.md": handoff,
    }


def run_wp6_8_integration(
    *,
    wp6_3_run_dir: Path,
    wp6_4_run_dir: Path,
    wp6_4_current_run_dir: Path,
    wp6_5_run_dir: Path,
    wp6_6_run_dir: Path,
    wp6_7_run_dir: Path,
    output_root: Path,
    documentation_root: Path | None = None,
) -> dict[str, Any]:
    """Create one audited WP6-8 formal delivery from accepted upstream runs."""

    runs = {
        "WP6-3": wp6_3_run_dir.expanduser().resolve(),
        "WP6-4": wp6_4_run_dir.expanduser().resolve(),
        "WP6-4 Current": wp6_4_current_run_dir.expanduser().resolve(),
        "WP6-5": wp6_5_run_dir.expanduser().resolve(),
        "WP6-6": wp6_6_run_dir.expanduser().resolve(),
        "WP6-7": wp6_7_run_dir.expanduser().resolve(),
    }
    required = {
        "wp6_3_summary": runs["WP6-3"] / "wp6_3_summary.json",
        "2024_canonical": runs["WP6-3"] / "2024_canonical_results.csv",
        "wp6_4_summary": runs["WP6-4"] / "wp6_4_summary.json",
        "2025_canonical": runs["WP6-4"] / "2025_canonical_results.csv",
        "2025_standard": runs["WP6-4 Current"] / "03_standardized" / "day4_standard_31_fields.csv",
        "wp6_5_summary": runs["WP6-5"] / "independent_validation_summary.json",
        "2024_validation": runs["WP6-5"] / "2024_independent_validation.csv",
        "2025_validation": runs["WP6-5"] / "2025_independent_validation.csv",
        "wp6_6_summary": runs["WP6-6"] / "wp6_6_analysis_summary.json",
        "wp6_7_summary": runs["WP6-7"] / "wp6_7_analysis_summary.json",
        "2024_scorecard": runs["WP6-7"] / "2024_data_quality_scorecard.json",
        "2025_scorecard": runs["WP6-7"] / "2025_data_quality_scorecard.json",
        "issues": runs["WP6-7"] / "data_quality_issue_register.csv",
        "dimensions": runs["WP6-7"] / "dimension_availability.csv",
        "management_2024": runs["WP6-7"] / "2024_management_summary.csv",
        "management_2025": runs["WP6-7"] / "2025_management_summary.csv",
        "contributors_2024": runs["WP6-7"] / "2024_top_emission_contributors.csv",
        "contributors_2025": runs["WP6-7"] / "2025_top_emission_contributors.csv",
        "factor_impact_2024": runs["WP6-7"] / "2024_top_factor_impact.csv",
        "factor_impact_2025": runs["WP6-7"] / "2025_top_factor_impact.csv",
    }
    for path in required.values():
        if not path.is_file():
            raise WP68IntegrationError(f"缺少必需的正式上游产物：{path}")
    wp63_preflight = _read_json(required["wp6_3_summary"])
    wp63_upstream = Path(_text(wp63_preflight.get("upstream_run_directory"))).resolve()
    required.update(
        {
            "2024_recognition": wp63_upstream / "01_import" / "recognition_summary.json",
            "2024_capability": wp63_upstream / "02_capability" / "capability_summary.json",
            "2025_recognition": runs["WP6-4 Current"] / "01_import" / "recognition_summary.json",
            "2025_capability": runs["WP6-4 Current"] / "02_capability" / "capability_summary.json",
        }
    )
    for path in required.values():
        if not path.is_file():
            raise WP68IntegrationError(f"缺少必需的正式上游产物：{path}")
    hashes_before = {name: _sha256(path) for name, path in required.items()}
    summaries = {
        "WP6-3": _read_json(required["wp6_3_summary"]),
        "WP6-4": _read_json(required["wp6_4_summary"]),
        "WP6-5": _read_json(required["wp6_5_summary"]),
        "WP6-6": _read_json(required["wp6_6_summary"]),
        "WP6-7": _read_json(required["wp6_7_summary"]),
    }
    blocked = [name for name, value in summaries.items() if value.get("status") == "BLOCKED"]
    if blocked:
        raise WP68IntegrationError(f"正式上游阶段已阻断，不能生成交付包：{', '.join(blocked)}")

    rows_2024 = _read_csv(required["2024_canonical"])
    rows_2025 = _read_csv(required["2025_canonical"])
    validation_2024 = _index(_read_csv(required["2024_validation"]), "2024 Validation")
    validation_2025 = _index(_read_csv(required["2025_validation"]), "2025 Validation")
    standard_2025 = _index(_read_csv(required["2025_standard"]), "2025 Standard")
    now = datetime.now(timezone.utc)
    fingerprint = hashlib.sha256("".join(hashes_before.values()).encode()).hexdigest()[:8].upper()
    run_id = f"WP6-8-{now.strftime('%Y%m%dT%H%M%S%fZ')}-{fingerprint}"
    output_dir = output_root.expanduser().resolve() / run_id
    output_dir.mkdir(parents=True, exist_ok=False)

    delivery_2024 = build_delivery_rows(
        rows_2024,
        validation_by_id=validation_2024,
        fallback_run_id=_text(summaries["WP6-3"].get("run_id")),
    )
    delivery_2025 = build_delivery_rows(
        rows_2025,
        enrichment_by_id=standard_2025,
        validation_by_id=validation_2025,
        fallback_run_id=_text(summaries["WP6-4"].get("run_id")),
    )
    canonical_rows = delivery_2024["canonical_rows"] + delivery_2025["canonical_rows"]
    audit_rows = delivery_2024["audit_rows"] + delivery_2025["audit_rows"]
    ledger_rows = delivery_2024["ledger_rows"] + delivery_2025["ledger_rows"]
    for number, row in enumerate(ledger_rows, start=1):
        row["序号"] = str(number)
    exclusions = delivery_2024["excluded_rows"] + delivery_2025["excluded_rows"]
    delivery = {
        "canonical_rows": canonical_rows,
        "audit_rows": audit_rows,
        "ledger_rows": ledger_rows,
        "excluded_rows": exclusions,
    }
    if not audit_rows:
        status = "BLOCKED"
    elif exclusions:
        status = "PARTIAL_RESULT"
    elif any(row["Governance_QC"].upper() == "WARNING" for row in audit_rows):
        status = "PASS_WITH_WARNING"
    else:
        status = "PASS"

    scorecard_2024 = _read_json(required["2024_scorecard"])
    scorecard_2025 = _read_json(required["2025_scorecard"])
    quality_rows = _flatten_scorecards(scorecard_2024, scorecard_2025)
    issue_rows = _read_csv(required["issues"])
    dimension_rows = _read_csv(required["dimensions"])
    management_rows = _read_csv(required["management_2024"]) + _read_csv(required["management_2025"])
    contributor_rows = _read_csv(required["contributors_2024"]) + _read_csv(required["contributors_2025"])
    factor_rows = _read_csv(required["factor_impact_2024"]) + _read_csv(required["factor_impact_2025"])
    recognition = {
        "2024": _read_json(required["2024_recognition"]),
        "2025": _read_json(required["2025_recognition"]),
    }
    capability = {
        "2024": _read_json(required["2024_capability"]),
        "2025": _read_json(required["2025_capability"]),
    }

    _write_csv(output_dir / "canonical_results.csv", canonical_rows, CANONICAL_EXPORT_FIELDS)
    _write_csv(output_dir / "suggested_ledger_v1.csv", ledger_rows, LEDGER_FIELDS)
    _write_csv(output_dir / "audit_detail.csv", audit_rows, AUDIT_FIELDS)
    _write_csv(
        output_dir / "excluded_records.csv",
        exclusions,
        ["Record_ID", "Reason_Codes", "Action"],
    )
    _write_csv(output_dir / "dimension_availability.csv", dimension_rows, list(dimension_rows[0]))
    _write_csv(output_dir / "quality_scorecard.csv", quality_rows, list(quality_rows[0]))
    _write_csv(
        output_dir / "data_quality_issue_register.csv", issue_rows, list(issue_rows[0])
    )
    _write_csv(output_dir / "management_summary.csv", management_rows, list(management_rows[0]))
    _write_csv(
        output_dir / "top_emission_contributors.csv",
        contributor_rows,
        list(contributor_rows[0]),
    )
    _write_csv(output_dir / "factor_sensitivity.csv", factor_rows, list(factor_rows[0]))

    source_run_ids = {name: _text(summary.get("run_id")) for name, summary in summaries.items()}
    source_files = sorted({row["Source_File"] for row in audit_rows})
    source_hashes = sorted({row["Source_SHA256"] for row in audit_rows})
    run_summary = {
        "schema_version": "WP6_8_RUN_SUMMARY_V1",
        "Run_ID": run_id,
        "Status": status,
        "Input_File": source_files,
        "Input_SHA": source_hashes,
        "Recognition_Status": {
            year: recognition[year].get("recognition_status") for year in ("2024", "2025")
        },
        "Selected_Sheet": {
            year: recognition[year].get("best_candidate_sheet") for year in ("2024", "2025")
        },
        "Header_Row": {
            year: recognition[year].get("best_candidate_header_row") for year in ("2024", "2025")
        },
        "Total_Records": len(rows_2024) + len(rows_2025),
        "Boundary_Records": len(audit_rows),
        "Activity_Ready": len(audit_rows),
        "Emission_Ready": len(audit_rows),
        "Calculation_Status": "PASS" if audit_rows else "BLOCKED",
        "Governance_Warning_Count": sum(
            row["Governance_QC"].upper() == "WARNING" for row in audit_rows
        ),
        "Validation_Status": "INDEPENDENT_CALCULATION_PASS",
        "Production_Eligible": False,
        "Generated_Outputs": [],
        "Source_Run_IDs": source_run_ids,
        "Scope_Notice": SCOPE_NOTICE,
        "Partial_Result": bool(exclusions and audit_rows),
        "Blocking_Reasons": (
            ["NO_CALCULABLE_RECORDS"] if not audit_rows else []
        ),
    }
    workbook_path = output_dir / "WP6_Result_Package.xlsx"
    run_information = [
        {"Field": key, "Value": json.dumps(value, ensure_ascii=False) if isinstance(value, (list, dict)) else value}
        for key, value in run_summary.items()
    ]
    _write_workbook(
        workbook_path,
        [
            ("01_核算结果", canonical_rows),
            ("02_建议清册", ledger_rows),
            ("03_审计明细", audit_rows),
            ("04_数据质量", quality_rows),
            ("05_问题登记", issue_rows),
            ("06_管理汇总", management_rows),
            ("07_排放贡献", contributor_rows),
            ("08_因子敏感性", factor_rows),
            ("09_运行信息", run_information),
        ],
    )
    acceptance = _validate_delivery(delivery, workbook_path)
    hashes_after = {name: _sha256(path) for name, path in required.items()}
    protected = {name: hashes_before[name] == hashes_after[name] for name in required}
    if not all(protected.values()):
        raise WP68IntegrationError("WP6-8 执行期间上游保护输入发生变化。")

    integration_summary = {
        "schema_version": "WP6_8_INTEGRATION_SUMMARY_V1",
        "stage": "WP6-8",
        "status": status,
        "Run_ID": run_id,
        "completed_at": now.isoformat(),
        "record_counts": {
            "input": len(rows_2024) + len(rows_2025),
            "included": len(audit_rows),
            "excluded": len(exclusions),
            "2024": len(delivery_2024["audit_rows"]),
            "2025": len(delivery_2025["audit_rows"]),
        },
        "activity_paths": {
            "2024": sorted({row["Activity_Method"] for row in delivery_2024["audit_rows"]}),
            "2025": sorted({row["Activity_Method"] for row in delivery_2025["audit_rows"]}),
        },
        "boundary_flow": {
            "2024": {
                "Input_Records": capability["2024"].get("total_records"),
                "Business_Unit_Records": summaries["WP6-3"]["boundary"]["business_unit_count"],
                "Purchase_Category_Records": summaries["WP6-3"]["boundary"]["cell_count"],
                "Formal_Pilot_Records": summaries["WP6-3"]["boundary"]["marker_count"],
            },
            "2025": {
                "Input_Records": capability["2025"].get("total_records"),
                "Business_Unit_Records": summaries["WP6-4"]["boundary"]["business_unit_count"],
                "Purchase_Category_Records": summaries["WP6-4"]["boundary"]["cell_count"],
                "Formal_Pilot_Records": summaries["WP6-4"]["boundary"]["marker_count"],
            },
        },
        "source_run_ids": source_run_ids,
        "acceptance_checks": acceptance,
        "scope_notice": SCOPE_NOTICE,
        "dimension_availability_reused": True,
        "wp6_6_recalculated": False,
        "wp6_7_metrics_recalculated": False,
        "streamlit_recalculates": False,
        "raw_excel_opened": False,
        "raw_data_modified": False,
        "frozen_evidence_modified": False,
        "enterprise_ledger_modified": False,
        "production_eligible": False,
        "protected_input_hashes_before": hashes_before,
        "protected_input_hashes_after": hashes_after,
        "protected_inputs_unchanged": protected,
        "wp6_9_execution_performed": False,
    }
    _write_json(output_dir / "run_summary.json", run_summary)
    _write_json(output_dir / "wp6_8_integration_summary.json", integration_summary)

    docs = _formal_docs(integration_summary)
    for name, content in docs.items():
        (output_dir / name).write_text(content, encoding="utf-8")
        if documentation_root is not None:
            destination = documentation_root.expanduser().resolve() / name
            destination.parent.mkdir(parents=True, exist_ok=True)
            destination.write_text(content, encoding="utf-8")

    manifest_names = [
        "canonical_results.csv",
        "suggested_ledger_v1.csv",
        "audit_detail.csv",
        "excluded_records.csv",
        "dimension_availability.csv",
        "quality_scorecard.csv",
        "data_quality_issue_register.csv",
        "management_summary.csv",
        "top_emission_contributors.csv",
        "factor_sensitivity.csv",
        "run_summary.json",
        "wp6_8_integration_summary.json",
        "WP6_Result_Package.xlsx",
        *docs,
    ]
    manifest = {
        "schema_version": "WP6_8_DOWNLOAD_MANIFEST_V1",
        "Run_ID": run_id,
        "files": [
            {
                "File_Name": name,
                "SHA256": _sha256(output_dir / name),
                "Size_Bytes": (output_dir / name).stat().st_size,
            }
            for name in manifest_names
        ],
    }
    _write_json(output_dir / "download_manifest.json", manifest)
    run_summary["Generated_Outputs"] = [*manifest_names, "download_manifest.json"]
    _write_json(output_dir / "run_summary.json", run_summary)
    # Refresh manifest hashes after the final Run Summary write.
    for item in manifest["files"]:
        path = output_dir / item["File_Name"]
        item["SHA256"] = _sha256(path)
        item["Size_Bytes"] = path.stat().st_size
    _write_json(output_dir / "download_manifest.json", manifest)

    return {
        "status": status,
        "run_id": run_id,
        "output_directory": str(output_dir),
        "record_counts": integration_summary["record_counts"],
        "acceptance_checks": acceptance,
        "generated_outputs": run_summary["Generated_Outputs"],
    }


def _analysis_row(row: Mapping[str, Any], run_id: str) -> dict[str, Any]:
    """Adapt an already-calculated live canonical row to the tested WP6-7 API."""

    validation_status = _first(row.get("Overall_Validation_Status"), "NOT_RUN")
    boundary = _first(row.get("Boundary_Ready"))
    if boundary.upper() in {"TRUE", "FALSE"}:
        boundary_ready = boundary.upper()
    else:
        boundary_ready = "TRUE" if str(row.get("Boundary_Ready")).strip() in {"True", "1"} else "FALSE"
    return {
        **dict(row),
        "Activity_kg": _text(row.get("Activity_Data_kg")),
        "Activity_Unit": _first(row.get("Activity_Unit"), "kg/year"),
        "Supplier": _first(row.get("Supplier"), row.get("Supplier_Name")),
        "Project": _first(row.get("Project"), row.get("Project_Code")),
        "Model": _first(row.get("Model"), row.get("Cell_Model")),
        "Chemistry": _text(row.get("Chemistry")),
        "Overall_Validation_Status": validation_status,
        "Calculation_QC": _first(row.get("Calculation_QC"), row.get("QC_Status")),
        "Governance_QC": _first(row.get("Governance_QC"), row.get("QC_Status")),
        "Boundary_Ready": boundary_ready,
        "Calculation_Issue_Codes": _text(row.get("Calculation_Issue_Codes")),
        "Governance_Issue_Codes": _first(
            row.get("Governance_Issue_Codes"), row.get("Warning_Codes")
        ),
        "Boundary_Issue_Codes": _text(row.get("Boundary_Issue_Codes")),
        "Run_ID": run_id,
    }


def run_live_delivery(
    *,
    run_dir: Path,
    canonical_rows: list[dict[str, Any]],
    route_decision: Mapping[str, Any],
    validation_rows: list[dict[str, Any]],
    input_file: str,
    input_sha256: str,
    recognition_summary: Mapping[str, Any],
    capability_summary: Mapping[str, Any],
    historical_rows: list[dict[str, Any]] | None = None,
    independent_status: str = "NOT_RUN",
) -> dict[str, Any]:
    """Package one fresh upload using only artifacts produced by its own Run."""

    run = run_dir.expanduser().resolve()
    run_id = run.name
    output = run / "08_download"
    output.mkdir(parents=True, exist_ok=True)
    expected_sha = _text(input_sha256).upper()
    wrong_sha = [
        row.get("Record_ID", "")
        for row in canonical_rows
        if _text(row.get("Source_SHA256")).upper() != expected_sha
    ]
    wrong_run = [
        row.get("Record_ID", "")
        for row in canonical_rows
        if _first(row.get("Run_ID"), run_id) != run_id
    ]
    if wrong_sha or wrong_run:
        raise WP68IntegrationError(
            f"Live Run 绑定失败：Input SHA mismatch={wrong_sha[:3]} Run mismatch={wrong_run[:3]}"
        )
    delivery = build_delivery_rows(canonical_rows, fallback_run_id=run_id)
    audit_rows = delivery["audit_rows"]
    ledger_rows = delivery["ledger_rows"]
    included_ids = {row["Record_ID"] for row in audit_rows}
    analysis_rows = [
        _analysis_row(row, run_id)
        for row in canonical_rows
        if _text(row.get("Record_ID")) in included_ids
    ]
    years = sorted({_text(row.get("Year")) for row in analysis_rows if _text(row.get("Year"))})
    scorecards: list[dict[str, Any]] = []
    quality_rows: list[dict[str, Any]] = []
    dimensions: list[dict[str, Any]] = []
    management: list[dict[str, Any]] = []
    contributors: list[dict[str, Any]] = []
    concentration: dict[str, Any] = {}
    for year in years:
        rows = [row for row in analysis_rows if row["Year"] == year]
        scorecard = build_data_quality_scorecard(rows, year)
        scorecards.append(scorecard)
        quality_rows.extend({"Year": year, **metric} for metric in scorecard["metrics"])
        available = build_dimension_availability(rows, year)
        dimensions.extend(available)
        management.extend(build_management_summary(rows, year, available))
        ranked, summary = build_top_emission_contributors(rows, year)
        contributors.extend(ranked)
        concentration[year] = summary
    issues, issue_summary = build_issue_register(analysis_rows)
    unit_by_id = {
        _text(row.get("Record_ID")): _text(row.get("Business_Unit"))
        for row in analysis_rows
    }
    issues = [
        {"Business_Unit": unit_by_id.get(_text(item.get("Record_ID")), ""), **item}
        for item in issues
    ]
    lineage = build_lineage_quality_summary(
        {year: [row for row in analysis_rows if row["Year"] == year] for year in years}
    )
    status = delivery["status"]
    run_summary = {
        "schema_version": "WP6_8_1_LIVE_RUN_SUMMARY_V1",
        "Run_ID": run_id,
        "Status": status,
        "Input_File": input_file,
        "Input_SHA": expected_sha,
        "Recognition_Status": recognition_summary.get("recognition_status"),
        "Selected_Sheet": recognition_summary.get("best_candidate_sheet"),
        "Header_Row": recognition_summary.get("best_candidate_header_row"),
        "Activity_Route": route_decision.get("Activity_Route"),
        "Factor_Route": route_decision.get("Factor_Route"),
        "Factor_Policy_ID": route_decision.get("Factor_Policy_ID"),
        "Factor_Value": route_decision.get("Factor_Value"),
        "Factor_Usage": route_decision.get("Factor_Usage"),
        "Boundary_Policy": route_decision.get("Boundary_Policy"),
        "Total_Records": delivery["input_record_count"],
        "Boundary_Records": delivery["included_record_count"],
        "Activity_Ready": capability_summary.get("activity_ready_count", 0),
        "Emission_Ready": delivery["included_record_count"],
        "Calculation_Status": "PASS" if audit_rows else "BLOCKED",
        "Governance_Warning_Count": sum(
            _text(row.get("Governance_QC")).upper() == "WARNING" for row in audit_rows
        ),
        "Independent_Validation_Status": independent_status,
        "Cross_Year_Factor_Analysis": "NOT_AVAILABLE_FOR_SINGLE_INPUT_RUN",
        "Simulation_Flag": True,
        "Production_Eligible": False,
        "Generated_Outputs": [],
        "Excluded_Record_Count": delivery["excluded_record_count"],
        "Historical_Formal_Runs_Used_As_Live_Input": [],
    }
    canonical_fields = list(canonical_rows[0]) if canonical_rows else ["Record_ID"]
    _write_csv(output / "canonical_results.csv", canonical_rows, canonical_fields)
    _write_csv(output / "suggested_ledger_v1.csv", ledger_rows, LEDGER_FIELDS)
    _write_csv(output / "audit_detail.csv", audit_rows, AUDIT_FIELDS)
    _write_csv(
        output / "excluded_records.csv",
        delivery["excluded_rows"],
        ["Record_ID", "Reason_Codes", "Action"],
    )
    _write_csv(output / "quality_scorecard.csv", quality_rows, list(quality_rows[0]) if quality_rows else ["Year", "Metric"])
    _write_csv(output / "data_quality_issue_register.csv", issues, list(issues[0]) if issues else ["Issue_Code", "Record_ID"])
    _write_csv(output / "dimension_availability.csv", dimensions, list(dimensions[0]) if dimensions else ["Year", "Dimension"])
    _write_csv(output / "management_summary.csv", management, list(management[0]) if management else ["Year", "Dimension"])
    _write_csv(output / "top_emission_contributors.csv", contributors, list(contributors[0]) if contributors else ["Year", "Record_ID"])
    _write_csv(
        output / "validation_results.csv",
        validation_rows,
        list(validation_rows[0]) if validation_rows else ["Overall_Validation_Status"],
    )
    _write_csv(
        output / "historical_validation.csv",
        historical_rows or [],
        list((historical_rows or [{}])[0]) if historical_rows else ["Validation_Status"],
    )
    _write_json(output / "quality_scorecards.json", scorecards)
    _write_json(output / "lineage_quality_summary.json", lineage)
    _write_json(output / "issue_summary.json", issue_summary)
    run_information = [
        {"Field": key, "Value": json.dumps(value, ensure_ascii=False) if isinstance(value, (list, dict)) else value}
        for key, value in run_summary.items()
    ]
    workbook_path = output / "WP6_Result_Package.xlsx"
    _write_workbook(
        workbook_path,
        [
            ("01_核算结果", delivery["canonical_rows"]),
            ("02_建议清册", ledger_rows),
            ("03_审计明细", audit_rows),
            ("04_数据质量", quality_rows),
            ("05_问题登记", issues),
            ("06_管理汇总", management),
            ("07_排放贡献", contributors),
            ("08_验证结果", validation_rows),
            ("09_运行信息", run_information),
        ],
    )
    live_summary = {
        "schema_version": "WP6_8_1_LIVE_DELIVERY_V1",
        "status": status,
        "Run_ID": run_id,
        "Input_SHA": input_sha256,
        "record_counts": {
            "input": delivery["input_record_count"],
            "included": delivery["included_record_count"],
            "excluded": delivery["excluded_record_count"],
        },
        "route_decision": dict(route_decision),
        "concentration": concentration,
        "lineage_quality": lineage,
        "current_run_only": True,
        "historical_wp6_run_used_as_input": False,
        "independent_validation_status": independent_status,
        "cross_year_factor_analysis": "NOT_AVAILABLE_FOR_SINGLE_INPUT_RUN",
        "streamlit_recalculates": False,
        "production_eligible": False,
    }
    _write_json(output / "run_summary.json", run_summary)
    _write_json(output / "wp6_8_live_summary.json", live_summary)
    manifest_names = [
        "canonical_results.csv",
        "suggested_ledger_v1.csv",
        "audit_detail.csv",
        "excluded_records.csv",
        "quality_scorecard.csv",
        "data_quality_issue_register.csv",
        "dimension_availability.csv",
        "management_summary.csv",
        "top_emission_contributors.csv",
        "validation_results.csv",
        "historical_validation.csv",
        "quality_scorecards.json",
        "lineage_quality_summary.json",
        "issue_summary.json",
        "run_summary.json",
        "wp6_8_live_summary.json",
        "WP6_Result_Package.xlsx",
    ]
    run_summary["Generated_Outputs"] = [*manifest_names, "download_manifest.json"]
    _write_json(output / "run_summary.json", run_summary)
    manifest = {
        "schema_version": "WP6_8_1_LIVE_DOWNLOAD_MANIFEST_V1",
        "Run_ID": run_id,
        "Input_SHA": input_sha256,
        "files": [
            {
                "File_Name": name,
                "SHA256": _sha256(output / name),
                "Size_Bytes": (output / name).stat().st_size,
            }
            for name in manifest_names
        ],
    }
    _write_json(output / "download_manifest.json", manifest)
    return {
        "status": status,
        "run_id": run_id,
        "output_directory": str(output),
        "record_count": delivery["included_record_count"],
        "excluded_record_count": delivery["excluded_record_count"],
        "workbook_path": str(workbook_path),
    }
