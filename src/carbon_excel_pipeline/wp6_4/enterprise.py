"""Read-only extraction of the 2025 enterprise historical cell scope."""

from __future__ import annotations

import hashlib
import warnings
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def _decimal(value: Any) -> Decimal | None:
    if isinstance(value, bool) or value is None:
        return None
    try:
        parsed = Decimal(str(value))
    except Exception:
        return None
    return parsed if parsed.is_finite() else None


def _unit(value: Any) -> str:
    return str(value or "").strip().casefold().replace(" ", "").replace("₂", "2")


def extract_enterprise_cell_scope(
    workbook_path: Path, *, category_label: str = "电芯"
) -> dict[str, Any]:
    """Extract Activity, EF and historical GHG from separate workbook tables."""

    path = workbook_path.expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(path)
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Print area cannot be set to Defined name.*",
            category=UserWarning,
            module="openpyxl.reader.workbook",
        )
        workbook = load_workbook(
            path, read_only=True, data_only=True, keep_links=False
        )
    activity_candidates: list[dict[str, Any]] = []
    ef_candidates: list[dict[str, Any]] = []
    matching_rows: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            for cells in sheet.iter_rows():
                values = [cell.value for cell in cells]
                if not any(str(value or "").strip() == category_label for value in values):
                    continue
                row_info = {"sheet": sheet.title, "row": cells[0].row, "values": values, "cells": cells}
                matching_rows.append(row_info)
                for index, value in enumerate(values[:-1]):
                    parsed = _decimal(value)
                    unit = _unit(values[index + 1])
                    if parsed is None or parsed <= 0:
                        continue
                    if unit in {"kg", "kg/year", "kg/年"}:
                        activity_candidates.append(
                            {
                                "value": parsed,
                                "sheet": sheet.title,
                                "row": cells[index].row,
                                "cell": cells[index].coordinate,
                                "unit_cell": cells[index + 1].coordinate,
                                "unit": str(values[index + 1]),
                            }
                        )
                    if unit in {"kgco2/kg", "kgco2e/kg"}:
                        ef_candidates.append(
                            {
                                "value": parsed,
                                "sheet": sheet.title,
                                "row": cells[index].row,
                                "cell": cells[index].coordinate,
                                "unit_cell": cells[index + 1].coordinate,
                                "unit": str(values[index + 1]),
                            }
                        )
    finally:
        workbook.close()
    if len(activity_candidates) != 1 or len(ef_candidates) != 1:
        raise ValueError(
            "Enterprise cell Activity/EF could not be identified uniquely: "
            f"activity={len(activity_candidates)}, ef={len(ef_candidates)}"
        )
    activity = activity_candidates[0]
    factor = ef_candidates[0]
    expected_t = activity["value"] * factor["value"] / Decimal("1000")
    emission_candidates: list[dict[str, Any]] = []
    for row in matching_rows:
        row_decimals = [_decimal(value) for value in row["values"]]
        contains_activity = any(value == activity["value"] for value in row_decimals)
        contains_factor = any(value == factor["value"] for value in row_decimals)
        if not (contains_activity and contains_factor):
            continue
        for index, value in enumerate(row_decimals):
            if value is None:
                continue
            difference = abs(value - expected_t)
            if difference <= Decimal("0.000001"):
                cell = row["cells"][index]
                emission_candidates.append(
                    {
                        "value": value,
                        "sheet": row["sheet"],
                        "row": row["row"],
                        "cell": cell.coordinate,
                        "difference_from_activity_times_ef_t": difference,
                    }
                )
    if not emission_candidates:
        raise ValueError("Enterprise historical cell emission was not found.")
    emission = sorted(
        emission_candidates,
        key=lambda item: (item["difference_from_activity_times_ef_t"], item["sheet"], item["cell"]),
    )[0]
    return {
        "source_file": path.name,
        "source_path": str(path),
        "source_sha256": _sha256(path),
        "scope": "ENTERPRISE_ALL_CELL_CATEGORY",
        "category": category_label,
        "activity_kg": format(activity["value"], "f"),
        "ef_kgco2e_per_kg": format(factor["value"], "f"),
        "historical_emission_tco2e": format(emission["value"], "f"),
        "calculated_cross_check_tco2e": format(expected_t, "f"),
        "formula_cache_difference_tco2e": format(emission["value"] - expected_t, "f"),
        "activity_source": {key: str(value) for key, value in activity.items() if key != "value"},
        "ef_source": {key: str(value) for key, value in factor.items() if key != "value"},
        "emission_source": {
            key: str(value)
            for key, value in emission.items()
            if key not in {"value", "difference_from_activity_times_ef_t"}
        },
        "read_only": True,
    }
