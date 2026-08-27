"""WP6-6 A/B/C/D counterfactual analysis over WP6-5 validated records."""

from __future__ import annotations

import csv
import hashlib
import json
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation, localcontext
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


PASS_STATUS = "INDEPENDENT_CALCULATION_PASS"
TRUE_VALUES = {"true", "1", "yes", "pass"}
COUNTERFACTUAL_FIELDS = [
    "Year",
    "Record_ID",
    "Source_File",
    "Source_SHA256",
    "Source_Sheet",
    "Source_Row",
    "Activity_Method",
    "Validated_Activity_kg",
    "EF_2024",
    "EF_2025",
    "Emission_Using_EF_2024_kgCO2e",
    "Emission_Using_EF_2025_kgCO2e",
    "Actual_Emission_kgCO2e",
    "Counterfactual_Emission_kgCO2e",
    "Factor_Impact_kgCO2e",
    "Factor_Impact_tCO2e",
    "Factor_Impact_Percent",
    "Governance_QC",
    "Boundary_Ready",
    "Lineage_Validation_Status",
    "Historical_Reproduction",
    "Simulation_Flag",
    "Production_Eligible",
]


class WP66AnalysisError(RuntimeError):
    """Raised when validated inputs cannot support the formal scenario model."""


def _decimal(value: Any, field: str) -> Decimal:
    if isinstance(value, bool) or value is None or str(value).strip() == "":
        raise WP66AnalysisError(f"{field} is missing")
    try:
        parsed = Decimal(str(value).strip())
    except (InvalidOperation, ValueError) as error:
        raise WP66AnalysisError(f"{field} is not a finite decimal: {value}") from error
    if not parsed.is_finite():
        raise WP66AnalysisError(f"{field} is not a finite decimal: {value}")
    return parsed


def _format(value: Decimal) -> str:
    return format(value, "f")


def _truth(value: Any) -> bool:
    return str(value or "").strip().casefold() in TRUE_VALUES


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def _validation_reasons(row: dict[str, str], expected_year: str) -> list[str]:
    reasons: list[str] = []
    if row.get("Year") != expected_year:
        reasons.append("YEAR_MISMATCH")
    if row.get("Overall_Validation_Status") != PASS_STATUS:
        reasons.append("WP6_5_INDEPENDENT_VALIDATION_NOT_PASS")
    if not _truth(row.get("Boundary_Ready")):
        reasons.append("BOUNDARY_NOT_READY")
    if row.get("Lineage_Validation_Status") != "PASS":
        reasons.append("LINEAGE_NOT_PASS")
    for field in (
        "Activity_Validation_Status",
        "EF_Validation_Status",
        "Emission_Validation_Status",
    ):
        if row.get(field) != "PASS_EXACT":
            reasons.append(f"{field.upper()}_NOT_PASS_EXACT")
    if not row.get("Record_ID", "").startswith(f"{expected_year}-"):
        reasons.append("RECORD_ID_YEAR_NAMESPACE_MISMATCH")
    return reasons


def _partition_rows(
    rows: list[dict[str, str]], year: str
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    eligible: list[dict[str, str]] = []
    excluded: list[dict[str, str]] = []
    seen: set[str] = set()
    for row in rows:
        record_id = row.get("Record_ID", "")
        reasons = _validation_reasons(row, year)
        if not record_id:
            reasons.append("RECORD_ID_MISSING")
        elif record_id in seen:
            reasons.append("DUPLICATE_RECORD_ID")
        seen.add(record_id)
        if reasons:
            excluded.append(
                {
                    "Year": year,
                    "Record_ID": record_id,
                    "Excluded_From_Scenario": "TRUE",
                    "Reason_Codes": "|".join(reasons),
                }
            )
        else:
            eligible.append(row)
    return eligible, excluded


def _ef_audit(rows: list[dict[str, str]], year: str) -> dict[str, Any]:
    if not rows:
        raise WP66AnalysisError(f"{year} has no eligible WP6-5 records")
    factors = [_decimal(row.get("Independent_EF"), f"{year} Independent_EF") for row in rows]
    activities = [
        _decimal(row.get("Independent_Activity_kg"), f"{year} Independent_Activity_kg")
        for row in rows
    ]
    unique = sorted(set(factors))
    activity_total = sum(activities, Decimal("0"))
    if activity_total == 0:
        raise WP66AnalysisError(f"{year} eligible activity total is zero")
    with localcontext() as context:
        context.prec = 50
        weighted = sum(
            (activity * factor for activity, factor in zip(activities, factors)),
            Decimal("0"),
        ) / activity_total
    audit = {
        "Year": year,
        "EF_Unique_Count": len(unique),
        "EF_Min": _format(min(unique)),
        "EF_Max": _format(max(unique)),
        "Activity_Weighted_EF": _format(weighted),
        "Single_EF_Assumption_Eligible": len(unique) == 1,
        "Warning_Code": "" if len(unique) == 1 else "MULTIPLE_EF_REQUIRES_MODEL_REASSESSMENT",
    }
    if len(unique) != 1:
        raise WP66AnalysisError(
            f"{year} MULTIPLE_EF_REQUIRES_MODEL_REASSESSMENT: {len(unique)} distinct EF values"
        )
    return audit


def _counterfactual_rows(
    rows: list[dict[str, str]], year: str, ef_2024: Decimal, ef_2025: Decimal
) -> list[dict[str, str]]:
    output: list[dict[str, str]] = []
    with localcontext() as context:
        context.prec = 50
        percent = (ef_2025 / ef_2024 - Decimal("1")) * Decimal("100")
        for row in rows:
            activity = _decimal(
                row.get("Independent_Activity_kg"), f"{year} Independent_Activity_kg"
            )
            emission_2024 = activity * ef_2024
            emission_2025 = activity * ef_2025
            impact = emission_2025 - emission_2024
            actual = emission_2024 if year == "2024" else emission_2025
            counterfactual = emission_2025 if year == "2024" else emission_2024
            validated_actual = _decimal(
                row.get("Independent_Emission_kgCO2e"),
                f"{year} Independent_Emission_kgCO2e",
            )
            if actual != validated_actual:
                raise WP66AnalysisError(
                    f"{year} validated emission no longer reconciles for {row.get('Record_ID')}"
                )
            output.append(
                {
                    "Year": year,
                    "Record_ID": row.get("Record_ID", ""),
                    "Source_File": row.get("Source_File", ""),
                    "Source_SHA256": row.get("Source_SHA256", ""),
                    "Source_Sheet": row.get("Source_Sheet", ""),
                    "Source_Row": row.get("Source_Row", ""),
                    "Activity_Method": row.get("Activity_Method", ""),
                    "Validated_Activity_kg": _format(activity),
                    "EF_2024": _format(ef_2024),
                    "EF_2025": _format(ef_2025),
                    "Emission_Using_EF_2024_kgCO2e": _format(emission_2024),
                    "Emission_Using_EF_2025_kgCO2e": _format(emission_2025),
                    "Actual_Emission_kgCO2e": _format(actual),
                    "Counterfactual_Emission_kgCO2e": _format(counterfactual),
                    "Factor_Impact_kgCO2e": _format(impact),
                    "Factor_Impact_tCO2e": _format(impact / Decimal("1000")),
                    "Factor_Impact_Percent": _format(percent),
                    "Governance_QC": row.get("Governance_QC", ""),
                    "Boundary_Ready": row.get("Boundary_Ready", ""),
                    "Lineage_Validation_Status": row.get(
                        "Lineage_Validation_Status", ""
                    ),
                    "Historical_Reproduction": "TRUE",
                    "Simulation_Flag": "TRUE",
                    "Production_Eligible": "FALSE",
                }
            )
    return output


def _effect_payload(name: str, value_kg: Decimal) -> dict[str, str]:
    return {
        "Effect_Name": name,
        "Effect_kgCO2e": _format(value_kg),
        "Effect_tCO2e": _format(value_kg / Decimal("1000")),
    }


def build_scenario_analysis(
    rows_2024: list[dict[str, str]], rows_2025: list[dict[str, str]]
) -> dict[str, Any]:
    """Build exact scenario results without matching records across years."""

    eligible_2024, excluded_2024 = _partition_rows(rows_2024, "2024")
    eligible_2025, excluded_2025 = _partition_rows(rows_2025, "2025")
    audit_2024 = _ef_audit(eligible_2024, "2024")
    audit_2025 = _ef_audit(eligible_2025, "2025")
    ef_2024 = _decimal(audit_2024["EF_Min"], "EF_2024")
    ef_2025 = _decimal(audit_2025["EF_Min"], "EF_2025")
    if ef_2024 == 0:
        raise WP66AnalysisError("EF_2024 is zero; relative factor change is undefined")

    activity_2024 = sum(
        (_decimal(row["Independent_Activity_kg"], "2024 activity") for row in eligible_2024),
        Decimal("0"),
    )
    activity_2025 = sum(
        (_decimal(row["Independent_Activity_kg"], "2025 activity") for row in eligible_2025),
        Decimal("0"),
    )
    with localcontext() as context:
        context.prec = 50
        a = activity_2024 * ef_2024
        b = activity_2024 * ef_2025
        c = activity_2025 * ef_2024
        d = activity_2025 * ef_2025
        factor_2024 = b - a
        factor_2025 = d - c
        activity_scope_2024_ef = c - a
        activity_scope_2025_ef = d - b
        observed = d - a
        symmetric_factor = (factor_2024 + factor_2025) / Decimal("2")
        symmetric_activity_scope = (
            activity_scope_2024_ef + activity_scope_2025_ef
        ) / Decimal("2")
        decomposition_difference = symmetric_factor + symmetric_activity_scope - observed
        ef_absolute_change = ef_2025 - ef_2024
        ef_relative_percent = (ef_2025 / ef_2024 - Decimal("1")) * Decimal("100")

    scenarios: dict[str, dict[str, Any]] = {}
    for code, activity_year, ef_year, records, activity, factor, emission, meaning in (
        ("A", "2024", "2024", len(eligible_2024), activity_2024, ef_2024, a, "2024 actual validated scenario"),
        ("B", "2024", "2025", len(eligible_2024), activity_2024, ef_2025, b, "2024 activity/scope held constant with 2025 EF"),
        ("C", "2025", "2024", len(eligible_2025), activity_2025, ef_2024, c, "2025 activity/scope held constant with 2024 EF"),
        ("D", "2025", "2025", len(eligible_2025), activity_2025, ef_2025, d, "2025 actual validated scenario"),
    ):
        scenarios[code] = {
            "Scenario": code,
            "Activity_Year": activity_year,
            "EF_Year": ef_year,
            "Record_Count": records,
            "Activity_kg": _format(activity),
            "EF_kgCO2e_per_kg": _format(factor),
            "Emission_kgCO2e": _format(emission),
            "Emission_tCO2e": _format(emission / Decimal("1000")),
            "Meaning": meaning,
        }

    counterfactual_2024 = _counterfactual_rows(
        eligible_2024, "2024", ef_2024, ef_2025
    )
    counterfactual_2025 = _counterfactual_rows(
        eligible_2025, "2025", ef_2024, ef_2025
    )
    factor_effect = {
        "EF_2024": _format(ef_2024),
        "EF_2025": _format(ef_2025),
        "Absolute_EF_Change": _format(ef_absolute_change),
        "Relative_EF_Change_Percent": _format(ef_relative_percent),
        "Factor_Effect_At_2024_Activity": _effect_payload("B_MINUS_A", factor_2024),
        "Factor_Effect_At_2025_Activity": _effect_payload("D_MINUS_C", factor_2025),
        "Counterfactual_Condition": "holding activity/scope constant",
        "Historical_Reproduction": True,
        "Simulation_Flag": True,
        "Production_Eligible": False,
    }
    activity_scope_effect = {
        "Activity_Scope_Combined_Effect_At_2024_EF": _effect_payload(
            "C_MINUS_A", activity_scope_2024_ef
        ),
        "Activity_Scope_Combined_Effect_At_2025_EF": _effect_payload(
            "D_MINUS_B", activity_scope_2025_ef
        ),
        "Observed_Cross_Year_Difference": _effect_payload("D_MINUS_A", observed),
        "Interpretation": "May jointly reflect activity, record count, product structure, and pilot scope changes.",
    }
    symmetric = {
        "Method": "Symmetric / Shapley-style two-factor decomposition",
        "Symmetric_Factor_Contribution": _effect_payload(
            "AVERAGE_FACTOR_PATHS", symmetric_factor
        ),
        "Symmetric_Activity_Scope_Contribution": _effect_payload(
            "AVERAGE_ACTIVITY_SCOPE_PATHS", symmetric_activity_scope
        ),
        "Observed_Cross_Year_Difference": _effect_payload("D_MINUS_A", observed),
        "Reconciliation_Difference_kgCO2e": _format(decomposition_difference),
        "Reconciliation_Status": "PASS_EXACT" if decomposition_difference == 0 else "FAIL",
    }
    excluded = excluded_2024 + excluded_2025
    input_total = len(rows_2024) + len(rows_2025)
    eligible_total = len(eligible_2024) + len(eligible_2025)
    coverage = (
        Decimal(eligible_total) / Decimal(input_total) * Decimal("100")
        if input_total
        else Decimal("0")
    )
    return {
        "status": "PASS" if not excluded else "PASS_WITH_WARNING",
        "scenario_abcd": scenarios,
        "factor_effect": factor_effect,
        "activity_scope_effect": activity_scope_effect,
        "symmetric_decomposition": symmetric,
        "ef_audit": {"2024": audit_2024, "2025": audit_2025},
        "record_counterfactuals": {
            "2024": counterfactual_2024,
            "2025": counterfactual_2025,
        },
        "excluded_records": excluded,
        "coverage": {
            "Input_Record_Count": input_total,
            "Eligible_Record_Count": eligible_total,
            "Excluded_Record_Count": len(excluded),
            "Coverage_Percent": _format(coverage),
            "2024_Eligible": len(eligible_2024),
            "2025_Eligible": len(eligible_2025),
        },
        "identifiability": {
            "FACTOR_EFFECT": "IDENTIFIABLE",
            "ACTIVITY_SCOPE_COMBINED_EFFECT": "IDENTIFIABLE",
            "PURE_ACTIVITY_EFFECT": "NOT_SEPARATELY_IDENTIFIABLE",
            "PRODUCT_STRUCTURE_EFFECT": "NOT_SEPARATELY_IDENTIFIABLE",
            "PURE_SCOPE_EFFECT": "NOT_SEPARATELY_IDENTIFIABLE",
        },
        "scope_safety": {
            "cross_year_record_matching_performed": False,
            "record_id_suffix_matching_performed": False,
            "scope_warning": "2024 and 2025 pilot record scopes differ; activity, product structure, and scope cannot be separated with current evidence.",
        },
    }


def _report(summary: dict[str, Any]) -> str:
    scenarios = summary["scenario_abcd"]
    factor = summary["factor_effect"]
    activity_scope = summary["activity_scope_effect"]
    symmetric = summary["symmetric_decomposition"]
    top_2024 = sorted(
        summary["record_counterfactuals"]["2024"],
        key=lambda row: abs(Decimal(row["Factor_Impact_kgCO2e"])),
        reverse=True,
    )[:5]
    top_2025 = sorted(
        summary["record_counterfactuals"]["2025"],
        key=lambda row: abs(Decimal(row["Factor_Impact_kgCO2e"])),
        reverse=True,
    )[:5]
    lines = [
        "# WP6-6 跨年度因子影响分析报告",
        "",
        f"> 正式 Run：`{summary['run_id']}`  ",
        f"> 状态：{summary['status']}  ",
        "> 历史因子情景分析；Simulation_Flag=TRUE；Production_Eligible=FALSE",
        "",
        "## 数据来源与覆盖",
        "",
        f"输入仅来自 WP6-5 正式 Run `{summary['source_wp6_5_run_id']}` 的独立验证产物。2024 纳入 {summary['coverage']['2024_Eligible']} 条，2025 纳入 {summary['coverage']['2025_Eligible']} 条，覆盖率 {summary['coverage']['Coverage_Percent']}%。",
        "",
        "## Scope 限制",
        "",
        "2024 与 2025 试点记录范围不同。当前可严格识别排放因子影响，但 Activity、产品结构和 Scope 影响不能完全独立拆分。程序未执行跨年 Record_ID、尾号或模糊匹配。",
        "",
        "## 四情景",
        "",
        "|情景|Activity 年度|EF 年度|记录数|排放量（tCO2e）|解释|",
        "|---|---:|---:|---:|---:|---|",
    ]
    for code in ("A", "B", "C", "D"):
        scenario = scenarios[code]
        lines.append(
            f"|{code}|{scenario['Activity_Year']}|{scenario['EF_Year']}|{scenario['Record_Count']}|{scenario['Emission_tCO2e']}|{scenario['Meaning']}|"
        )
    lines.extend(
        [
            "",
            "## EF 变化与固定 Activity/Scope 因子效应",
            "",
            f"EF 从 {factor['EF_2024']} 变为 {factor['EF_2025']} kgCO2e/kg，绝对变化 {factor['Absolute_EF_Change']}，相对变化 {factor['Relative_EF_Change_Percent']}%。",
            "",
            f"- 固定 2024 Activity/Scope：B-A = {factor['Factor_Effect_At_2024_Activity']['Effect_tCO2e']} tCO2e。",
            f"- 固定 2025 Activity/Scope：D-C = {factor['Factor_Effect_At_2025_Activity']['Effect_tCO2e']} tCO2e。",
            "",
            "以上结论只在 holding activity/scope constant 的反事实条件下成立，不代表供应商真实减排。",
            "",
            "## Activity/Scope 综合效应与总差异",
            "",
            f"- 固定 2024 EF：C-A = {activity_scope['Activity_Scope_Combined_Effect_At_2024_EF']['Effect_tCO2e']} tCO2e。",
            f"- 固定 2025 EF：D-B = {activity_scope['Activity_Scope_Combined_Effect_At_2025_EF']['Effect_tCO2e']} tCO2e。",
            f"- 两年度实际试点结果总差异：D-A = {activity_scope['Observed_Cross_Year_Difference']['Effect_tCO2e']} tCO2e。",
            "",
            "这些变化可能同时来自采购 Activity、记录数量、产品组合和试点 Scope，不能称为纯 Activity 效应。",
            "",
            "## 对称分解",
            "",
            f"- Symmetric Factor Contribution：{symmetric['Symmetric_Factor_Contribution']['Effect_tCO2e']} tCO2e。",
            f"- Symmetric Activity/Scope Contribution：{symmetric['Symmetric_Activity_Scope_Contribution']['Effect_tCO2e']} tCO2e。",
            f"- 与 D-A 精确勾稽：{symmetric['Reconciliation_Status']}，差异 {symmetric['Reconciliation_Difference_kgCO2e']} kgCO2e。",
            "",
            "## 记录级高影响项",
            "",
            "|年度|Record_ID|Activity（kg）|因子影响（tCO2e）|",
            "|---:|---|---:|---:|",
        ]
    )
    for row in top_2024 + top_2025:
        lines.append(
            f"|{row['Year']}|{row['Record_ID']}|{row['Validated_Activity_kg']}|{row['Factor_Impact_tCO2e']}|"
        )
    lines.extend(
        [
            "",
            "## 可识别程度",
            "",
            "- FACTOR_EFFECT：IDENTIFIABLE",
            "- ACTIVITY_SCOPE_COMBINED_EFFECT：IDENTIFIABLE",
            "- PURE_ACTIVITY_EFFECT：NOT_SEPARATELY_IDENTIFIABLE",
            "- PRODUCT_STRUCTURE_EFFECT：NOT_SEPARATELY_IDENTIFIABLE",
            "- PURE_SCOPE_EFFECT：NOT_SEPARATELY_IDENTIFIABLE",
            "",
            "## 阶段结论",
            "",
            f"WP6-6 已用 WP6-5 验证通过的记录完成 A/B/C/D、记录级反事实和对称分解。2025 EF 相比 2024 变化 {factor['Relative_EF_Change_Percent']}%；若保持 2025 Activity/Scope 不变，该变化对应 {factor['Factor_Effect_At_2025_Activity']['Effect_tCO2e']} tCO2e 的历史模拟影响。",
            "",
        ]
    )
    return "\n".join(lines)


def _acceptance_report(summary: dict[str, Any]) -> str:
    scenarios = summary["scenario_abcd"]
    checks = [
        ("只使用 WP6-5 已验证且 Boundary Ready 的记录", summary["coverage"]["Excluded_Record_Count"] == 0),
        ("2024/2025 Scope 差异显式保留", bool(summary["scope_safety"]["scope_warning"])),
        ("未执行跨年 Record_ID 或尾号匹配", not summary["scope_safety"]["cross_year_record_matching_performed"] and not summary["scope_safety"]["record_id_suffix_matching_performed"]),
        ("A/B/C/D 四情景完整", set(scenarios) == {"A", "B", "C", "D"}),
        ("2024 固定 Activity 因子效应为 B-A", summary["factor_effect"]["Factor_Effect_At_2024_Activity"]["Effect_Name"] == "B_MINUS_A"),
        ("2025 固定 Activity 因子效应为 D-C", summary["factor_effect"]["Factor_Effect_At_2025_Activity"]["Effect_Name"] == "D_MINUS_C"),
        ("总差异命名为 Observed Cross-Year Difference", summary["activity_scope_effect"]["Observed_Cross_Year_Difference"]["Effect_Name"] == "D_MINUS_A"),
        ("Activity/Scope Combined Effect 明确标识", "Activity_Scope_Combined_Effect_At_2024_EF" in summary["activity_scope_effect"]),
        ("未无证据拆分纯 Activity/Structure/Scope", all(summary["identifiability"][key] == "NOT_SEPARATELY_IDENTIFIABLE" for key in ("PURE_ACTIVITY_EFFECT", "PRODUCT_STRUCTURE_EFFECT", "PURE_SCOPE_EFFECT"))),
        ("对称分解精确勾稽", summary["symmetric_decomposition"]["Reconciliation_Status"] == "PASS_EXACT"),
        ("2024 记录级反事实完整", len(summary["record_counterfactuals"]["2024"]) == summary["coverage"]["2024_Eligible"]),
        ("2025 记录级反事实完整", len(summary["record_counterfactuals"]["2025"]) == summary["coverage"]["2025_Eligible"]),
        ("EF 审计确认年度单一因子", all(summary["ef_audit"][year]["EF_Unique_Count"] == 1 for year in ("2024", "2025"))),
        ("历史模拟属性保留", summary["simulation_flag"] and not summary["production_eligible"]),
        ("输入哈希前后不变", all(summary["protected_inputs_unchanged"].values())),
        ("未提前执行 WP6-7", not summary["wp6_7_execution_performed"]),
    ]
    rows = [
        f"|{index}|{label}|{'PASS' if passed else 'FAIL'}|"
        for index, (label, passed) in enumerate(checks, start=1)
    ]
    status = "PASS" if all(passed for _, passed in checks) else "BLOCKED"
    return "\n".join(
        [
            "# WP6-6 验收报告 V1.0",
            "",
            f"> 验收日期：{summary['completed_at'][:10]}",
            f"> 最终状态：{status}",
            f"> 正式 Run：`{summary['run_id']}`",
            "",
            "## 验收结论",
            "",
            f"WP6-6 四情景、固定 Activity/Scope 因子效应、Activity/Scope 综合效应、记录级反事实与对称分解均完成；共纳入 {summary['coverage']['Eligible_Record_Count']} 条，排除 {summary['coverage']['Excluded_Record_Count']} 条。",
            "",
            "## 正式验收项",
            "",
            "|序号|验收项|结果|",
            "|---:|---|---|",
            *rows,
            "",
        ]
    )


def _handoff(summary: dict[str, Any]) -> str:
    factor = summary["factor_effect"]
    return "\n".join(
        [
            "# WP6-6 交接摘要",
            "",
            f"> WP6-6 状态：{summary['status']}",
            f"> 正式 Run：`{summary['run_id']}`",
            "> WP6-7 状态：NOT STARTED",
            "",
            "## 已完成",
            "",
            f"- 2024：{summary['coverage']['2024_Eligible']} 条记录级因子反事实",
            f"- 2025：{summary['coverage']['2025_Eligible']} 条记录级因子反事实",
            "- A/B/C/D 四情景与对称分解精确勾稽",
            f"- EF 相对变化：{factor['Relative_EF_Change_Percent']}%",
            "",
            "## WP6-7 可用接口",
            "",
            "`2024_factor_counterfactual.csv`、`2025_factor_counterfactual.csv`、`scenario_abcd_summary.json`、`factor_effect_summary.json`、`activity_scope_effect_summary.json`、`symmetric_decomposition.json` 与 `wp6_6_analysis_summary.json`。",
            "",
            "## 保留限制",
            "",
            "历史因子仅用于情景模拟，Production_Eligible=FALSE。跨年度未做记录匹配；Activity、产品结构与 Scope 仍不可分别识别。",
            "",
        ]
    )


def _write_workbook(path: Path, summary: dict[str, Any]) -> None:
    workbook = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    def add_sheet(name: str, rows: list[dict[str, Any]]) -> None:
        sheet = workbook.create_sheet(name) if workbook.sheetnames != ["Sheet"] else workbook.active
        sheet.title = name
        if not rows:
            sheet.append(["No data"])
            return
        fields = list(rows[0])
        sheet.append(fields)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in rows:
            sheet.append([row.get(field, "") for field in fields])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
            sheet.column_dimensions[column[0].column_letter].width = width

    add_sheet("Scenarios", list(summary["scenario_abcd"].values()))
    add_sheet("2024 Counterfactual", summary["record_counterfactuals"]["2024"])
    add_sheet("2025 Counterfactual", summary["record_counterfactuals"]["2025"])
    effect_rows = [
        {"Metric": "EF_2024", "Value": summary["factor_effect"]["EF_2024"]},
        {"Metric": "EF_2025", "Value": summary["factor_effect"]["EF_2025"]},
        {"Metric": "Relative_EF_Change_Percent", "Value": summary["factor_effect"]["Relative_EF_Change_Percent"]},
        {"Metric": "Factor_Effect_At_2024_Activity_tCO2e", "Value": summary["factor_effect"]["Factor_Effect_At_2024_Activity"]["Effect_tCO2e"]},
        {"Metric": "Factor_Effect_At_2025_Activity_tCO2e", "Value": summary["factor_effect"]["Factor_Effect_At_2025_Activity"]["Effect_tCO2e"]},
        {"Metric": "Observed_Cross_Year_Difference_tCO2e", "Value": summary["activity_scope_effect"]["Observed_Cross_Year_Difference"]["Effect_tCO2e"]},
        {"Metric": "Symmetric_Factor_Contribution_tCO2e", "Value": summary["symmetric_decomposition"]["Symmetric_Factor_Contribution"]["Effect_tCO2e"]},
        {"Metric": "Symmetric_Activity_Scope_Contribution_tCO2e", "Value": summary["symmetric_decomposition"]["Symmetric_Activity_Scope_Contribution"]["Effect_tCO2e"]},
    ]
    add_sheet("Effects", effect_rows)
    add_sheet("EF Audit", list(summary["ef_audit"].values()))
    workbook.save(path)


def run_wp6_6_analysis(
    *,
    wp6_5_run_dir: Path,
    output_root: Path,
    documentation_root: Path | None = None,
    run_id: str | None = None,
) -> dict[str, Any]:
    """Run the formal WP6-6 analysis from a completed WP6-5 evidence directory."""

    source = wp6_5_run_dir.expanduser().resolve()
    input_paths = {
        "wp6_5_summary": source / "independent_validation_summary.json",
        "2024_validation": source / "2024_independent_validation.csv",
        "2025_validation": source / "2025_independent_validation.csv",
    }
    missing = [str(path) for path in input_paths.values() if not path.is_file()]
    if missing:
        raise WP66AnalysisError(f"WP6-5 formal inputs are missing: {missing}")
    source_summary = json.loads(
        input_paths["wp6_5_summary"].read_text(encoding="utf-8")
    )
    if source_summary.get("status") != "PASS":
        raise WP66AnalysisError("WP6-5 formal summary status is not PASS")
    before_hashes = {name: _sha256(path) for name, path in input_paths.items()}
    analysis = build_scenario_analysis(
        _read_csv(input_paths["2024_validation"]),
        _read_csv(input_paths["2025_validation"]),
    )
    after_hashes = {name: _sha256(path) for name, path in input_paths.items()}
    hash_checks = {name: before_hashes[name] == after_hashes[name] for name in input_paths}
    if not all(hash_checks.values()):
        raise WP66AnalysisError("WP6-5 protected inputs changed during analysis")
    current_run_id = run_id or (
        "WP6-6-"
        + datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%fZ")
        + "-"
        + before_hashes["2024_validation"][:4]
        + before_hashes["2025_validation"][:4]
    )
    output_dir = output_root.expanduser().resolve() / current_run_id
    if output_dir.exists():
        raise WP66AnalysisError(f"formal run already exists: {output_dir}")
    output_dir.mkdir(parents=True)
    summary = {
        "schema_version": "WP6_6_ANALYSIS_SUMMARY_V1",
        "stage": "WP6-6",
        "status": analysis["status"],
        "run_id": current_run_id,
        "completed_at": datetime.now(timezone.utc).isoformat(),
        "source_wp6_5_run_id": source_summary["run_id"],
        "source_wp6_5_run_directory": str(source),
        **analysis,
        "protected_input_hashes_before": before_hashes,
        "protected_input_hashes_after": after_hashes,
        "protected_inputs_unchanged": hash_checks,
        "raw_excel_rescanned": False,
        "raw_data_modified": False,
        "frozen_evidence_modified": False,
        "historical_reproduction": True,
        "simulation_flag": True,
        "production_eligible": False,
        "streamlit_recalculates": False,
        "wp6_7_execution_performed": False,
    }

    _write_json(output_dir / "scenario_abcd_summary.json", summary["scenario_abcd"])
    _write_csv(
        output_dir / "2024_factor_counterfactual.csv",
        summary["record_counterfactuals"]["2024"],
        COUNTERFACTUAL_FIELDS,
    )
    _write_csv(
        output_dir / "2025_factor_counterfactual.csv",
        summary["record_counterfactuals"]["2025"],
        COUNTERFACTUAL_FIELDS,
    )
    _write_json(output_dir / "factor_effect_summary.json", summary["factor_effect"])
    _write_json(
        output_dir / "activity_scope_effect_summary.json",
        summary["activity_scope_effect"],
    )
    _write_json(
        output_dir / "symmetric_decomposition.json",
        summary["symmetric_decomposition"],
    )
    if summary["excluded_records"]:
        _write_csv(
            output_dir / "excluded_records.csv",
            summary["excluded_records"],
            ["Year", "Record_ID", "Excluded_From_Scenario", "Reason_Codes"],
        )
    _write_json(output_dir / "wp6_6_analysis_summary.json", summary)
    _write_workbook(output_dir / "scenario_analysis.xlsx", summary)
    documents = {
        "WP6-6_跨年度因子影响分析报告.md": _report(summary),
        "WP6-6_验收报告_V1.0.md": _acceptance_report(summary),
        "WP6-6_交接摘要.md": _handoff(summary),
    }
    for name, content in documents.items():
        (output_dir / name).write_text(content, encoding="utf-8")
    if documentation_root is not None:
        documentation = documentation_root.expanduser().resolve()
        documentation.mkdir(parents=True, exist_ok=True)
        for name, content in documents.items():
            (documentation / name).write_text(content, encoding="utf-8")
    return {
        "stage": "WP6-6",
        "status": summary["status"],
        "run_id": current_run_id,
        "output_directory": str(output_dir),
        "eligible_record_count": summary["coverage"]["Eligible_Record_Count"],
        "excluded_record_count": summary["coverage"]["Excluded_Record_Count"],
        "reconciliation_status": summary["symmetric_decomposition"]["Reconciliation_Status"],
        "wp6_7_execution_performed": False,
    }
