"""Read-only WP2 mapping adapters and public inline mapping catalogs."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from carbon_excel_pipeline.cleaning.raw_cleaner import clean_text


def _sheet_records(
    workbook_path: Path,
    sheet_name: str,
    *,
    required_header: str,
) -> list[dict[str, Any]]:
    workbook = load_workbook(
        workbook_path, read_only=True, data_only=True, keep_links=False
    )
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Required worksheet is missing: {sheet_name}")
        rows = list(workbook[sheet_name].iter_rows(values_only=True))
    finally:
        workbook.close()
    header_index = next(
        (
            index
            for index, row in enumerate(rows[:10])
            if required_header in row
        ),
        None,
    )
    if header_index is None:
        raise ValueError(
            f"Cannot find header {required_header!r} in worksheet {sheet_name!r}."
        )
    headers = [str(value) if value is not None else "" for value in rows[header_index]]
    return [
        dict(zip(headers, row))
        for row in rows[header_index + 1 :]
        if any(value is not None for value in row)
    ]


def load_frozen_id_map(
    baseline_path: Path,
    *,
    sheet_name: str,
) -> dict[tuple[str, str, int], str]:
    records = _sheet_records(
        baseline_path, sheet_name, required_header="Record_ID"
    )
    result: dict[tuple[str, str, int], str] = {}
    seen_ids: set[str] = set()
    for record in records:
        key = (
            str(record["Source_File"]),
            str(record["Source_Sheet"]),
            int(record["Source_Row"]),
        )
        record_id = str(record["Record_ID"])
        if key in result:
            raise ValueError(f"Duplicate frozen source key: {key!r}")
        if record_id in seen_ids:
            raise ValueError(f"Duplicate frozen Record_ID: {record_id}")
        result[key] = record_id
        seen_ids.add(record_id)
    return result


def _supplier_key(value: Any) -> str:
    return clean_text(value).casefold()


@dataclass(frozen=True, slots=True)
class MappingCatalog:
    suppliers: dict[str, dict[str, Any]]
    customers_by_id: dict[str, dict[str, Any]]
    project_cells: dict[tuple[int, str, str], dict[str, Any]]


def load_private_mapping_catalog(
    workbook_path: Path,
    *,
    mapping_config: dict[str, Any],
) -> MappingCatalog:
    settings = mapping_config["mapping_workbook"]
    active = settings["active_status"]
    supplier_rows = _sheet_records(
        workbook_path,
        settings["supplier_sheet"],
        required_header="Mapping_ID",
    )
    customer_rows = _sheet_records(
        workbook_path,
        settings["customer_sheet"],
        required_header="Mapping_ID",
    )
    project_rows = _sheet_records(
        workbook_path,
        settings["project_cell_sheet"],
        required_header="Mapping_ID",
    )
    suppliers: dict[str, dict[str, Any]] = {}
    for row in supplier_rows:
        if row.get("Record_Status") != active:
            continue
        key = _supplier_key(row["Normalized_Raw_Value"])
        if key in suppliers:
            raise ValueError(f"Duplicate active supplier mapping: {key}")
        suppliers[key] = row

    customers_by_id: dict[str, dict[str, Any]] = {}
    for row in customer_rows:
        if row.get("Record_Status") != active:
            continue
        customer_id = str(row["Customer_ID"])
        customers_by_id.setdefault(customer_id, row)

    project_cells: dict[tuple[int, str, str], dict[str, Any]] = {}
    for row in project_rows:
        if row.get("Record_Status") != active:
            continue
        key = (
            int(row["Year"]),
            str(row["Supplier_ID"]),
            str(row["Product_Description_Key"]),
        )
        if key in project_cells:
            raise ValueError("Duplicate active project/cell exact mapping.")
        project_cells[key] = row
    return MappingCatalog(suppliers, customers_by_id, project_cells)


def load_inline_mapping_catalog(mapping_config: dict[str, Any]) -> MappingCatalog:
    inline = mapping_config.get("inline_mappings", {})
    suppliers = {
        _supplier_key(row["Raw_Supplier_Value"]): dict(row)
        for row in inline.get("suppliers", [])
        if row.get("Record_Status") == "ACTIVE"
    }
    customers_by_id: dict[str, dict[str, Any]] = {}
    for row in inline.get("customers", []):
        if row.get("Record_Status") == "ACTIVE":
            customers_by_id.setdefault(str(row["Customer_ID"]), dict(row))
    project_cells = {
        (int(row["Year"]), str(row["Supplier_ID"]), str(row["Product_Description_Key"])): dict(row)
        for row in inline.get("project_cells", [])
        if row.get("Record_Status") == "ACTIVE"
    }
    return MappingCatalog(suppliers, customers_by_id, project_cells)


def find_supplier(catalog: MappingCatalog, alias: Any) -> dict[str, Any] | None:
    return catalog.suppliers.get(_supplier_key(alias))
