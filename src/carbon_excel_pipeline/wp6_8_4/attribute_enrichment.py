"""Deterministic attribute enrichment. Never overwrites Activity / EF / Emission."""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from carbon_excel_pipeline.io.semantic_registry import SemanticFieldRegistry


PROTECTED_FIELDS = {
    "Activity_Data_kg",
    "Original_Activity_Value",
    "Original_Activity_Unit",
    "EF_Value",
    "EF_Unit",
    "EF_Source",
    "Historical_GHG_Value",
    "Historical_GHG_Unit",
    "Emission_kgCO2e",
    "Emission_tCO2e",
    "Display_Emission_kgCO2e",
}
FILL_FIELDS = {
    "Chemistry": ("Chemistry",),
    "Supplier": ("Supplier", "Supplier_Name"),
    "Project": ("Project", "Project_Code"),
    "Model": ("Model", "Cell_Model"),
    "Customer": ("Customer",),
}


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _empty(value: Any) -> bool:
    text = _text(value)
    return text == "" or text.upper() in {"NONE", "UNKNOWN", "NA", "N/A"}


def _sheet_scope(sheet_name: str) -> tuple[str, str]:
    text = _text(sheet_name)
    year = "2025" if "2025" in text else ("2024" if "2024" in text else "")
    if "二部" in text or "事业二部" in text:
        unit = "合成二部"
    elif "一部" in text or "事业一部" in text:
        unit = "合成一部"
    else:
        unit = ""
    return year, unit


def _unit_token(value: Any) -> str:
    text = _text(value)
    if "二部" in text:
        return "合成二部"
    if "一部" in text:
        return "合成一部"
    return text


def load_attribute_rows(
    path: Path,
    alias_config: dict[str, Any],
) -> list[dict[str, Any]]:
    registry = SemanticFieldRegistry(alias_config)
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            attribute_year, attribute_unit = _sheet_scope(sheet.title)
            iterator = sheet.iter_rows(values_only=True)
            try:
                header_row = next(iterator)
            except StopIteration:
                continue
            mapping: dict[int, str] = {}
            for index, header in enumerate(header_row):
                if header is None or str(header).strip() == "":
                    continue
                field = registry.map_header(header, column_index=index + 1).semantic_field
                if field in FILL_FIELDS or field in {"Product_Description", "Purchase_Category"}:
                    mapping[index] = field
            if "Product_Description" not in mapping.values() and "Model" not in mapping.values():
                continue
            for source_row, values in enumerate(iterator, start=2):
                payload = {
                    "Attribute_Source_File": path.name,
                    "Attribute_Source_Sheet": sheet.title,
                    "Attribute_Source_Row": source_row,
                    "Attribute_Year": attribute_year,
                    "Attribute_Business_Unit": attribute_unit,
                }
                for index, field in mapping.items():
                    if index < len(values):
                        payload[field] = values[index]
                if any(_text(payload.get(field)) for field in FILL_FIELDS):
                    rows.append(payload)
    finally:
        workbook.close()
    return rows


def _index_rows(rows: list[dict[str, Any]], key: str) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        token = _text(row.get(key))
        if token:
            grouped[token].append(row)
    return grouped


def _scoped_candidates(
    candidates: list[dict[str, Any]], canonical: dict[str, Any]
) -> list[dict[str, Any]]:
    year = _text(canonical.get("Year"))
    unit = _unit_token(canonical.get("Business_Unit"))
    return [
        row
        for row in candidates
        if (not _text(row.get("Attribute_Year")) or _text(row.get("Attribute_Year")) == year)
        and (
            not _unit_token(row.get("Attribute_Business_Unit"))
            or _unit_token(row.get("Attribute_Business_Unit")) == unit
        )
    ]


def _attribute_signature(row: dict[str, Any]) -> tuple[str, ...]:
    return tuple(_text(row.get(field)) for field in FILL_FIELDS)


def enrich_canonical_with_attributes(
    canonical: list[dict[str, Any]],
    attribute_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    by_description = _index_rows(attribute_rows, "Product_Description")
    by_model = _index_rows(attribute_rows, "Model")
    by_project = _index_rows(attribute_rows, "Project")
    output: list[dict[str, Any]] = []
    matched = 0
    unmatched = 0
    ambiguous = 0
    for row in canonical:
        item = dict(row)
        method = ""
        status = "UNMATCHED"
        candidates: list[dict[str, Any]] = []
        description = _text(item.get("Product_Description"))
        model = _text(item.get("Model") or item.get("Cell_Model"))
        project = _text(item.get("Project") or item.get("Project_Code"))
        if description and description in by_description:
            candidates = _scoped_candidates(by_description[description], item)
            method = "Product_Description"
        elif model and model in by_model:
            candidates = _scoped_candidates(by_model[model], item)
            method = "Model"
        elif project and project in by_project:
            candidates = _scoped_candidates(by_project[project], item)
            method = "Project"
        if len(candidates) > 1 and len({_attribute_signature(row) for row in candidates}) == 1:
            candidates = [min(candidates, key=lambda row: int(row.get("Attribute_Source_Row") or 0))]
        if len(candidates) == 1:
            source = candidates[0]
            for semantic, targets in FILL_FIELDS.items():
                incoming = source.get(semantic)
                if _empty(incoming):
                    continue
                for target in targets:
                    if target in PROTECTED_FIELDS:
                        continue
                    if _empty(item.get(target)):
                        item[target] = incoming
            status = "MATCHED"
            matched += 1
            item["Attribute_Source_File"] = source.get("Attribute_Source_File")
            item["Attribute_Source_Sheet"] = source.get("Attribute_Source_Sheet")
            item["Attribute_Source_Row"] = source.get("Attribute_Source_Row")
        elif len(candidates) > 1:
            status = "AMBIGUOUS"
            method = method or "UNSAFE"
            ambiguous += 1
        else:
            unmatched += 1
        item["Attribute_Match_Method"] = method or "NONE"
        key_value = description if method == "Product_Description" else (model if method == "Model" else project)
        item["Attribute_Match_Key"] = f"{method}={key_value}" if method and key_value else "NONE"
        item["Attribute_Match_Status"] = status
        output.append(item)
    return output, {
        "matched": matched,
        "unmatched": unmatched,
        "ambiguous": ambiguous,
        "attribute_rows": len(attribute_rows),
        "message": (
            "部分补充属性无法可靠匹配，请确认关联字段。"
            if unmatched or ambiguous
            else "属性已按确定字段完成补充。"
        ),
    }
