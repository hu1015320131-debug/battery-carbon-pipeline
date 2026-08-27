"""Classify uploaded workbooks by mapped semantics, never by filename."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from carbon_excel_pipeline.io.semantic_registry import SemanticFieldRegistry
from carbon_excel_pipeline.wp6_8_4.input_set import (
    ROLE_ATTRIBUTE,
    ROLE_LEDGER,
    ROLE_PRIMARY,
    ROLE_UNKNOWN,
)
from carbon_excel_pipeline.wp6_8_5.ledger_reference import looks_like_historical_ledger


ACTIVITY_FIELDS = {
    "Reported_Activity_Value",
    "Quantity_PCS",
    "Unit_Weight",
    "Reported_Purchase_Quantity",
}
ATTRIBUTE_FIELDS = {
    "Chemistry",
    "Supplier",
    "Project",
    "Model",
    "Customer",
}


def _normalize(value: Any) -> str:
    return "" if value is None else " ".join(str(value).strip().lower().split())


def _registry_match(registry: SemanticFieldRegistry, header: str) -> str | None:
    result = registry.map_header(header, column_index=1)
    return result.semantic_field


def scan_mapped_fields(path: Path, alias_config: dict[str, Any], *, max_rows: int = 12) -> set[str]:
    registry = SemanticFieldRegistry(alias_config)
    workbook = load_workbook(path, read_only=True, data_only=True)
    mapped: set[str] = set()
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True):
                for value in row:
                    if value is None or str(value).strip() == "":
                        continue
                    field = _registry_match(registry, str(value).strip())
                    if field:
                        mapped.add(field)
    finally:
        workbook.close()
    return mapped


def classify_mapped_fields(mapped: set[str]) -> str:
    has_activity = bool(mapped & ACTIVITY_FIELDS)
    has_attribute = bool(mapped & ATTRIBUTE_FIELDS)
    if has_activity:
        return ROLE_PRIMARY
    if has_attribute:
        return ROLE_ATTRIBUTE
    return ROLE_UNKNOWN


def classify_workbook_role(path: Path, alias_config: dict[str, Any]) -> dict[str, Any]:
    mapped = scan_mapped_fields(path, alias_config)
    role = ROLE_LEDGER if looks_like_historical_ledger(path) else classify_mapped_fields(mapped)
    return {
        "path": str(path),
        "name": path.name,
        "role": role,
        "mapped_fields": sorted(mapped),
        "suggested": True,
    }


def reconcile_roles(files: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Keep user overrides. If several files look primary, keep the first as primary."""

    output = [dict(item) for item in files]
    primaries = [index for index, item in enumerate(output) if item.get("role") == ROLE_PRIMARY]
    if len(primaries) > 1:
        for index in primaries[1:]:
            output[index]["role"] = ROLE_ATTRIBUTE
            output[index]["role_note"] = "同一批只能有一个主核算文件，其余已改为属性补充。"
    if not primaries and output:
        output[0]["role"] = ROLE_PRIMARY
        output[0]["role_note"] = "未能自动区分时，请确认主核算文件。"
    return output
