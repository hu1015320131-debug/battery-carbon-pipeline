"""Recognize a historical ledger and extract controlled cell-category evidence."""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from carbon_excel_pipeline.io.excel_importer import sha256_file


ROLE_LEDGER = "历史清册/因子参考"


def _text(value: Any) -> str:
    return "" if value is None else " ".join(str(value).strip().split())


def _decimal(value: Any) -> Decimal | None:
    try:
        return Decimal(str(value)) if value not in (None, "") else None
    except (InvalidOperation, ValueError):
        return None


def _contains_sheet_semantics(sheet_names: list[str]) -> bool:
    joined = "|".join(sheet_names)
    markers = ("活动数据管理", "排放系数管理", "排放量计算")
    return sum(marker in joined for marker in markers) >= 2


def looks_like_historical_ledger(path: Path) -> bool:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return _contains_sheet_semantics(workbook.sheetnames)
    finally:
        workbook.close()


def _cell_row(worksheet) -> tuple[int, tuple[Any, ...]] | None:
    for row_number, values in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if any(_text(value) == "电芯" for value in values):
            return row_number, values
    return None


def _activity_evidence(workbook) -> dict[str, Any]:
    sheet = next((ws for ws in workbook.worksheets if "活动数据管理" in ws.title), None)
    found = _cell_row(sheet) if sheet is not None else None
    if not found:
        return {}
    row_number, values = found
    for index, value in enumerate(values):
        if _text(value) != "电芯":
            continue
        for offset in range(index + 1, min(len(values) - 1, index + 6)):
            amount = _decimal(values[offset])
            unit = _text(values[offset + 1])
            if amount is not None and unit.lower() == "kg":
                return {
                    "Historical_Activity_kg": format(amount, "f"),
                    "Activity_Source_Sheet": sheet.title,
                    "Activity_Source_Row": row_number,
                    "Activity_Source_Cell": sheet.cell(row_number, offset + 1).coordinate,
                }
    return {}


def _factor_evidence(workbook) -> dict[str, Any]:
    sheet = next((ws for ws in workbook.worksheets if "排放系数管理" in ws.title), None)
    found = _cell_row(sheet) if sheet is not None else None
    if not found:
        return {}
    row_number, values = found
    for index in range(len(values) - 1):
        factor = _decimal(values[index])
        unit = _text(values[index + 1]).lower().replace(" ", "")
        if factor is not None and "kgco2/kg" in unit:
            return {
        "EF_Value": format(factor.quantize(Decimal("0.000001")), "f"),
                "EF_Unit": "kgCO2e/kg",
                "Factor_Source_Sheet": sheet.title,
                "Factor_Source_Row": row_number,
                "Factor_Source_Cell": sheet.cell(row_number, index + 1).coordinate,
            }
    return {}


def _emission_evidence(workbook) -> dict[str, Any]:
    sheet = next((ws for ws in workbook.worksheets if "排放量计算" in ws.title), None)
    found = _cell_row(sheet) if sheet is not None else None
    if not found:
        return {}
    row_number, values = found
    header_labels: dict[int, str] = {}
    for column in range(1, sheet.max_column + 1):
        header_labels[column] = "|".join(
            _text(sheet.cell(row, column).value) for row in range(1, 5)
        )
    target_column = next(
        (
            column
            for column, label in header_labels.items()
            if "年总排放量" in label and "t CO2当量" in label
        ),
        None,
    )
    if target_column is None:
        return {}
    value = _decimal(values[target_column - 1])
    if value is None:
        return {}
    return {
        "Historical_Emission_tCO2e": format(value, "f"),
        "Emission_Source_Sheet": sheet.title,
        "Emission_Source_Row": row_number,
        "Emission_Source_Cell": sheet.cell(row_number, target_column).coordinate,
    }


def extract_cell_ledger_evidence(path: Path) -> dict[str, Any]:
    source = path.expanduser().resolve()
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        semantics_ready = _contains_sheet_semantics(workbook.sheetnames)
        evidence = {
            **_activity_evidence(workbook),
            **_factor_evidence(workbook),
            **_emission_evidence(workbook),
        }
    finally:
        workbook.close()
    factor = _decimal(evidence.get("EF_Value"))
    status = "PASS" if semantics_ready and factor is not None and factor > 0 else "WARNING"
    return {
        "schema_version": "WP6_8_5_LEDGER_EVIDENCE_V1",
        "status": status,
        "role": ROLE_LEDGER,
        "Category": "电芯",
        "Source_File": source.name,
        "Source_SHA256": sha256_file(source),
        "semantic_sheet_detection": semantics_ready,
        **evidence,
        "Factor_Usage": "HISTORICAL_SIMULATION",
        "Simulation_Flag": True,
        "Production_Eligible": False,
    }
