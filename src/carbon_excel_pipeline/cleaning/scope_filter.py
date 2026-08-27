"""Day 3 mapping confirmation, scope filtering, exclusion audit and raw cleaning."""

from __future__ import annotations

import csv
import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from carbon_excel_pipeline.cleaning.raw_cleaner import (
    clean_text,
    contains_marker,
    extract_cell_model_candidate,
    extract_project_prefix_candidate,
    map_unit_strict,
    parse_strict_positive_decimal,
)
from carbon_excel_pipeline.errors import PipelineUserError
from carbon_excel_pipeline.mapping.field_confirmation import confirm_sheet_mapping


RAW_CANDIDATE_FIELDS = [
    "Run_ID",
    "Source_File",
    "Source_Sheet",
    "Source_Row",
    "Scope_Status",
    "Scope_Rule_ID",
    "Purchase_Category_Raw",
    "Product_Description_Raw",
    "PCS_Raw",
    "Source_Unit_Raw",
    "Unit_Weight_Raw",
    "Annual_Activity_Formula_Raw",
    "Annual_Activity_Cached_Raw",
]
EXCLUSION_FIELDS = [
    "Run_ID",
    "Source_File",
    "Source_Sheet",
    "Source_Row",
    "Scope_Status",
    "Exclusion_Reason_Code",
    "Exclusion_Reason_CN",
    "Purchase_Category_Raw",
    "Product_Description_Raw",
]
CLEANED_CANDIDATE_FIELDS = [
    "Run_ID",
    "Source_File",
    "Source_Sheet",
    "Source_Row",
    "Scope_Status",
    "Scope_Rule_ID",
    "Purchase_Category_Raw",
    "Purchase_Category_Clean",
    "Product_Description_Raw",
    "Product_Description_Clean",
    "Supplier_Alias_Candidate",
    "Supplier_Match_Method",
    "Project_Model_Prefix_Candidate",
    "Cell_Model_Candidate",
    "PCS_Raw",
    "PCS_Clean",
    "PCS_Unit_Raw",
    "PCS_Unit_Clean",
    "Unit_Weight_Raw",
    "Unit_Weight_Clean",
    "Unit_Weight_Unit_Raw",
    "Unit_Weight_Unit_Clean",
    "Annual_Activity_Formula_Raw",
    "Annual_Activity_Cached_Raw",
    "Annual_Activity_Clean",
    "Annual_Activity_Unit_Raw",
    "Annual_Activity_Unit_Clean",
    "Cleaning_Status",
    "Cleaning_Issue_Codes",
]


def _load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def _write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _safe_raw(value: Any) -> Any:
    return "" if value is None else value


def _pipeline_error(
    *, code: str, message: str, location: str, value: Any, rule: str, suggestion: str
) -> PipelineUserError:
    return PipelineUserError(
        stage="SCOPE_AND_CLEANING",
        error_code=code,
        message_cn=message,
        source_location=location,
        original_value=value,
        rule=rule,
        impact="阻断整次Day 3运行",
        fix_suggestion=suggestion,
    )


def _scope_decision(
    *, category_clean: str, description_clean: str, config: dict[str, Any]
) -> tuple[bool, str, str]:
    if category_clean != config["target_purchase_category"]:
        return False, "CATEGORY_OUT_OF_SCOPE", "采购分类不属于正式试点范围"
    markers = config["supplier_markers"]
    if not any(contains_marker(description_clean, marker) for marker in markers):
        return False, "SUPPLIER_MARKER_NOT_FOUND", "物料描述未识别到试点供应商标记"
    return True, "", ""


def _clean_candidate(
    *,
    base: dict[str, Any],
    formula_value: Any,
    cached_value: Any,
    unit_config: dict[str, Any],
    scope_config: dict[str, Any],
) -> dict[str, Any]:
    category_clean = clean_text(base["Purchase_Category_Raw"])
    description_clean = clean_text(base["Product_Description_Raw"])
    pcs = parse_strict_positive_decimal(
        base["PCS_Raw"], field_code="PCS", integer_required=True
    )
    weight = parse_strict_positive_decimal(
        base["Unit_Weight_Raw"], field_code="UNIT_WEIGHT"
    )
    annual = parse_strict_positive_decimal(
        cached_value, field_code="ANNUAL_ACTIVITY"
    )
    pcs_unit = map_unit_strict(
        base["Source_Unit_Raw"],
        field_code="PCS",
        mappings=unit_config["fields"]["PCS"]["mappings"],
    )
    weight_unit_raw = scope_config["context_units"]["unit_weight_raw"]
    annual_unit_raw = scope_config["context_units"]["annual_activity_raw"]
    weight_unit = map_unit_strict(
        weight_unit_raw,
        field_code="UNIT_WEIGHT",
        mappings=unit_config["fields"]["UNIT_WEIGHT"]["mappings"],
    )
    annual_unit = map_unit_strict(
        annual_unit_raw,
        field_code="ANNUAL_ACTIVITY",
        mappings=unit_config["fields"]["ANNUAL_ACTIVITY"]["mappings"],
    )
    issue_codes = [
        item
        for item in (
            pcs.issue_code,
            weight.issue_code,
            annual.issue_code,
            pcs_unit.issue_code,
            weight_unit.issue_code,
            annual_unit.issue_code,
        )
        if item
    ]
    supplier_marker = next(
        (
            marker
            for marker in scope_config["supplier_markers"]
            if contains_marker(description_clean, marker)
        ),
        "",
    )
    return {
        **base,
        "Purchase_Category_Clean": category_clean,
        "Product_Description_Clean": description_clean,
        "Supplier_Alias_Candidate": supplier_marker.upper(),
        "Supplier_Match_Method": scope_config["supplier_match_method"],
        "Project_Model_Prefix_Candidate": extract_project_prefix_candidate(
            description_clean
        ),
        "Cell_Model_Candidate": extract_cell_model_candidate(description_clean),
        "PCS_Clean": pcs.cleaned_value,
        "PCS_Unit_Raw": base["Source_Unit_Raw"],
        "PCS_Unit_Clean": pcs_unit.cleaned_unit,
        "Unit_Weight_Clean": weight.cleaned_value,
        "Unit_Weight_Unit_Raw": weight_unit_raw,
        "Unit_Weight_Unit_Clean": weight_unit.cleaned_unit,
        "Annual_Activity_Formula_Raw": _safe_raw(formula_value),
        "Annual_Activity_Cached_Raw": _safe_raw(cached_value),
        "Annual_Activity_Clean": annual.cleaned_value,
        "Annual_Activity_Unit_Raw": annual_unit_raw,
        "Annual_Activity_Unit_Clean": annual_unit.cleaned_unit,
        "Cleaning_Status": "PASS" if not issue_codes else "ERROR",
        "Cleaning_Issue_Codes": ";".join(issue_codes),
    }


def run_day3_scope_and_cleaning(
    run_dir: Path,
    *,
    scope_config_path: Path,
    unit_config_path: Path,
    mapping_overrides: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    run_dir = run_dir.expanduser().resolve()
    import_dir = run_dir / "01_import"
    scope_dir = run_dir / "02_scope_filter"
    cleaning_dir = run_dir / "03_standardized"
    for path in (import_dir, scope_dir, cleaning_dir):
        if not path.is_dir():
            raise _pipeline_error(
                code="RUN_STAGE_DIRECTORY_MISSING",
                message="运行目录缺少Day 3所需阶段目录。",
                location=path.name,
                value=path.name,
                rule="Day 3必须接续通过Day 2的隔离运行目录。",
                suggestion="先用Day 2 inspect命令创建完整运行目录。",
            )

    scope_config = _load_json(scope_config_path)
    unit_config = _load_json(unit_config_path)
    mapping_reports = _load_json(import_dir / "field_mapping_preview.json")
    receipt = _load_json(import_dir / "file_receipt_report.json")
    confirmation = confirm_sheet_mapping(
        mapping_reports,
        target_sheet=scope_config["target_sheet"],
        required_targets=scope_config["required_targets"],
        mapping_overrides=mapping_overrides,
    )
    confirmation.update(
        {
            "config_id": scope_config["config_id"],
            "profile_id": scope_config["profile_id"],
        }
    )
    _write_json(scope_dir / "field_mapping_confirmation.json", confirmation)
    if confirmation["status"] != "CONFIRMED":
        raise _pipeline_error(
            code="FIELD_MAPPING_CONFIRMATION_REQUIRED",
            message="字段映射未达到自动确认条件。",
            location=scope_config["target_sheet"],
            value={"error_count": len(confirmation["errors"])},
            rule="六个Day 3目标字段必须全部唯一映射。",
            suggestion="查看field_mapping_confirmation.json并人工确认冲突字段。",
        )

    input_copy = run_dir / "00_input_copy" / receipt["copy_file_name"]
    if not input_copy.is_file():
        raise _pipeline_error(
            code="INPUT_COPY_MISSING",
            message="Day 2隔离输入副本不存在。",
            location=receipt["copy_file_name"],
            value=receipt["copy_file_name"],
            rule="Day 3只能处理Day 2已经校验的隔离副本。",
            suggestion="重新执行Day 2文件接收。",
        )

    formula_workbook = load_workbook(
        input_copy, read_only=False, data_only=False, keep_links=False
    )
    value_workbook = load_workbook(
        input_copy, read_only=False, data_only=True, keep_links=False
    )
    target_sheet = scope_config["target_sheet"]
    if target_sheet not in formula_workbook.sheetnames:
        formula_workbook.close()
        value_workbook.close()
        raise _pipeline_error(
            code="TARGET_SHEET_MISSING",
            message="工作簿中不存在配置的目标工作表。",
            location=target_sheet,
            value=target_sheet,
            rule="私有正式路线只处理配置指定的工作表。",
            suggestion="检查Profile中的target_sheet或重新选择工作簿。",
        )

    formula_sheet = formula_workbook[target_sheet]
    value_sheet = value_workbook[target_sheet]
    all_sheet_names = list(formula_workbook.sheetnames)
    columns = {
        item["target_field"]: int(item["column_index"])
        for item in confirmation["confirmed_fields"]
    }
    header_row = int(confirmation["header_row"])
    source_data_records = formula_sheet.max_row - header_row
    raw_candidates: list[dict[str, Any]] = []
    cleaned_candidates: list[dict[str, Any]] = []
    exclusions: list[dict[str, Any]] = []
    try:
        for source_row in range(header_row + 1, formula_sheet.max_row + 1):
            category_raw = formula_sheet.cell(
                source_row, columns["Purchase_Category"]
            ).value
            description_raw = formula_sheet.cell(
                source_row, columns["Product_Description"]
            ).value
            category_clean = clean_text(category_raw)
            description_clean = clean_text(description_raw)
            included, reason_code, reason_cn = _scope_decision(
                category_clean=category_clean,
                description_clean=description_clean,
                config=scope_config,
            )
            common = {
                "Run_ID": run_dir.name,
                "Source_File": receipt["source_file_name"],
                "Source_Sheet": target_sheet,
                "Source_Row": source_row,
                "Purchase_Category_Raw": _safe_raw(category_raw),
                "Product_Description_Raw": _safe_raw(description_raw),
            }
            if not included:
                exclusions.append(
                    {
                        **common,
                        "Scope_Status": "EXCLUDED",
                        "Exclusion_Reason_Code": reason_code,
                        "Exclusion_Reason_CN": reason_cn,
                    }
                )
                continue

            formula_value = formula_sheet.cell(
                source_row, columns["Annual_Purchase_g_per_year"]
            ).value
            cached_value = value_sheet.cell(
                source_row, columns["Annual_Purchase_g_per_year"]
            ).value
            candidate = {
                **common,
                "Scope_Status": "IN_SCOPE",
                "Scope_Rule_ID": scope_config["config_id"],
                "PCS_Raw": _safe_raw(
                    formula_sheet.cell(source_row, columns["PCS"]).value
                ),
                "Source_Unit_Raw": _safe_raw(
                    formula_sheet.cell(source_row, columns["Source_Unit"]).value
                ),
                "Unit_Weight_Raw": _safe_raw(
                    formula_sheet.cell(
                        source_row, columns["Unit_Weight_g_per_PCS"]
                    ).value
                ),
                "Annual_Activity_Formula_Raw": _safe_raw(formula_value),
                "Annual_Activity_Cached_Raw": _safe_raw(cached_value),
            }
            raw_candidates.append(candidate)
            cleaned_candidates.append(
                _clean_candidate(
                    base=candidate,
                    formula_value=formula_value,
                    cached_value=cached_value,
                    unit_config=unit_config,
                    scope_config=scope_config,
                )
            )
    finally:
        formula_workbook.close()
        value_workbook.close()

    input_records = len(raw_candidates) + len(exclusions)
    exclusion_counts = Counter(
        item["Exclusion_Reason_Code"] for item in exclusions
    )
    cleaning_counts = Counter(
        item["Cleaning_Status"] for item in cleaned_candidates
    )
    issue_counts = Counter(
        code
        for item in cleaned_candidates
        for code in item["Cleaning_Issue_Codes"].split(";")
        if code
    )
    validation_errors: list[dict[str, Any]] = []
    expected_input = scope_config.get("expected_input_records")
    expected_candidates = scope_config.get("expected_candidate_records")
    if expected_input is not None and input_records != int(expected_input):
        validation_errors.append(
            {
                "error_code": "INPUT_RECORD_COUNT_MISMATCH",
                "expected": int(expected_input),
                "actual": input_records,
            }
        )
    if expected_candidates is not None and len(raw_candidates) != int(
        expected_candidates
    ):
        validation_errors.append(
            {
                "error_code": "CANDIDATE_RECORD_COUNT_MISMATCH",
                "expected": int(expected_candidates),
                "actual": len(raw_candidates),
            }
        )
    if input_records != source_data_records:
        validation_errors.append(
            {
                "error_code": "DESTINATION_BALANCE_FAILED",
                "expected": source_data_records,
                "actual": input_records,
            }
        )

    _write_csv(scope_dir / "candidate_records.csv", raw_candidates, RAW_CANDIDATE_FIELDS)
    _write_csv(
        scope_dir / "excluded_records_audit.csv", exclusions, EXCLUSION_FIELDS
    )
    _write_csv(
        cleaning_dir / "day3_cleaned_candidates.csv",
        cleaned_candidates,
        CLEANED_CANDIDATE_FIELDS,
    )
    _write_csv(
        cleaning_dir / "cleaning_issue_records.csv",
        [
            {
                "Run_ID": item["Run_ID"],
                "Source_Sheet": item["Source_Sheet"],
                "Source_Row": item["Source_Row"],
                "Cleaning_Status": item["Cleaning_Status"],
                "Cleaning_Issue_Codes": item["Cleaning_Issue_Codes"],
            }
            for item in cleaned_candidates
            if item["Cleaning_Status"] != "PASS"
        ],
        [
            "Run_ID",
            "Source_Sheet",
            "Source_Row",
            "Cleaning_Status",
            "Cleaning_Issue_Codes",
        ],
    )
    sheet_decisions = [
        {
            "sheet_name": sheet_name,
            "decision": "ROW_LEVEL_SCOPE_FILTER"
            if sheet_name == target_sheet
            else "SHEET_OUT_OF_SCOPE",
            "reason_code": "TARGET_SHEET"
            if sheet_name == target_sheet
            else "NON_TARGET_SHEET",
        }
        for sheet_name in all_sheet_names
    ]
    _write_json(scope_dir / "sheet_scope_decisions.json", sheet_decisions)
    summary = {
        "run_id": run_dir.name,
        "status": "PASS" if not validation_errors else "FAIL",
        "profile_id": scope_config["profile_id"],
        "scope_config_id": scope_config["config_id"],
        "target_sheet": target_sheet,
        "field_mapping_status": confirmation["status"],
        "input_records": input_records,
        "candidate_records": len(raw_candidates),
        "excluded_records": len(exclusions),
        "destination_balance": input_records == source_data_records,
        "exclusion_reason_counts": dict(sorted(exclusion_counts.items())),
        "cleaning_status_counts": dict(sorted(cleaning_counts.items())),
        "cleaning_issue_counts": dict(sorted(issue_counts.items())),
        "source_sha256": receipt["source_sha256"],
        "source_unchanged_in_day2": receipt["source_unchanged"],
        "validation_errors": validation_errors,
        "outputs": {
            "mapping_confirmation": "02_scope_filter/field_mapping_confirmation.json",
            "raw_candidates": "02_scope_filter/candidate_records.csv",
            "exclusion_audit": "02_scope_filter/excluded_records_audit.csv",
            "cleaned_candidates": "03_standardized/day3_cleaned_candidates.csv",
            "cleaning_issues": "03_standardized/cleaning_issue_records.csv",
        },
    }
    _write_json(scope_dir / "day3_scope_summary.json", summary)
    _write_json(
        cleaning_dir / "day3_cleaning_summary.json",
        {
            "run_id": run_dir.name,
            "status": summary["status"],
            "candidate_records": len(cleaned_candidates),
            "cleaning_status_counts": summary["cleaning_status_counts"],
            "cleaning_issue_counts": summary["cleaning_issue_counts"],
            "raw_and_clean_values_preserved": True,
            "decimal_semantics": True,
            "unit_comparison_case_sensitive": True,
            "unit_trim_before_compare": False,
        },
    )
    return {**summary, "run_directory": str(run_dir)}
