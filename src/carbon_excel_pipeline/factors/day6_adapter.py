"""Day 6 strict eight-field receipt and auditable WP5-D1 adaptation."""

from __future__ import annotations

import csv
import hashlib
import json
import re
import shutil
import uuid
from collections import Counter
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from carbon_excel_pipeline.errors import PipelineUserError


RESULT_NAMESPACE = uuid.UUID("4165f964-36d8-4b08-9903-f81358ca4268")
ADAPTATION_AUDIT_FIELDS = [
    "Record_ID",
    "Result_ID",
    "Input_Row",
    "Receipt_Source_File",
    "Receipt_Source_SHA256",
    "Source_Field_Count",
    "Target_Field_Count",
    "Enrichment_Source",
    "Adaptation_Status",
    "Issue_Code",
]


def _error(code: str, message: str, location: str, value: Any, rule: str) -> PipelineUserError:
    return PipelineUserError(
        stage="DAY6_FACTOR_ADAPTER",
        error_code=code,
        message_cn=message,
        source_location=location,
        original_value=value,
        rule=rule,
        impact="阻断因子适配、精确匹配和后续排放计算",
        fix_suggestion="按Day 6八字段模板修正输入后重新执行。",
    )


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest().upper()


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def _read_xlsx(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    header_index = next(
        (index for index, row in enumerate(rows[:10]) if "Record_ID" in row), None
    )
    if header_index is None:
        raise _error(
            "FACTOR_HEADER_NOT_FOUND",
            "前10行未找到Record_ID表头。",
            path.name,
            "Record_ID",
            "CSV或首个工作表前10行必须包含完整八字段表头。",
        )
    fields = [str(value) for value in rows[header_index] if value is not None]
    records = [
        {field: "" if value is None else value for field, value in zip(fields, row)}
        for row in rows[header_index + 1 :]
        if any(value is not None for value in row)
    ]
    return fields, records


def read_factor_input(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix == ".xlsx":
        return _read_xlsx(path)
    raise _error(
        "FACTOR_FILE_TYPE_UNSUPPORTED",
        "因子输入仅支持CSV或XLSX。",
        path.name,
        suffix,
        "允许的扩展名为.csv和.xlsx。",
    )


def receive_factor_input(
    source: Path,
    *,
    stage_dir: Path,
    contract: dict[str, Any],
) -> tuple[Path, list[dict[str, str]], dict[str, str]]:
    source = source.expanduser().resolve()
    if not source.is_file():
        raise _error(
            "FACTOR_INPUT_MISSING", "因子输入文件不存在。", str(source), "MISSING", "输入必须是可读文件。"
        )
    fields, raw_records = read_factor_input(source)
    if fields != contract["fields"]:
        raise _error(
            "FACTOR_8_FIELD_SCHEMA_MISMATCH",
            "因子输入字段名称或顺序与八字段契约不一致。",
            source.name,
            fields,
            "必须严格使用8个字段且顺序不变。",
        )
    digest = sha256(source)
    raw_dir = stage_dir / "raw_upload"
    raw_dir.mkdir(parents=True, exist_ok=True)
    receipt_path = raw_dir / f"{digest[:12]}_{source.name}"
    if source != receipt_path.resolve():
        if receipt_path.exists() and sha256(receipt_path) != digest:
            raise _error(
                "RAW_RECEIPT_COLLISION",
                "原始接收文件名冲突且哈希不同。",
                str(receipt_path),
                digest,
                "原始上传副本不可覆盖。",
            )
        if not receipt_path.exists():
            shutil.copy2(source, receipt_path)
    received_at = datetime.now(timezone.utc).isoformat()
    records: list[dict[str, str]] = []
    for index, raw in enumerate(raw_records, start=2):
        record = {field: "" if raw.get(field) is None else str(raw.get(field)) for field in fields}
        _validate_factor_row(record, index=index, contract=contract)
        records.append(record)
    metadata = {
        "source_file": source.name,
        "receipt_file": receipt_path.name,
        "receipt_sha256": digest,
        "received_at": received_at,
        "raw_payload_path": str(receipt_path.relative_to(stage_dir.parent)).replace("\\", "/"),
    }
    return receipt_path, records, metadata


def _validate_factor_row(record: dict[str, str], *, index: int, contract: dict[str, Any]) -> None:
    record_id = record["Record_ID"]
    if not record_id or record_id != record_id.strip():
        raise _error(
            "RECORD_ID_INVALID", "Record_ID为空或带前后空格。", f"row {index}/Record_ID", record_id, "Record_ID必须非空且原样匹配。"
        )
    value = record["EF_Value"]
    if not value or value != value.strip():
        raise _error(
            "EF_VALUE_INVALID", "EF_Value为空或带前后空格。", f"row {index}/EF_Value", value, "EF_Value必须为严格正Decimal。"
        )
    try:
        decimal_value = Decimal(value)
    except InvalidOperation as exc:
        raise _error(
            "EF_VALUE_INVALID", "EF_Value不是有效Decimal。", f"row {index}/EF_Value", value, "EF_Value必须为严格正Decimal。"
        ) from exc
    if not decimal_value.is_finite() or decimal_value <= 0:
        raise _error(
            "EF_VALUE_NOT_POSITIVE", "EF_Value必须严格大于0。", f"row {index}/EF_Value", value, "EF_Value必须为严格正Decimal。"
        )
    if record["EF_Unit"] not in contract["allowed_ef_units"]:
        raise _error(
            "EF_UNIT_INVALID_EXACT", "EF_Unit未通过精确匹配。", f"row {index}/EF_Unit", record["EF_Unit"], "单位不裁剪空格、不转换大小写。"
        )
    if record["Result_Type"] not in contract["allowed_result_types"]:
        raise _error(
            "RESULT_TYPE_INVALID", "Result_Type不受支持。", f"row {index}/Result_Type", record["Result_Type"], "仅允许EMISSION_FACTOR。"
        )
    if not record["Source_Name"] or not record["Source_Version"]:
        raise _error(
            "FACTOR_SOURCE_INCOMPLETE", "因子来源名称或版本为空。", f"row {index}", "", "Source_Name和Source_Version必须非空。"
        )
    for field in ("Simulation_Flag", "Production_Eligible"):
        if record[field] not in contract["boolean_values"]:
            raise _error(
                "BOOLEAN_VALUE_INVALID", f"{field}不是受控布尔值。", f"row {index}/{field}", record[field], "仅允许大写TRUE或FALSE。"
            )


def generate_historical_eight_field_input(
    activity_records: list[dict[str, str]],
    *,
    config: dict[str, Any],
    output_path: Path,
    fields: list[str],
) -> Path:
    rows = [
        {
            "Record_ID": record["Record_ID"],
            "EF_Value": config["ef_value"],
            "EF_Unit": config["ef_unit"],
            "Result_Type": config["result_type"],
            "Source_Name": config["source_name"],
            "Source_Version": config["source_version"],
            "Simulation_Flag": "TRUE",
            "Production_Eligible": "FALSE",
        }
        for record in activity_records
    ]
    write_csv(output_path, rows, fields)
    return output_path


def candidate_anomalies(
    activity_records: list[dict[str, Any]], factor_records: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    activity_counts = Counter(str(row.get("Record_ID", "")) for row in activity_records)
    factor_counts = Counter(str(row.get("Record_ID", "")) for row in factor_records)
    anomalies: list[dict[str, Any]] = []
    for record_id in sorted(set(activity_counts) | set(factor_counts)):
        activity_count = activity_counts[record_id]
        factor_count = factor_counts[record_id]
        if activity_count == factor_count == 1:
            continue
        if activity_count == 1 and factor_count == 0:
            issue = "UNMATCHED_FACTOR_RESULT"
        elif activity_count == 0 and factor_count >= 1:
            issue = "ORPHAN_FACTOR_RESULT"
        elif activity_count > 1:
            issue = "DUPLICATE_ACTIVITY_RECORD_ID"
        else:
            issue = "MULTIPLE_FACTOR_CANDIDATES"
        anomalies.append(
            {
                "Record_ID": record_id,
                "Activity_Candidate_Count": activity_count,
                "Result_Candidate_Count": factor_count,
                "Issue_Code": issue,
                "Match_Status": "MANUAL_REVIEW_REQUIRED",
                "Fallback_Attempted": "FALSE",
            }
        )
    return anomalies


def _relationship_id(record_id: str) -> str:
    matched = re.fullmatch(r"2025-DY2-SYNA-DX([0-9]{6})", record_id)
    if matched:
        return f"REL-2025-{matched.group(1)}"
    token = uuid.uuid5(RESULT_NAMESPACE, f"relationship|{record_id}").hex[:12].upper()
    return f"REL-DEMO-{token}"


def adapt_to_d1(
    factor_records: list[dict[str, str]],
    activity_records: list[dict[str, str]],
    *,
    metadata: dict[str, str],
    factor_config: dict[str, Any],
    target_fields: list[str],
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    activity_by_id = {row["Record_ID"]: row for row in activity_records}
    occurrences: Counter[str] = Counter()
    output: list[dict[str, str]] = []
    audit: list[dict[str, str]] = []
    for input_index, factor in enumerate(factor_records, start=2):
        record_id = factor["Record_ID"]
        activity = activity_by_id.get(record_id)
        if activity is None:
            raise _error(
                "TRUSTED_ENRICHMENT_MISSING", "Record_ID无法从活动数据补全受控字段。", f"row {input_index}/Record_ID", record_id, "每个因子Record_ID必须在当前活动数据中唯一存在。"
            )
        if factor["Production_Eligible"] != "FALSE":
            raise _error(
                "SYNTHETIC_PRODUCTION_FLAG_FORBIDDEN", "合成演示因子不得标记为可生产使用。", f"row {input_index}/Production_Eligible", factor["Production_Eligible"], "合成演示固定为FALSE。"
            )
        if factor["Simulation_Flag"] != "TRUE":
            raise _error(
                "HISTORICAL_SIMULATION_FLAG_REQUIRED", "历史模拟输入必须标记Simulation_Flag=TRUE。", f"row {input_index}/Simulation_Flag", factor["Simulation_Flag"], "Day 6历史模拟路径固定为TRUE。"
            )
        occurrences[record_id] += 1
        identity = "|".join(
            [
                "DAY6-D1-V1",
                record_id,
                factor["EF_Value"],
                factor["EF_Unit"],
                factor["Source_Name"],
                factor["Source_Version"],
                str(occurrences[record_id]),
            ]
        )
        result_id = str(uuid.uuid5(RESULT_NAMESPACE, identity))
        inherited_warning = activity["QC_Status"] == "WARNING"
        issue_code = (
            "D1_SIMULATION_ONLY;D1_CATEGORY_SCOPE_APPLIED;"
            "D1_CHEMISTRY_APPLICABILITY_UNCONFIRMED;D1_SOURCE_METADATA_INCOMPLETE"
        )
        if inherited_warning:
            issue_code += ";D1_UPSTREAM_WARNING_INHERITED"
        row = {
            "Result_ID": result_id,
            "Record_ID": record_id,
            "Historical_EF_ID": factor_config["historical_ef_id"],
            "PACT_PF_ID": "",
            "Result_Type": factor["Result_Type"],
            "Receipt_Source_Type": factor_config["receipt_source_type"],
            "Receipt_Source_Organization": factor_config["receipt_source_organization"],
            "Receipt_Source_File": metadata["receipt_file"],
            "Receipt_Source_SHA256": metadata["receipt_sha256"],
            "Received_At": metadata["received_at"],
            "Receipt_Schema_Version": factor_config["receipt_schema_version"],
            "Raw_Payload_Path": metadata["raw_payload_path"],
            "EF_Value": format(Decimal(factor["EF_Value"]), "f"),
            "EF_Unit": factor["EF_Unit"],
            "EF_Source_Type": factor_config["ef_source_type"],
            "Source_Name": factor["Source_Name"],
            "Source_Version": factor["Source_Version"],
            "EF_Description": factor_config["ef_description"],
            "Technology": "",
            "Geography": factor_config["geography"],
            "Reference_Year": "",
            "System_Boundary": "",
            "Assessment_Date": "",
            "Remarks": factor_config["remarks"],
            "Evidence_File": metadata["receipt_file"],
            "Updated_Date": "",
            "Activity_Year": activity["Year"],
            "Simulation_Trial_Scope": factor_config["simulation_trial_scope"],
            "Historical_EF_Trial_Scope": factor_config["historical_ef_trial_scope"],
            "Activity_Data_Context": activity["Total_Weight_kg"],
            "Activity_Unit_Context": "kg/year",
            "Chemistry_Context": activity["Chemistry"],
            "Upstream_QC_Status": activity["QC_Status"],
            "Upstream_Issue_Code": activity["Issue_Code"],
            "Historical_Relationship_ID": _relationship_id(record_id),
            "Historical_Link_Method": factor_config["historical_link_method"],
            "Simulation_Flag": factor["Simulation_Flag"],
            "Simulation_Type": factor_config["simulation_type"],
            "Simulation_Source": factor_config["simulation_source"],
            "Simulation_Purpose": factor_config["simulation_purpose"],
            "Data_Truth_Class": factor_config["data_truth_class"],
            "Synthetic_Test_Flag": factor_config["synthetic_test_flag"],
            "Production_Eligible": factor["Production_Eligible"],
            "D1_Record_Status": "WARNING",
            "D1_Issue_Code": issue_code,
        }
        if set(row) != set(target_fields):
            raise _error(
                "D1_ADAPTER_INTERNAL_SCHEMA_ERROR", "45字段适配器内部字段不完整。", record_id, sorted(set(target_fields) - set(row)), "适配结果必须完整覆盖D1 45字段。"
            )
        output.append({field: str(row[field]) for field in target_fields})
        audit.append(
            {
                "Record_ID": record_id,
                "Result_ID": result_id,
                "Input_Row": str(input_index),
                "Receipt_Source_File": metadata["receipt_file"],
                "Receipt_Source_SHA256": metadata["receipt_sha256"],
                "Source_Field_Count": "8",
                "Target_Field_Count": "45",
                "Enrichment_Source": "DAY5_ACTIVITY_36_FIELDS+WP4_FROZEN_RELATIONSHIP_CONFIG",
                "Adaptation_Status": "PASS_WITH_SIMULATION_WARNING",
                "Issue_Code": issue_code,
            }
        )
    if len({row["Result_ID"] for row in output}) != len(output):
        raise _error(
            "RESULT_ID_NOT_UNIQUE", "生成的Result_ID不唯一。", "D1 output", len(output), "Result_ID必须确定性且唯一。"
        )
    return output, audit
