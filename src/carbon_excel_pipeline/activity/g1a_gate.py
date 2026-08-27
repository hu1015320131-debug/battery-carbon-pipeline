"""Independent WP2/WP3 reconciliation for the Day 5 upstream gate."""

from __future__ import annotations

import hashlib
from collections import Counter
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


STANDARD_NUMERIC_FIELDS = {
    "Year",
    "Source_Row",
    "PCS",
    "Unit_Weight_g",
    "Original_Activity_Value",
}
ACTIVITY_NUMERIC_FIELDS = STANDARD_NUMERIC_FIELDS | {
    "Total_Weight_g",
    "Total_Weight_kg",
    "Total_Weight_t",
    "Activity_Diff_g",
    "Activity_Diff_Rate",
}
THIRD_PARTY_NUMERIC_FIELDS = {"Year", "Activity_Data", "PCS", "Unit_Weight_g"}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def read_xlsx_table(
    path: Path,
    *,
    sheet_name: str,
    required_header: str = "Record_ID",
) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        rows = list(workbook[sheet_name].iter_rows(values_only=True))
    finally:
        workbook.close()
    header_index = next(
        index
        for index, row in enumerate(rows[:12])
        if required_header in row
    )
    fields = [str(value) for value in rows[header_index] if value is not None]
    records = [
        dict(zip(fields, row))
        for row in rows[header_index + 1 :]
        if any(value is not None for value in row)
    ]
    return fields, records


def _decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        result = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None
    return result if result.is_finite() else None


def _equal_value(actual: Any, expected: Any, field: str, numeric_fields: set[str]) -> bool:
    if field not in numeric_fields:
        return ("" if actual is None else str(actual)) == (
            "" if expected is None else str(expected)
        )
    return _decimal(actual) == _decimal(expected)


def compare_table(
    actual: list[dict[str, Any]],
    actual_fields: list[str],
    expected: list[dict[str, Any]],
    expected_fields: list[str],
    *,
    numeric_fields: set[str],
    float_tail_tolerances: dict[str, Decimal] | None = None,
) -> dict[str, Any]:
    mismatch_counts: Counter[str] = Counter()
    semantic_tail_matches = 0
    compared_rows = min(len(actual), len(expected))
    for index in range(compared_rows):
        for field in expected_fields:
            actual_value = actual[index].get(field)
            expected_value = expected[index].get(field)
            if (
                float_tail_tolerances is not None
                and field in float_tail_tolerances
                and _decimal(actual_value) is not None
                and _decimal(expected_value) is not None
                and abs(
                    (_decimal(actual_value) or Decimal("0"))
                    - (_decimal(expected_value) or Decimal("0"))
                )
                <= float_tail_tolerances[field]
            ):
                semantic_tail_matches += int(
                    _decimal(actual_value) != _decimal(expected_value)
                )
                continue
            if (
                float_tail_tolerances is not None
                and field == "Activity_Diff_Rate"
                and _decimal(actual_value) == Decimal("0")
            ):
                expected_diff = _decimal(expected[index].get("Activity_Diff_g"))
                diff_tolerance = float_tail_tolerances.get(
                    "Activity_Diff_g", Decimal("0")
                )
                if expected_diff is not None and abs(expected_diff) <= diff_tolerance:
                    semantic_tail_matches += int(_decimal(expected_value) != Decimal("0"))
                    continue
            if not _equal_value(actual_value, expected_value, field, numeric_fields):
                mismatch_counts[field] += 1
    actual_ids = [str(record.get("Record_ID", "")) for record in actual]
    expected_ids = [str(record.get("Record_ID", "")) for record in expected]
    return {
        "actual_record_count": len(actual),
        "baseline_record_count": len(expected),
        "actual_field_count": len(actual_fields),
        "baseline_field_count": len(expected_fields),
        "field_order_equal": actual_fields == expected_fields,
        "record_id_order_equal": actual_ids == expected_ids,
        "mismatched_cell_count": sum(mismatch_counts.values()),
        "field_mismatch_counts": dict(sorted(mismatch_counts.items())),
        "registered_float_tail_semantic_matches": semantic_tail_matches,
    }


def evaluate_g1a(
    *,
    standard_records: list[dict[str, Any]],
    standard_fields: list[str],
    activity_records: list[dict[str, Any]],
    activity_fields: list[str],
    third_party_records: list[dict[str, Any]],
    third_party_fields: list[str],
    quality_counts: dict[str, int],
    record_open_item_count: int,
    interface_open_item_count: int,
    standard_baseline_path: Path,
    activity_baseline_path: Path,
    third_party_baseline_path: Path,
    quality_config: dict[str, Any],
) -> dict[str, Any]:
    baseline_standard_fields, baseline_standard = read_xlsx_table(
        standard_baseline_path, sheet_name="2025_Standard"
    )
    baseline_activity_fields, baseline_activity = read_xlsx_table(
        activity_baseline_path, sheet_name="2025_Activity_Data"
    )
    baseline_third_fields, baseline_third = read_xlsx_table(
        third_party_baseline_path, sheet_name="2025_Third_Party_Input"
    )
    tolerances = {
        field: Decimal(value)
        for field, value in quality_config[
            "registered_float_tail_tolerance_by_field"
        ].items()
    }
    comparisons = {
        "standard_31": compare_table(
            standard_records,
            standard_fields,
            baseline_standard,
            baseline_standard_fields,
            numeric_fields=STANDARD_NUMERIC_FIELDS,
        ),
        "activity_36": compare_table(
            activity_records,
            activity_fields,
            baseline_activity,
            baseline_activity_fields,
            numeric_fields=ACTIVITY_NUMERIC_FIELDS,
            float_tail_tolerances=tolerances,
        ),
        "third_party_20": compare_table(
            third_party_records,
            third_party_fields,
            baseline_third,
            baseline_third_fields,
            numeric_fields=THIRD_PARTY_NUMERIC_FIELDS,
        ),
    }
    total_kg = sum(
        (Decimal(str(record["Total_Weight_kg"])) for record in activity_records),
        Decimal("0"),
    )
    expected_counts = quality_config.get("expected_private_counts")
    expected_total = quality_config.get("expected_private_activity_total_kg")
    checks = {
        "standard_31_reconciled": all(
            (
                comparisons["standard_31"]["actual_record_count"]
                == comparisons["standard_31"]["baseline_record_count"],
                comparisons["standard_31"]["field_order_equal"],
                comparisons["standard_31"]["record_id_order_equal"],
                comparisons["standard_31"]["mismatched_cell_count"] == 0,
            )
        ),
        "activity_36_reconciled": all(
            (
                comparisons["activity_36"]["actual_record_count"]
                == comparisons["activity_36"]["baseline_record_count"],
                comparisons["activity_36"]["field_order_equal"],
                comparisons["activity_36"]["record_id_order_equal"],
                comparisons["activity_36"]["mismatched_cell_count"] == 0,
            )
        ),
        "third_party_20_reconciled": all(
            (
                comparisons["third_party_20"]["actual_record_count"]
                == comparisons["third_party_20"]["baseline_record_count"],
                comparisons["third_party_20"]["field_order_equal"],
                comparisons["third_party_20"]["record_id_order_equal"],
                comparisons["third_party_20"]["mismatched_cell_count"] == 0,
            )
        ),
        "quality_counts_reconciled": expected_counts is None or quality_counts == expected_counts,
        "activity_total_reconciled": expected_total is None or total_kg == Decimal(str(expected_total)),
        "record_open_items_reconciled": quality_config.get("expected_private_record_open_items") is None
        or record_open_item_count == quality_config.get("expected_private_record_open_items"),
        "interface_open_items_reconciled": quality_config.get("expected_private_interface_open_items") is None
        or interface_open_item_count == quality_config.get("expected_private_interface_open_items"),
        "error_records_zero": quality_counts.get("ERROR", 0) == 0,
    }
    passed = all(checks.values())
    return {
        "gate_status": quality_config["g1a_pass_status"]
        if passed
        else quality_config["g1a_fail_status"],
        "status": "PASS" if passed else "FAIL",
        "checks": checks,
        "quality_counts": quality_counts,
        "activity_total_kg_per_year": str(total_kg),
        "comparisons": comparisons,
        "baseline_evidence_hashes": {
            "standard_31_sha256": sha256_file(standard_baseline_path),
            "activity_36_sha256": sha256_file(activity_baseline_path),
            "third_party_20_sha256": sha256_file(third_party_baseline_path),
        },
        "baselines_used_as_calculation_input": False,
        "baselines_used_for_post_generation_reconciliation": True,
        "row_values_exported": False,
    }
