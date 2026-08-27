"""WP6-8.3 UI helpers: recognition display, current-run isolation, business downloads.

Display-only. Does not change Router, Boundary, EF Policy, or Decimal calculation cores.
"""

from __future__ import annotations

import csv
import io
import json
from decimal import Decimal, InvalidOperation, localcontext
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from carbon_excel_pipeline.ui.reason_mapper import (
    display_historical_status,
    display_reason_code,
    display_route,
    display_status,
)
from carbon_excel_pipeline.wp6_8_4.business_units import filter_by_business_unit


PROJECT_ROOT = Path(__file__).resolve().parents[3]
DISPLAY_POLICY = PROJECT_ROOT / "config" / "wp6" / "wp6_8_3_display_policies.json"

SEMANTIC_FIELD_LABELS = {
    "Business_Unit": "事业部",
    "Purchase_Type": "采购类型",
    "Purchase_Category": "物料类别",
    "Product_Description": "物料描述",
    "Reported_Activity_Value": "年度采购量",
    "EF_Value": "排放因子",
    "EF_Source": "排放因子来源",
    "Historical_GHG_Value": "历史排放量",
    "Quantity_PCS": "采购数量",
    "Quantity_Unit": "采购数量单位",
    "Unit_Weight": "单件重量",
    "Reported_Purchase_Quantity": "年度采购量（重量）",
    "Chemistry": "化学体系",
    "Supplier": "供应商",
    "Project": "项目",
    "Model": "型号",
}

DIRECT_MASS_FIELDS = (
    "Business_Unit",
    "Purchase_Type",
    "Purchase_Category",
    "Product_Description",
    "Reported_Activity_Value",
    "EF_Value",
    "EF_Source",
    "Historical_GHG_Value",
)
PCS_FIELDS = (
    "Purchase_Category",
    "Product_Description",
    "Quantity_PCS",
    "Quantity_Unit",
    "Unit_Weight",
    "Reported_Purchase_Quantity",
    "Chemistry",
    "Supplier",
    "Project",
    "Model",
)
NOT_APPLICABLE_FIELDS = {
    "DIRECT_REPORTED_MASS": ("Quantity_PCS", "Unit_Weight", "Quantity_Unit"),
    "PCS_WEIGHT_DERIVED": ("Reported_Activity_Value", "Historical_GHG_Value"),
}


def _load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _read_csv(path: Path) -> list[dict[str, str]]:
    if not path.is_file():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _decimal(value: Any) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        number = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None
    return number if number.is_finite() else None


def current_run_bound(
    *,
    run_id: str | None,
    input_sha256: str | None,
    current_run_id: str | None,
    current_sha256: str | None,
    input_set_sha256: str | None = None,
    current_set_sha256: str | None = None,
) -> bool:
    if not run_id or not current_run_id:
        return False
    if run_id != current_run_id:
        return False
    set_expected = (current_set_sha256 or "").upper()
    set_actual = (input_set_sha256 or "").upper()
    if set_expected and set_actual:
        return set_expected == set_actual
    expected = (current_sha256 or "").upper()
    actual = (input_sha256 or "").upper()
    return bool(expected) and expected == actual


def activity_route_from_view(view: dict[str, Any]) -> str | None:
    capability = view.get("capability_summary") or {}
    paths = capability.get("supported_activity_paths") or []
    if isinstance(paths, list) and len(paths) == 1:
        return str(paths[0])
    if int(capability.get("direct_reported_mass_count") or 0) > 0 and int(
        capability.get("pcs_weight_derived_count") or 0
    ) == 0:
        return "DIRECT_REPORTED_MASS"
    if int(capability.get("pcs_weight_derived_count") or 0) > 0 and int(
        capability.get("direct_reported_mass_count") or 0
    ) == 0:
        return "PCS_WEIGHT_DERIVED"
    return None


def recognition_mapping_rows(view: dict[str, Any], sheet_name: str, activity_route: str | None) -> list[dict[str, str]]:
    sheets = view.get("semantic_mappings") or []
    sheet = next((item for item in sheets if item.get("sheet_name") == sheet_name), None)
    mapped = {
        item.get("semantic_field"): item
        for item in (sheet or {}).get("field_mappings", [])
        if item.get("semantic_field")
    }
    if activity_route == "DIRECT_REPORTED_MASS":
        wanted = DIRECT_MASS_FIELDS + ("Quantity_PCS", "Unit_Weight")
    elif activity_route == "PCS_WEIGHT_DERIVED":
        wanted = PCS_FIELDS
    else:
        wanted = tuple(dict.fromkeys(DIRECT_MASS_FIELDS + PCS_FIELDS))
    hidden = NOT_APPLICABLE_FIELDS.get(activity_route or "", ())
    rows: list[dict[str, str]] = []
    for field in wanted:
        label = SEMANTIC_FIELD_LABELS.get(field, field)
        if field in hidden:
            rows.append(
                {
                    "业务字段": label,
                    "识别列": "不适用",
                    "原始表头": "—",
                    "状态": "不适用",
                }
            )
            continue
        item = mapped.get(field)
        if not item:
            rows.append(
                {
                    "业务字段": label,
                    "识别列": "未识别",
                    "原始表头": "—",
                    "状态": "未识别",
                }
            )
            continue
        rows.append(
            {
                "业务字段": label,
                "识别列": f"{item.get('column_letter', '')}列",
                "原始表头": str(item.get("raw_header") or ""),
                "状态": display_status(item.get("mapping_status")),
            }
        )
    return rows


def comparison_ef_for_current_run(input_sha256: str) -> dict[str, str] | None:
    if not DISPLAY_POLICY.is_file():
        return None
    payload = _load_json(DISPLAY_POLICY)
    digest = (input_sha256 or "").upper()
    for policy in payload.get("comparison_policies", []):
        allowed = [str(item).upper() for item in policy.get("match", {}).get("input_sha256", [])]
        if digest in allowed:
            return {
                "comparison_ef": str(policy["comparison_ef"]),
                "comparison_label": str(policy.get("comparison_label") or "对比排放因子"),
                "current_label": str(policy.get("current_label") or "当前排放因子"),
            }
    return None


def factor_improvement_from_canonical(
    rows: list[dict[str, Any]],
    *,
    comparison_ef: str,
) -> dict[str, str] | None:
    if not rows:
        return None
    compare = _decimal(comparison_ef)
    if compare is None or compare == 0:
        return None
    with localcontext() as context:
        context.prec = 50
        activity = Decimal("0")
        current = Decimal("0")
        for row in rows:
            activity_kg = _decimal(row.get("Activity_Data_kg"))
            ef_now = _decimal(row.get("EF_Value"))
            if activity_kg is None or ef_now is None:
                continue
            activity += activity_kg
            current += activity_kg * ef_now
        if activity == 0:
            return None
        simulated = activity * compare
        reduction = simulated - current
        relative = (current / simulated - Decimal("1")) * Decimal("100") if simulated else Decimal("0")
        current_ef = _decimal(rows[0].get("EF_Value"))
        decline = (
            (Decimal("1") - current_ef / compare) * Decimal("100")
            if current_ef is not None
            else (Decimal("-1") * relative)
        )
        return {
            "activity_kg": format(activity, "f"),
            "comparison_ef": format(compare, "f"),
            "current_ef": format(current_ef, "f") if current_ef is not None else "",
            "simulated_emission_kgco2e": format(simulated, "f"),
            "current_emission_kgco2e": format(current, "f"),
            "reduction_kgco2e": format(reduction, "f"),
            "reduction_tco2e": format(reduction / Decimal("1000"), "f"),
            "relative_percent": format(relative, "f"),
            "ef_decline_percent": format(decline, "f"),
        }


def _csv_bytes(rows: list[dict[str, Any]], fields: list[str] | None = None) -> bytes:
    output = io.StringIO()
    columns = fields or (list(rows[0].keys()) if rows else ["说明"])
    writer = csv.DictWriter(output, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    if rows:
        writer.writerows(rows)
    else:
        writer.writerow({columns[0]: "暂无数据"})
    return output.getvalue().encode("utf-8-sig")


def _xlsx_bytes(sheets: list[tuple[str, list[dict[str, Any]], list[str]]]) -> bytes:
    workbook = Workbook()
    first = True
    for title, rows, fields in sheets:
        sheet = workbook.active if first else workbook.create_sheet()
        first = False
        sheet.title = title[:31]
        sheet.append(fields)
        if rows:
            for row in rows:
                sheet.append([row.get(field, "") for field in fields])
        else:
            sheet.append(["暂无数据"] + [""] * (len(fields) - 1))
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for index, header in enumerate(fields, start=1):
            width = max([len(str(header))] + [len(str(row.get(header, ""))) for row in rows[:80]])
            sheet.column_dimensions[get_column_letter(index)].width = min(width + 2, 42)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def historical_validation_display_rows(rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    return [
        {
            "记录编号": str(row.get("Record_ID") or ""),
            "原始数据行": str(row.get("Source_Row") or ""),
            "重新计算排放量（tCO2e）": str(row.get("Calculated_Emission_tCO2e") or ""),
            "历史排放量（tCO2e）": str(row.get("Historical_Emission_tCO2e") or ""),
            "差异（tCO2e）": str(row.get("Difference_tCO2e") or ""),
            "验证结果": display_historical_status(row.get("Validation_Status")),
        }
        for row in rows
    ]


def record_detail_display(row: dict[str, Any]) -> list[tuple[str, str]]:
    purchase = row.get("Quantity_PCS") or row.get("Original_Activity_Value") or ""
    emission = row.get("Emission_tCO2e") or row.get("Emission_kgCO2e") or ""
    return [
        ("记录编号", str(row.get("Record_ID") or "")),
        ("事业部", str(row.get("Business_Unit") or "")),
        ("物料类别", str(row.get("Purchase_Category") or "")),
        ("物料描述", str(row.get("Product_Description") or "")),
        ("供应商", str(row.get("Supplier") or row.get("Supplier_Name") or "")),
        ("化学体系", str(row.get("Chemistry") or "")),
        ("采购数量 / 年度采购量", str(purchase)),
        ("活动数据", str(row.get("Activity_Data_kg") or "")),
        ("活动数据生成方式", display_route(row.get("Activity_Method"))),
        ("排放因子", str(row.get("EF_Value") or "")),
        ("排放因子来源", str(row.get("EF_Source") or display_route(row.get("Factor_Route")))),
        ("排放量", str(emission)),
        ("核算状态", display_status(row.get("Calculation_QC") or row.get("Overall_Status"))),
        ("数据质量状态", display_status(row.get("Governance_QC"))),
        ("原始文件", str(row.get("Source_File") or "")),
        ("原始工作表", str(row.get("Source_Sheet") or "")),
        ("原始行号", str(row.get("Source_Row") or "")),
    ]


def detected_unit_options(rows: list[dict[str, Any]], summary_units: list[str] | None = None) -> list[str]:
    found: list[str] = []
    for value in [*(summary_units or []), *[str(row.get("Business_Unit") or "").strip() for row in rows]]:
        if value and value not in found:
            found.append(value)
    return ["全部", *found]


def build_business_download_pack(run_dir: Path, business_unit: str | None = None) -> dict[str, Any]:
    run = run_dir.expanduser().resolve()
    download = run / "08_download"
    canonical = filter_by_business_unit(_read_csv(download / "canonical_results.csv"), business_unit)
    scoped = bool(business_unit) and business_unit not in {"全部", "ALL"}
    third_party = _read_csv(run / "06_third_party_input" / "day5_third_party_20_fields.csv")
    if not third_party:
        third_party = _read_csv(download / "suggested_ledger_v1.csv")
    issues = _read_csv(download / "data_quality_issue_register.csv")
    summary = _load_json(download / "run_summary.json") if (download / "run_summary.json").is_file() else {}
    route = _load_json(run / "route_decision.json") if (run / "route_decision.json").is_file() else {}

    detail_fields = [
        "Record_ID",
        "物料描述",
        "事业部",
        "物料类别",
        "Activity",
        "Activity_Unit",
        "核算路径",
    ]
    result_fields = [
        "Record_ID",
        "物料描述",
        "Activity",
        "Activity_Unit",
        "EF",
        "EF_Unit",
        "EF_Source",
        "排放量",
        "核算状态",
        "异常提示",
    ]
    detail_rows = []
    result_rows = []
    for row in canonical:
        detail_rows.append(
            {
                "Record_ID": row.get("Record_ID", ""),
                "物料描述": row.get("Product_Description", ""),
                "事业部": row.get("Business_Unit", ""),
                "物料类别": row.get("Purchase_Category", ""),
                "Activity": row.get("Activity_Data_kg", ""),
                "Activity_Unit": row.get("Activity_Unit") or "kg/year",
                "核算路径": row.get("Activity_Method", ""),
            }
        )
        result_rows.append(
            {
                "Record_ID": row.get("Record_ID", ""),
                "物料描述": row.get("Product_Description", ""),
                "Activity": row.get("Activity_Data_kg", ""),
                "Activity_Unit": row.get("Activity_Unit") or "kg/year",
                "EF": row.get("EF_Value", ""),
                "EF_Unit": row.get("EF_Unit", ""),
                "EF_Source": row.get("EF_Source", ""),
                "排放量": row.get("Emission_kgCO2e", ""),
                "核算状态": display_status(row.get("Calculation_QC") or row.get("Overall_Status")),
                "异常提示": display_reason_code(row.get("Warning_Codes") or row.get("Governance_Issue_Codes") or "NONE"),
            }
        )
    run_info = [
        {"项目": key, "内容": json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else str(value)}
        for key, value in {
            "运行编号": summary.get("Run_ID") or run.name,
            "输入文件": summary.get("Input_File", ""),
            "文件指纹": summary.get("Input_SHA") or summary.get("Input_SHA256", ""),
            "活动路径": route.get("Activity_Route", ""),
            "排放因子路径": route.get("Factor_Route", ""),
            "记录数": summary.get("Total_Records") or summary.get("Boundary_Records") or len(canonical),
            "当前事业部": business_unit or "全部",
            "因子用途": "历史模拟",
            "生产使用": "否",
        }.items()
    ]
    third_fields = list(third_party[0].keys()) if third_party else ["说明"]
    issue_fields = list(issues[0].keys()) if issues else ["说明"]
    package = _xlsx_bytes(
        [
            ("01_电芯数据", detail_rows, detail_fields),
            ("02_碳核算结果", result_rows, result_fields),
            ("03_第三方输入", third_party, third_fields),
            ("04_异常记录", issues, issue_fields),
            ("05_运行信息", run_info, ["项目", "内容"]),
        ]
    )
    return {
        "cell_detail": {
            "display_name": "下载当前事业部电芯数据明细" if scoped else "下载电芯数据明细",
            "download_name": "当前事业部电芯数据明细.csv" if scoped else "电芯数据明细.csv",
            "data": _csv_bytes(detail_rows, detail_fields),
            "mime": "text/csv",
        },
        "carbon_result": {
            "display_name": "下载当前事业部碳核算结果" if scoped else "下载碳核算结果",
            "download_name": "当前事业部碳核算结果.csv" if scoped else "碳核算结果.csv",
            "data": _csv_bytes(result_rows, result_fields),
            "mime": "text/csv",
        },
        "third_party": {
            "display_name": "下载第三方因子匹配输入表",
            "download_name": "第三方因子匹配输入表.csv",
            "data": _csv_bytes(third_party, third_fields),
            "mime": "text/csv",
        },
        "package": {
            "display_name": "下载完整结果包",
            "download_name": "完整结果包.xlsx",
            "data": package,
            "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    }
