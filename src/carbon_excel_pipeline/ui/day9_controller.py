"""Testable Day 9 UI orchestration built only from existing core pipeline APIs."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from carbon_excel_pipeline.activity.day5_pipeline import run_day5_upstream_rebuild
from carbon_excel_pipeline.capability.pipeline import run_wp6_2_capability_detection
from carbon_excel_pipeline.calculation.day7_pipeline import (
    run_day7_calculation_and_lineage,
)
from carbon_excel_pipeline.cleaning.scope_filter import run_day3_scope_and_cleaning
from carbon_excel_pipeline.errors import PipelineUserError
from carbon_excel_pipeline.export.day8_export import run_day8_export
from carbon_excel_pipeline.export.day8_run_all import run_day8_full_pipeline
from carbon_excel_pipeline.factors.day6_pipeline import (
    run_day6_factors_and_matching,
)
from carbon_excel_pipeline.io.excel_importer import inspect_excel_to_run
from carbon_excel_pipeline.standardization.standardizer import (
    run_day4_standardization,
)
from carbon_excel_pipeline.wp6_3.pipeline import run_wp6_3_historical_reproduction
from carbon_excel_pipeline.wp6_8_1.pipeline import run_end_to_end_pipeline
from carbon_excel_pipeline.wp6_8_5.cell_scope import apply_cell_scope_to_run
from carbon_excel_pipeline.ui.reason_mapper import display_reason_code, display_status


PROJECT_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_LOCAL_CONFIG = PROJECT_ROOT / "config" / "private" / "local_paths.json"
NON_PRODUCTION_NOTICE = (
    "本地历史模拟Demo：Production_Eligible=FALSE。禁止用于生产决策，"
    "不会上传外网，也不会执行GitHub操作。"
)
REQUIRED_TARGETS = (
    "Purchase_Category",
    "Product_Description",
    "PCS",
    "Source_Unit",
    "Unit_Weight_g_per_PCS",
    "Annual_Purchase_g_per_year",
)
TARGET_FIELD_LABELS = {
    "Purchase_Category": "采购分类",
    "Product_Description": "原始物料信息",
    "PCS": "采购数量",
    "Source_Unit": "采购数量单位",
    "Unit_Weight_g_per_PCS": "单件重量",
    "Annual_Purchase_g_per_year": "年度采购量",
}
BUSINESS_STAGE_LABELS = {
    "day2": "数据接收与结构识别",
    "day3": "范围筛选与原始清洗",
    "day4": "标准化与记录编号",
    "day5": "数据质量与活动量核验",
    "day6": "排放因子匹配与核算路由",
    "day7": "碳核算与数据追溯",
    "day8": "结果文件生成与验证",
}
STATUS_TRANSLATIONS = {
    "PASS": "通过",
    "WARNING": "警告",
    "ERROR": "错误",
    "BLOCKED": "已阻断",
    "PASS_WITH_WARNING": "通过（含治理警告）",
    "PARTIAL_RESULT": "部分结果",
    "OPEN": "待处理",
    "INDEPENDENT_CALCULATION_PASS": "独立核算验证通过",
    "NOT_AVAILABLE_FOR_SINGLE_INPUT_RUN": "单文件运行不可用",
    "NOT_RUN": "尚未运行",
    "ROUTED": "已完成路径选择",
    "UNKNOWN": "状态未知",
    "G1A_UPSTREAM_REBUILD_RECONCILED": "上游重建对账通过",
    "DAY6_HISTORICAL_FACTOR_EXACT_MATCH_LOCKED": "历史模拟因子已精确匹配并锁定",
    "DAY7_CALCULATION_LINEAGE_RECONCILED": "碳核算与两层追溯对账通过",
    "G2_CLI_END_TO_END_PASS": "完整结果文件生成与验证通过",
    "TRUE": "是",
    "FALSE": "否",
    "EXACT_MATCH": "精确匹配",
    "EXACT_LOCKED": "精确锁定",
    "CALCULATED_WITH_WARNING": "已核算（含限制说明）",
}
CHEMISTRY_TRANSLATIONS = {
    "LCO": "钴酸锂",
    "LFP": "磷酸铁锂",
    "NCM": "三元锂",
    "NCA": "镍钴铝",
    "UNKNOWN": "未知",
    "": "未知",
}
CHEMISTRY_DISPLAY_PRIORITY = ("钴酸锂", "磷酸铁锂", "三元锂", "镍钴铝")
ISSUE_TRANSLATIONS = {
    "NONE": "无",
    "CUSTOMER_UNMAPPED": "客户未完成映射",
    "CHEMISTRY_UNKNOWN": "化学体系未知",
    "SUPPLIER_UNMAPPED": "供应商未完成映射",
    "MISSING_REQUIRED_FIELD": "缺少必填字段",
}
STAGE_REPORTS = {
    "day2": "01_import/day2_run_summary.json",
    "day3": "02_scope_filter/day3_scope_summary.json",
    "day4": "03_standardized/day4_standardization_summary.json",
    "day5": "05_activity/day5_upstream_summary.json",
    "day6": "08_matching/day6_summary.json",
    "day7": "10_output/day7_summary.json",
    "day8": "10_output/day8_run_report.json",
}


@dataclass(frozen=True, slots=True)
class Day9Paths:
    project_root: Path
    local_config_path: Path
    raw_input: Path
    run_root: Path
    output_root: Path
    artifact_work_root: Path
    upload_root: Path
    node_executable: Path
    node_modules: Path
    private_id_baseline: Path
    private_mapping_workbook: Path
    standard_baseline: Path
    activity_baseline: Path
    third_party_baseline: Path
    wp5_open_items: Path

    def private_processing_paths(self) -> dict[str, Path]:
        return {
            "private_id_baseline": self.private_id_baseline,
            "private_mapping_workbook": self.private_mapping_workbook,
            "standard_baseline": self.standard_baseline,
            "activity_baseline": self.activity_baseline,
            "third_party_baseline": self.third_party_baseline,
        }

    def required_inputs(self) -> dict[str, Path]:
        return {
            "原始Excel": self.raw_input,
            "WP2正式编号基线": self.private_id_baseline,
            "WP2映射工作簿": self.private_mapping_workbook,
            "WP2标准数据基线": self.standard_baseline,
            "WP3活动数据基线": self.activity_baseline,
            "WP3第三方输入基线": self.third_party_baseline,
            "WP5治理Open Items": self.wp5_open_items,
            "Node运行时": self.node_executable,
            "Node依赖目录": self.node_modules,
        }

    def readiness(self) -> dict[str, Any]:
        missing = [label for label, path in self.required_inputs().items() if not path.exists()]
        external_roots = all(
            not _is_relative_to(path.resolve(), self.project_root.resolve())
            for path in (self.run_root, self.output_root, self.artifact_work_root, self.upload_root)
        )
        return {
            "status": "PASS" if not missing and external_roots else "BLOCKED",
            "missing": missing,
            "external_run_and_output_roots": external_roots,
            "github_operations_enabled": False,
            "production_eligible": False,
        }


def _is_relative_to(path: Path, root: Path) -> bool:
    try:
        path.relative_to(root)
        return True
    except ValueError:
        return False


def _load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def _find_named(paths: Iterable[str], fragment: str) -> Path:
    match = next((Path(item) for item in paths if fragment in Path(item).name), None)
    if match is None:
        return Path(f"__MISSING__/{fragment}")
    return match


def load_day9_paths(config_path: Path | None = None) -> Day9Paths:
    path = (config_path or DEFAULT_LOCAL_CONFIG).expanduser().resolve()
    payload = _load_json(path)
    wp2_files = payload.get("wp2_files", [])
    wp3_files = payload.get("wp3_files", [])
    development_root = PROJECT_ROOT.parent
    run_root = Path(payload["run_root"])
    wp5_root = Path(payload["wp5_root"])
    private_id = Path(
        payload.get("private_id_baseline")
        or _find_named(wp2_files, "WP2-D5_2025标准数据集")
    )
    mapping_workbook = Path(
        payload.get("private_mapping_workbook")
        or _find_named(wp2_files, "WP2-D3_项目与属性映射表")
    )
    standard_baseline = Path(payload.get("standard_baseline") or private_id)
    activity_baseline = Path(
        payload.get("activity_baseline")
        or _find_named(wp3_files, "WP3-D3_2025_Activity_Data")
    )
    third_party_baseline = Path(
        payload.get("third_party_baseline")
        or _find_named(wp3_files, "WP3-D5_2025_Third_Party_Input_V1")
    )
    return Day9Paths(
        project_root=PROJECT_ROOT,
        local_config_path=path,
        raw_input=Path(payload["raw_input"]),
        run_root=run_root,
        output_root=Path(
            payload.get("output_root") or development_root / "outputs" / "day9_streamlit"
        ),
        artifact_work_root=Path(
            payload.get("artifact_work_root")
            or development_root / "99_temp" / "wp5_day9_streamlit" / "artifact_work"
        ),
        upload_root=Path(
            payload.get("upload_root")
            or development_root / "99_temp" / "wp5_day9_streamlit" / "uploads"
        ),
        node_executable=Path(payload.get("node_executable", "__MISSING__/node.exe")),
        node_modules=Path(payload.get("node_modules", "__MISSING__/node_modules")),
        private_id_baseline=private_id,
        private_mapping_workbook=mapping_workbook,
        standard_baseline=standard_baseline,
        activity_baseline=activity_baseline,
        third_party_baseline=third_party_baseline,
        wp5_open_items=Path(
            payload.get("wp5_open_items")
            or wp5_root / "06_output" / "WP5-D5_Open_Items_V1.csv"
        ),
    )


def source_identity(path: Path) -> str:
    resolved = path.expanduser().resolve()
    digest = hashlib.sha256()
    with resolved.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def should_invalidate_run(active_source_identity: str | None, selected_identity: str) -> bool:
    return bool(active_source_identity and active_source_identity != selected_identity)


def save_uploaded_file(content: bytes, filename: str, paths: Day9Paths) -> Path:
    safe_name = re.sub(r"[^0-9A-Za-z._\-\u4e00-\u9fff ]", "_", Path(filename).name)
    if not safe_name:
        raise ValueError("上传文件名无效。")
    digest = hashlib.sha256(content).hexdigest().upper()
    destination = paths.upload_root / digest[:12] / safe_name
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_bytes(content)
    return destination


def inspect_stage(source_path: Path, paths: Day9Paths) -> dict[str, Any]:
    readiness = paths.readiness()
    if readiness["status"] != "PASS":
        raise ValueError(f"本地路径配置未通过：{readiness}")
    result = inspect_excel_to_run(
        source_path,
        run_root=paths.run_root,
        alias_config_path=PROJECT_ROOT / "config" / "import" / "field_aliases.json",
        max_size_bytes=50 * 1024 * 1024,
    )
    run_dir = Path(str(result.get("run_directory", "")))
    if (run_dir / "01_import" / "recognized_records.json").is_file():
        result["cell_scope"] = apply_cell_scope_to_run(run_dir)
        result["capability_detection"] = run_wp6_2_capability_detection(run_dir)
    return result


def run_end_to_end_stage(
    source_path: Path,
    paths: Day9Paths,
    existing_run_dir: Path | None = None,
    extra_source_paths: list[Path] | None = None,
    file_roles: dict[str, str] | None = None,
) -> dict[str, Any]:
    """Call the unified WP6-8.1 orchestrator; the UI does not calculate."""

    return run_end_to_end_pipeline(
        source_path,
        run_root=paths.run_root,
        existing_run_dir=existing_run_dir,
        private_paths=paths.private_processing_paths(),
        extra_source_paths=extra_source_paths,
        file_roles=file_roles,
    )


def load_e2e_view(run_dir: Path) -> dict[str, Any] | None:
    run = run_dir.expanduser().resolve()
    summary_path = run / "e2e_run_summary.json"
    if not summary_path.is_file():
        return None
    download = run / "08_download"
    route_path = run / "route_decision.json"
    return {
        "e2e_summary": _load_json(summary_path),
        "route": _load_json(route_path) if route_path.is_file() else {},
        "stage_status": _load_json(run / "pipeline_stage_status.json")
        if (run / "pipeline_stage_status.json").is_file()
        else {},
        "live_summary": _load_json(download / "wp6_8_live_summary.json")
        if (download / "wp6_8_live_summary.json").is_file()
        else {},
        "run_summary": _load_json(download / "run_summary.json")
        if (download / "run_summary.json").is_file()
        else {},
        "canonical": _read_csv_rows(download / "canonical_results.csv")
        if (download / "canonical_results.csv").is_file()
        else [],
        "ledger": _read_csv_rows(download / "suggested_ledger_v1.csv")
        if (download / "suggested_ledger_v1.csv").is_file()
        else [],
        "audit": _read_csv_rows(download / "audit_detail.csv")
        if (download / "audit_detail.csv").is_file()
        else [],
        "download_dir": download if download.is_dir() else None,
    }


def run_wp6_3_2024_stage(run_dir: Path, paths: Day9Paths) -> dict[str, Any]:
    """Run the shared WP6-3 backend; the UI performs no calculation itself."""

    return run_wp6_3_historical_reproduction(
        run_dir,
        output_root=paths.output_root / "wp6-3",
        policy_config_path=PROJECT_ROOT / "config" / "wp6" / "synthetic_direct_mass_policy.json",
    )


def load_wp6_3_view(output_dir: Path) -> dict[str, Any]:
    output = output_dir.expanduser().resolve()
    return {
        "summary": _load_json(output / "wp6_3_summary.json"),
        "records": _read_csv_rows(output / "2024_canonical_results.csv"),
    }


def latest_wp6_4_run(run_root: Path) -> Path | None:
    root = run_root.expanduser().resolve()
    if not root.is_dir():
        return None
    candidates = [
        path
        for path in root.iterdir()
        if path.is_dir() and (path / "wp6_4_summary.json").is_file()
    ]
    return max(candidates, key=lambda path: path.stat().st_mtime) if candidates else None


def load_wp6_4_view(output_dir: Path) -> dict[str, Any]:
    output = output_dir.expanduser().resolve()
    return {
        "summary": _load_json(output / "wp6_4_summary.json"),
        "strict_records": _read_csv_rows(output / "2025_strict_regression.csv"),
        "quality_records": _read_csv_rows(output / "2025_qc_regression.csv"),
        "scope_comparison": _read_csv_rows(output / "2025_scope_comparison.csv"),
        "forward_compatibility": _load_json(
            output / "2024_shared_model_forward_compatibility.json"
        ),
    }


def latest_wp6_5_run(run_root: Path) -> Path | None:
    root = run_root.expanduser().resolve()
    if not root.is_dir():
        return None
    candidates = [
        path
        for path in root.iterdir()
        if path.is_dir() and (path / "independent_validation_summary.json").is_file()
    ]
    return max(candidates, key=lambda path: path.stat().st_mtime) if candidates else None


def load_wp6_5_view(output_dir: Path) -> dict[str, Any]:
    """Load completed validation artifacts; the UI performs no recalculation."""

    output = output_dir.expanduser().resolve()
    return {
        "summary": _load_json(output / "independent_validation_summary.json"),
        "records_2024": _read_csv_rows(output / "2024_independent_validation.csv"),
        "records_2025": _read_csv_rows(output / "2025_independent_validation.csv"),
        "manual_samples": _read_csv_rows(output / "independent_manual_samples.csv"),
    }


def latest_wp6_6_run(run_root: Path) -> Path | None:
    root = run_root.expanduser().resolve()
    if not root.is_dir():
        return None
    candidates = [
        path
        for path in root.iterdir()
        if path.is_dir() and (path / "wp6_6_analysis_summary.json").is_file()
    ]
    return max(candidates, key=lambda path: path.stat().st_mtime) if candidates else None


def load_wp6_6_view(output_dir: Path) -> dict[str, Any]:
    """Load completed WP6-6 artifacts; Streamlit performs no scenario arithmetic."""

    output = output_dir.expanduser().resolve()
    return {
        "summary": _load_json(output / "wp6_6_analysis_summary.json"),
        "scenarios": _load_json(output / "scenario_abcd_summary.json"),
        "factor_effect": _load_json(output / "factor_effect_summary.json"),
        "activity_scope_effect": _load_json(
            output / "activity_scope_effect_summary.json"
        ),
        "symmetric_decomposition": _load_json(
            output / "symmetric_decomposition.json"
        ),
        "records_2024": _read_csv_rows(output / "2024_factor_counterfactual.csv"),
        "records_2025": _read_csv_rows(output / "2025_factor_counterfactual.csv"),
    }


def latest_wp6_7_run(run_root: Path) -> Path | None:
    root = run_root.expanduser().resolve()
    if not root.is_dir():
        return None
    candidates = [
        path
        for path in root.iterdir()
        if path.is_dir() and (path / "wp6_7_analysis_summary.json").is_file()
    ]
    return max(candidates, key=lambda path: path.stat().st_mtime) if candidates else None


def load_wp6_7_view(output_dir: Path) -> dict[str, Any]:
    """Load completed WP6-7 artifacts; the UI performs no analysis arithmetic."""

    output = output_dir.expanduser().resolve()
    return {
        "summary": _load_json(output / "wp6_7_analysis_summary.json"),
        "scorecard_2024": _load_json(output / "2024_data_quality_scorecard.json"),
        "scorecard_2025": _load_json(output / "2025_data_quality_scorecard.json"),
        "issues": _read_csv_rows(output / "data_quality_issue_register.csv"),
        "dimensions": _read_csv_rows(output / "dimension_availability.csv"),
        "management_2024": _read_csv_rows(output / "2024_management_summary.csv"),
        "management_2025": _read_csv_rows(output / "2025_management_summary.csv"),
        "contributors_2024": _read_csv_rows(
            output / "2024_top_emission_contributors.csv"
        ),
        "contributors_2025": _read_csv_rows(
            output / "2025_top_emission_contributors.csv"
        ),
        "factor_impact_2024": _read_csv_rows(output / "2024_top_factor_impact.csv"),
        "factor_impact_2025": _read_csv_rows(output / "2025_top_factor_impact.csv"),
        "lineage": _load_json(output / "lineage_quality_summary.json"),
    }


def latest_wp6_8_run(run_root: Path) -> Path | None:
    """Return the newest completed WP6-8 delivery without generating one."""

    root = run_root.expanduser().resolve()
    if not root.is_dir():
        return None
    candidates = [
        path
        for path in root.iterdir()
        if path.is_dir()
        and (path / "wp6_8_integration_summary.json").is_file()
        and (path / "download_manifest.json").is_file()
    ]
    return max(candidates, key=lambda path: path.stat().st_mtime) if candidates else None


def load_wp6_8_view(output_dir: Path) -> dict[str, Any]:
    """Load completed WP6-8 artifacts; the UI performs no business arithmetic."""

    output = output_dir.expanduser().resolve()
    return {
        "summary": _load_json(output / "wp6_8_integration_summary.json"),
        "run_summary": _load_json(output / "run_summary.json"),
        "manifest": _load_json(output / "download_manifest.json"),
        "canonical": _read_csv_rows(output / "canonical_results.csv"),
        "ledger": _read_csv_rows(output / "suggested_ledger_v1.csv"),
        "audit": _read_csv_rows(output / "audit_detail.csv"),
        "excluded": _read_csv_rows(output / "excluded_records.csv"),
        "dimensions": _read_csv_rows(output / "dimension_availability.csv"),
    }


def wp6_8_download_artifacts(output_dir: Path) -> list[dict[str, Any]]:
    """Read only files declared by the completed WP6-8 download manifest."""

    output = output_dir.expanduser().resolve()
    manifest = _load_json(output / "download_manifest.json")
    items = [*manifest.get("files", []), {"File_Name": "download_manifest.json"}]
    result: list[dict[str, Any]] = []
    seen: set[str] = set()
    for item in items:
        name = Path(str(item.get("File_Name", ""))).name
        if not name or name in seen:
            continue
        path = (output / name).resolve()
        if path.parent != output or not path.is_file():
            continue
        seen.add(name)
        result.append(
            {
                "display_name": name,
                "download_name": name,
                "data": path.read_bytes(),
                "mime": {
                    ".csv": "text/csv",
                    ".json": "application/json",
                    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    ".md": "text/markdown",
                }.get(path.suffix.lower(), "application/octet-stream"),
                "size_bytes": path.stat().st_size,
            }
        )
    return result


def load_inspection_view(run_dir: Path) -> dict[str, Any]:
    run = run_dir.expanduser().resolve()
    inventory = _load_json(run / "01_import" / "sheet_inventory.json")
    mappings = _load_json(run / "01_import" / "field_mapping_preview.json")
    previews = _load_json(run / "01_import" / "input_preview.json")
    recognition_summary_path = run / "01_import" / "recognition_summary.json"
    semantic_mapping_path = run / "01_import" / "semantic_field_mapping.json"
    capability_path = run / "02_capability" / "dataset_capabilities.json"
    recognition_summary = (
        _load_json(recognition_summary_path)
        if recognition_summary_path.is_file()
        else None
    )
    semantic_mappings = (
        _load_json(semantic_mapping_path) if semantic_mapping_path.is_file() else []
    )
    capability_summary = _load_json(capability_path) if capability_path.is_file() else None
    columns: dict[str, list[dict[str, Any]]] = {}
    for mapping in mappings:
        sheet = mapping["sheet_name"]
        header_row = int(mapping.get("header_row") or 1)
        preview = next((item for item in previews if item["sheet_name"] == sheet), None)
        row = next(
            (item for item in (preview or {}).get("rows", []) if int(item["source_row"]) == header_row),
            None,
        )
        columns[sheet] = [
            {
                "column_index": index,
                "column_letter": get_column_letter(index),
                "source_header": "" if value is None else str(value),
            }
            for index, value in enumerate((row or {}).get("values", []), start=1)
            if value is not None and str(value) != ""
        ]
    return {
        "inventory": inventory,
        "mappings": mappings,
        "previews": previews,
        "columns": columns,
        "recognition_summary": recognition_summary,
        "semantic_mappings": semantic_mappings,
        "capability_summary": capability_summary,
    }


def default_mapping_overrides(view: dict[str, Any], target_sheet: str) -> dict[str, dict[str, Any]]:
    report = next(item for item in view["mappings"] if item["sheet_name"] == target_sheet)
    result: dict[str, dict[str, Any]] = {}
    for item in report.get("mapping_preview", []):
        matched = item.get("matched_columns", [])
        if item.get("target_field") in REQUIRED_TARGETS and len(matched) == 1:
            result[item["target_field"]] = dict(matched[0])
    return result


def run_scope_stage(
    run_dir: Path,
    paths: Day9Paths,
    *,
    target_sheet: str,
    target_purchase_category: str,
    supplier_markers: list[str],
    mapping_overrides: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    base_path = PROJECT_ROOT / "config" / "scope" / "public_synthetic_scope.json"
    scope = _load_json(base_path)
    baseline_values = (
        scope["target_sheet"],
        scope["target_purchase_category"],
        scope["supplier_markers"],
    )
    scope.update(
        {
            "target_sheet": target_sheet,
            "target_purchase_category": target_purchase_category,
            "supplier_markers": supplier_markers,
        }
    )
    if baseline_values != (target_sheet, target_purchase_category, supplier_markers):
        scope["config_id"] = "DAY9_UI_RUNTIME_SCOPE_V1"
        scope["expected_input_records"] = None
        scope["expected_candidate_records"] = None
    runtime_path = run_dir / "11_logs" / "day9_scope_config_snapshot.json"
    _write_json(
        runtime_path,
        {
            **scope,
            "mapping_overrides": mapping_overrides,
            "user_confirmed": True,
        },
    )
    return run_day3_scope_and_cleaning(
        run_dir,
        scope_config_path=runtime_path,
        unit_config_path=PROJECT_ROOT / "config" / "cleaning" / "unit_mappings.json",
        mapping_overrides=mapping_overrides,
    )


def run_standardize_stage(run_dir: Path, paths: Day9Paths) -> dict[str, Any]:
    return run_day4_standardization(
        run_dir,
        profile_config_path=PROJECT_ROOT / "config" / "profiles" / "public_synthetic_profile.json",
        contract_path=PROJECT_ROOT / "config" / "standardization" / "standard_31_contract.json",
        mapping_config_path=PROJECT_ROOT / "config" / "mapping" / "public_synthetic_mapping_v1.json",
        private_id_baseline_path=paths.private_id_baseline,
        private_mapping_workbook_path=paths.private_mapping_workbook,
    )


def run_upstream_stage(run_dir: Path, paths: Day9Paths) -> dict[str, Any]:
    return run_day5_upstream_rebuild(
        run_dir,
        profile_config_path=PROJECT_ROOT / "config" / "profiles" / "public_synthetic_profile.json",
        standard_contract_path=PROJECT_ROOT / "config" / "standardization" / "standard_31_contract.json",
        activity_contract_path=PROJECT_ROOT / "config" / "activity" / "activity_36_contract.json",
        third_party_contract_path=PROJECT_ROOT / "config" / "activity" / "third_party_20_contract.json",
        quality_config_path=PROJECT_ROOT / "config" / "qc" / "day5_quality_rules.json",
        interface_open_items_path=PROJECT_ROOT / "config" / "open_items" / "wp3_interface_open_items.json",
        standard_baseline_path=paths.standard_baseline,
        activity_baseline_path=paths.activity_baseline,
        third_party_baseline_path=paths.third_party_baseline,
    )


def run_factor_stage(
    run_dir: Path,
    paths: Day9Paths,
    *,
    mode: str,
    factor_input_path: Path | None = None,
) -> dict[str, Any]:
    if mode not in {"historical_simulation", "uploaded_factor"}:
        raise ValueError("因子模式必须是historical_simulation或uploaded_factor。")
    if mode == "uploaded_factor" and factor_input_path is None:
        raise ValueError("用户上传因子模式必须提供8字段CSV或XLSX。")
    return run_day6_factors_and_matching(
        run_dir,
        profile_config_path=PROJECT_ROOT / "config" / "profiles" / "public_synthetic_profile.json",
        activity_contract_path=PROJECT_ROOT / "config" / "activity" / "activity_36_contract.json",
        external_contract_path=PROJECT_ROOT / "config" / "factors" / "external_8_contract.json",
        d1_contract_path=PROJECT_ROOT / "config" / "factors" / "wp5_d1_45_contract.json",
        factor_config_path=PROJECT_ROOT / "config" / "factors" / "public_synthetic_factor.json",
        d2_contract_path=PROJECT_ROOT / "config" / "matching" / "wp5_d2_57_contract.json",
        route_contract_path=PROJECT_ROOT / "config" / "matching" / "wp5_d3_route_36_contract.json",
        factor_input_path=factor_input_path,
        historical_simulation=mode == "historical_simulation",
    )


def run_calculation_stage(run_dir: Path, paths: Day9Paths) -> dict[str, Any]:
    return run_day7_calculation_and_lineage(
        run_dir,
        profile_config_path=PROJECT_ROOT / "config" / "profiles" / "public_synthetic_profile.json",
        calculation_rules_path=PROJECT_ROOT / "config" / "calculation" / "day7_calculation_rules.json",
        activity_contract_path=PROJECT_ROOT / "config" / "activity" / "activity_36_contract.json",
        d1_contract_path=PROJECT_ROOT / "config" / "factors" / "wp5_d1_45_contract.json",
        d2_contract_path=PROJECT_ROOT / "config" / "matching" / "wp5_d2_57_contract.json",
        d3_contract_path=PROJECT_ROOT / "config" / "matching" / "wp5_d3_route_36_contract.json",
        d4_contract_path=PROJECT_ROOT / "config" / "calculation" / "wp5_d4_48_contract.json",
        d5_contract_path=PROJECT_ROOT / "config" / "calculation" / "wp5_d5_56_contract.json",
        frozen_lineage_contract_path=PROJECT_ROOT / "config" / "lineage" / "wp5_frozen_32_contract.json",
        extended_lineage_contract_path=PROJECT_ROOT / "config" / "lineage" / "demo_extended_lineage_contract.json",
    )


def run_export_stage(run_dir: Path, paths: Day9Paths) -> dict[str, Any]:
    return run_day8_export(
        run_dir,
        output_dir=paths.output_root / run_dir.name,
        artifact_work_dir=paths.artifact_work_root / run_dir.name,
        node_executable=paths.node_executable,
        node_modules_path=paths.node_modules,
        builder_script_path=PROJECT_ROOT / "scripts" / "build_day8_workbook.mjs",
        wp5_open_items_path=paths.wp5_open_items,
    )


def run_full_historical(source_path: Path, paths: Day9Paths) -> dict[str, Any]:
    return run_day8_full_pipeline(
        source_path,
        run_root=paths.run_root,
        output_root=paths.output_root,
        artifact_work_root=paths.artifact_work_root,
        node_executable=paths.node_executable,
        node_modules_path=paths.node_modules,
        builder_script_path=PROJECT_ROOT / "scripts" / "build_day8_workbook.mjs",
        alias_config_path=PROJECT_ROOT / "config" / "import" / "field_aliases.json",
        scope_config_path=PROJECT_ROOT / "config" / "scope" / "public_synthetic_scope.json",
        unit_config_path=PROJECT_ROOT / "config" / "cleaning" / "unit_mappings.json",
        profile_config_path=PROJECT_ROOT / "config" / "profiles" / "public_synthetic_profile.json",
        standard_contract_path=PROJECT_ROOT / "config" / "standardization" / "standard_31_contract.json",
        mapping_config_path=PROJECT_ROOT / "config" / "mapping" / "public_synthetic_mapping_v1.json",
        private_id_baseline_path=paths.private_id_baseline,
        private_mapping_workbook_path=paths.private_mapping_workbook,
        activity_contract_path=PROJECT_ROOT / "config" / "activity" / "activity_36_contract.json",
        third_party_contract_path=PROJECT_ROOT / "config" / "activity" / "third_party_20_contract.json",
        quality_config_path=PROJECT_ROOT / "config" / "qc" / "day5_quality_rules.json",
        interface_open_items_path=PROJECT_ROOT / "config" / "open_items" / "wp3_interface_open_items.json",
        standard_baseline_path=paths.standard_baseline,
        activity_baseline_path=paths.activity_baseline,
        third_party_baseline_path=paths.third_party_baseline,
        external_factor_contract_path=PROJECT_ROOT / "config" / "factors" / "external_8_contract.json",
        d1_contract_path=PROJECT_ROOT / "config" / "factors" / "wp5_d1_45_contract.json",
        factor_config_path=PROJECT_ROOT / "config" / "factors" / "public_synthetic_factor.json",
        d2_contract_path=PROJECT_ROOT / "config" / "matching" / "wp5_d2_57_contract.json",
        d3_contract_path=PROJECT_ROOT / "config" / "matching" / "wp5_d3_route_36_contract.json",
        calculation_rules_path=PROJECT_ROOT / "config" / "calculation" / "day7_calculation_rules.json",
        d4_contract_path=PROJECT_ROOT / "config" / "calculation" / "wp5_d4_48_contract.json",
        d5_contract_path=PROJECT_ROOT / "config" / "calculation" / "wp5_d5_56_contract.json",
        frozen_lineage_contract_path=PROJECT_ROOT / "config" / "lineage" / "wp5_frozen_32_contract.json",
        extended_lineage_contract_path=PROJECT_ROOT / "config" / "lineage" / "demo_extended_lineage_contract.json",
        wp5_open_items_path=paths.wp5_open_items,
    )


def run_cleaning_and_quality(run_dir: Path, paths: Day9Paths) -> dict[str, Any]:
    """Run the two business-facing cleaning/quality steps in their fixed order."""
    standard = run_standardize_stage(run_dir, paths)
    upstream = run_upstream_stage(run_dir, paths)
    return {
        "status": "PASS" if standard.get("status") == upstream.get("status") == "PASS" else "BLOCKED",
        "operation": "清洗、标准化与质量检查",
        "standardization": standard,
        "quality_and_activity": upstream,
        "run_directory": str(run_dir.resolve()),
    }


def run_calculation_and_export(
    run_dir: Path,
    paths: Day9Paths,
    *,
    factor_mode: str,
    factor_input_path: Path | None = None,
) -> dict[str, Any]:
    """Run factor matching, Decimal calculation and delivery as one business action."""
    factor = run_factor_stage(
        run_dir,
        paths,
        mode=factor_mode,
        factor_input_path=factor_input_path,
    )
    calculation = run_calculation_stage(run_dir, paths)
    export = run_export_stage(run_dir, paths)
    return {
        **export,
        "operation": "电芯碳核算与结果文件生成",
        "factor_matching": factor,
        "calculation": calculation,
        "run_directory": str(run_dir.resolve()),
    }


def safe_error(error: Exception) -> dict[str, Any]:
    if isinstance(error, PipelineUserError):
        return {"status": "BLOCKED", "error": error.to_dict()}
    return {
        "status": "BLOCKED",
        "error": {
            "stage": "DAY9_UI",
            "error_code": type(error).__name__,
            "message_cn": str(error),
            "impact": "阻断当前页面操作",
            "fix_suggestion": "检查输入、路径和前置阶段后重试。",
        },
    }


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.is_file():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def translate_status(value: Any) -> str:
    return display_status(value)


def translate_chemistry(value: Any) -> str:
    text = "" if value is None else str(value)
    return CHEMISTRY_TRANSLATIONS.get(text, text)


def chemistry_display_order(values: Iterable[Any]) -> list[str]:
    unique = {str(value) for value in values if value not in (None, "")}
    known = [name for name in CHEMISTRY_DISPLAY_PRIORITY if name in unique]
    configured = set(CHEMISTRY_DISPLAY_PRIORITY) | {"未知"}
    known.extend(sorted(unique - configured))
    if "未知" in unique:
        known.append("未知")
    return known


def translate_issue_codes(value: Any) -> str:
    return display_reason_code(value)


def business_stage_statuses(snapshot: dict[str, Any]) -> list[dict[str, str]]:
    return [
        {
            "处理步骤": BUSINESS_STAGE_LABELS.get(stage, stage),
            "运行状态": translate_status(status),
        }
        for stage, status in snapshot["technical_status"].items()
    ]


def load_carbon_result_rows(
    run_dir: Path,
    *,
    base_url: str,
    include_record_links: bool = False,
) -> list[dict[str, str]]:
    run = run_dir.expanduser().resolve()
    standard_rows = _read_csv_rows(run / "03_standardized" / "day4_standard_31_fields.csv")
    result_rows = _read_csv_rows(run / "10_output" / "day7_d5_end_to_end_56_fields.csv")
    standard_by_id = {row["Record_ID"]: row for row in standard_rows}
    clean_base = base_url.split("?", 1)[0].rstrip("/")
    result: list[dict[str, str]] = []
    for row in result_rows:
        record_id = row["Record_ID"]
        standard = standard_by_id.get(record_id, {})
        result.append(
            {
                "记录编号": (
                    f"{clean_base}/?run_id={run.name}&record_id={record_id}"
                    if include_record_links
                    else record_id
                ),
                "碳排放量（kgCO2e/年）": row.get("Emission_kgCO2e", ""),
                "化学体系": translate_chemistry(standard.get("Chemistry", "")),
                "排放因子（kgCO2e/kg）": row.get(
                    "EF_Value_Normalized_kgCO2e_per_kg", ""
                ),
                "供应商": standard.get("Supplier_Name", ""),
            }
        )
    return result


def load_processed_rows(run_dir: Path) -> list[dict[str, str]]:
    rows = _read_csv_rows(
        run_dir.expanduser().resolve() / "03_standardized" / "day4_standard_31_fields.csv"
    )
    return [
        {
            "记录编号": row.get("Record_ID", ""),
            "年度": row.get("Year", ""),
            "化学体系": translate_chemistry(row.get("Chemistry", "")),
            "供应商": row.get("Supplier_Name", ""),
            "项目号": row.get("Project_Code", ""),
            "电芯型号": row.get("Cell_Model", ""),
            "采购数量（PCS）": row.get("PCS", ""),
            "单件重量（g/PCS）": row.get("Unit_Weight_g", ""),
            "采购量（g/年）": row.get("Original_Activity_Value", ""),
            "原始物料信息": row.get("Product_Description", ""),
            "质量状态": translate_status(row.get("QC_Status", "")),
            "问题说明": translate_issue_codes(row.get("Issue_Code", "")),
            "来源工作表": row.get("Source_Sheet", ""),
            "来源行号": row.get("Source_Row", ""),
        }
        for row in rows
    ]


def load_quality_issue_rows(run_dir: Path) -> list[dict[str, str]]:
    rows = _read_csv_rows(
        run_dir.expanduser().resolve() / "04_qc" / "day5_quality_issue_records.csv"
    )
    return [
        {
            "记录编号": row.get("Record_ID", ""),
            "质量状态": translate_status(row.get("QC_Status", row.get("Status", ""))),
            "问题代码": translate_issue_codes(row.get("Issue_Code", "")),
            "问题字段": row.get("Field", row.get("Field_Name", "")),
            "问题原因": row.get("Message_CN", row.get("Issue_Message", "")),
            "来源行号": row.get("Source_Row", ""),
        }
        for row in rows
    ]


def load_cleaning_quality_summary(run_dir: Path) -> dict[str, Any]:
    run = run_dir.expanduser().resolve()
    day3_path = run / "02_scope_filter" / "day3_scope_summary.json"
    day4_path = run / "03_standardized" / "day4_standardization_summary.json"
    day5_path = run / "05_activity" / "day5_upstream_summary.json"
    day3 = _load_json(day3_path) if day3_path.is_file() else {}
    day4 = _load_json(day4_path) if day4_path.is_file() else {}
    day5 = _load_json(day5_path) if day5_path.is_file() else {}
    quality = day5.get("quality_status_counts", {})
    return {
        "原始记录数": int(day3.get("input_records", 0)),
        "进入核算范围": int(day3.get("candidate_records", 0)),
        "排除记录数": int(day3.get("excluded_records", 0)),
        "标准化记录数": int(day4.get("output_records", 0)),
        "质量通过": int(quality.get("PASS", 0)),
        "质量警告": int(quality.get("WARNING", 0)),
        "质量错误": int(quality.get("ERROR", 0)),
        "阻断记录数": int(day5.get("blocked_records", 0)),
        "活动数据记录数": int(day5.get("activity_records", 0)),
        "上游对账状态": translate_status(day5.get("g1a_gate_status", "NOT_RUN")),
    }


def _business_csv_bytes(rows: list[dict[str, str]]) -> bytes:
    if not rows:
        return b""
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=list(rows[0]))
    writer.writeheader()
    writer.writerows(rows)
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def _business_workbook_bytes(
    result_rows: list[dict[str, str]],
    processed_rows: list[dict[str, str]],
    issue_rows: list[dict[str, str]],
) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = [
        ("碳核算结果", result_rows),
        ("清洗后数据", processed_rows),
        ("质检问题", issue_rows),
    ]
    for title, rows in sheets:
        sheet = workbook.create_sheet(title)
        if not rows:
            sheet.append(["说明"])
            sheet.append(["当前运行没有该类记录"])
            continue
        headers = list(rows[0])
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for index, header in enumerate(headers, start=1):
            values = [str(header)] + [str(row.get(header, "")) for row in rows[:200]]
            sheet.column_dimensions[get_column_letter(index)].width = min(
                max(len(value) for value in values) + 2,
                42,
            )
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def business_download_artifacts(run_dir: Path, paths: Day9Paths) -> list[dict[str, Any]]:
    del paths  # The localized downloads are derived only from the active isolated run.
    result_rows = load_carbon_result_rows(
        run_dir,
        base_url="http://127.0.0.1",
        include_record_links=False,
    )
    processed_rows = load_processed_rows(run_dir)
    issue_rows = load_quality_issue_rows(run_dir)
    if not result_rows:
        return []
    workbook_data = _business_workbook_bytes(result_rows, processed_rows, issue_rows)
    known = [
        (
            "电芯碳核算工作簿",
            "电芯碳核算.xlsx",
            workbook_data,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
        ("电芯碳核算明细", "电芯碳核算.csv", _business_csv_bytes(result_rows), "text/csv"),
        ("电芯清洗后数据", "电芯清洗后数据.csv", _business_csv_bytes(processed_rows), "text/csv"),
    ]
    if issue_rows:
        known.append(
            ("电芯质检问题", "电芯质检问题.csv", _business_csv_bytes(issue_rows), "text/csv")
        )
    return [
        {
            "display_name": display_name,
            "download_name": download_name,
            "data": data,
            "mime": mime,
            "size_bytes": len(data),
        }
        for display_name, download_name, data, mime in known
    ]


def load_run_snapshot(run_dir: Path, paths: Day9Paths) -> dict[str, Any]:
    run = run_dir.expanduser().resolve()
    reports = {
        stage: _load_json(run / relative)
        for stage, relative in STAGE_REPORTS.items()
        if (run / relative).is_file()
    }
    upstream = reports.get("day5", {}).get("quality_status_counts", {})
    technical = {stage: report.get("status", "UNKNOWN") for stage, report in reports.items()}
    wp5 = {
        "排放因子匹配与路由": reports.get("day6", {}).get("stage_status", "NOT_RUN"),
        "碳核算与数据追溯": reports.get("day7", {}).get("stage_status", "NOT_RUN"),
        "结果文件生成与验证": reports.get("day8", {}).get("gate_status", "NOT_RUN"),
    }
    record_items = _read_csv_rows(run / "05_activity" / "day5_record_open_items.csv")
    interface_items = _read_csv_rows(run / "05_activity" / "day5_interface_open_items.csv")
    governance_items = _read_csv_rows(paths.wp5_open_items)
    summary = reports.get("day7", {})
    return {
        "run_id": run.name,
        "run_directory": str(run),
        "technical_status": technical,
        "upstream_quality": {
            "PASS": int(upstream.get("PASS", 0)),
            "WARNING": int(upstream.get("WARNING", 0)),
            "ERROR": int(upstream.get("ERROR", 0)),
        },
        "wp5_status": wp5,
        "open_items": {
            "WP3_RECORD": record_items,
            "WP3_INTERFACE": interface_items,
            "WP5_GOVERNANCE": governance_items,
        },
        "open_item_counts": {
            "WP3_RECORD": len(record_items),
            "WP3_INTERFACE": len(interface_items),
            "WP5_GOVERNANCE": len(governance_items),
        },
        "metrics": {
            "activity_total_kg": summary.get("activity_total_kg"),
            "ef_value": summary.get("ef_value"),
            "raw_total_emission_kgco2e": summary.get("raw_total_emission_kgco2e"),
            "official_six_decimal_total": summary.get("official_six_decimal_total"),
            "row_six_decimal_sum": summary.get("row_six_decimal_sum"),
            "rounding_reconciliation_difference": summary.get(
                "rounding_reconciliation_difference"
            ),
            "production_eligible": False,
        },
        "factor_mode": (
            "historical_simulation"
            if reports.get("day6", {}).get("simulation_flag") == "TRUE"
            else "uploaded_factor"
            if "day6" in reports
            else "not_run"
        ),
        "github_operations_performed": False,
    }


def current_run_artifacts(run_dir: Path, paths: Day9Paths) -> list[dict[str, Any]]:
    run = run_dir.expanduser().resolve()
    allowed_roots = (run, (paths.output_root / run.name).resolve())
    candidates = list(run.rglob("*.csv")) + list(run.rglob("*.json"))
    report_path = run / "10_output" / "day8_run_report.json"
    if report_path.is_file():
        report = _load_json(report_path)
        workbook = Path(str(report.get("workbook_path", "")))
        if workbook.is_file():
            candidates.append(workbook)
    artifacts: list[dict[str, Any]] = []
    seen: set[Path] = set()
    for candidate in candidates:
        resolved = candidate.resolve()
        if resolved in seen or not any(_is_relative_to(resolved, root) for root in allowed_roots):
            continue
        seen.add(resolved)
        suffix = resolved.suffix.lower()
        artifacts.append(
            {
                "name": resolved.name,
                "path": str(resolved),
                "mime": {
                    ".csv": "text/csv",
                    ".json": "application/json",
                    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                }.get(suffix, "application/octet-stream"),
                "size_bytes": resolved.stat().st_size,
            }
        )
    return sorted(artifacts, key=lambda item: (Path(item["path"]).suffix, item["name"]))


def path_config_for_display(paths: Day9Paths) -> dict[str, str]:
    values = asdict(paths)
    return {key: str(value) for key, value in values.items()}
