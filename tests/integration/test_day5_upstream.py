from __future__ import annotations

import csv
import json
import tempfile
import unittest
from copy import deepcopy
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook

from carbon_excel_pipeline.activity.day5_builders import (
    build_activity_records,
    build_third_party_records,
)
from carbon_excel_pipeline.activity.day5_pipeline import run_day5_upstream_rebuild
from carbon_excel_pipeline.qc.day5_quality import run_quality_checks


ROOT = Path(__file__).resolve().parents[2]
STANDARD_CONTRACT = ROOT / "config/standardization/standard_31_contract.json"
ACTIVITY_CONTRACT = ROOT / "config/activity/activity_36_contract.json"
THIRD_CONTRACT = ROOT / "config/activity/third_party_20_contract.json"
PRIVATE_PROFILE = ROOT / "config/profiles/public_synthetic_profile.json"
INTERFACE_ITEMS = ROOT / "config/open_items/wp3_interface_open_items.json"


class Day5UpstreamTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.standard_contract = json.loads(STANDARD_CONTRACT.read_text(encoding="utf-8"))
        self.activity_contract = json.loads(ACTIVITY_CONTRACT.read_text(encoding="utf-8"))
        self.third_contract = json.loads(THIRD_CONTRACT.read_text(encoding="utf-8"))
        self.quality_config = json.loads(
            (ROOT / "config/qc/day5_quality_rules.json").read_text(encoding="utf-8")
        )
        self.quality_config["expected_private_counts"] = {
            "PASS": 1,
            "WARNING": 1,
            "ERROR": 0,
        }
        self.quality_config["expected_private_record_open_items"] = 1
        self.quality_config["expected_private_interface_open_items"] = 5
        self.records = [self._record("RID-001", 11, mapped=True), self._record("RID-002", 12, mapped=False)]
        self.quality_path = self.root / "quality.json"
        self.quality_path.write_text(json.dumps(self.quality_config), encoding="utf-8")
        self.checked, _, _ = run_quality_checks(self.records, config=self.quality_config)
        self.activity, _ = build_activity_records(
            self.checked, activity_fields=self.activity_contract["fields"]
        )
        self.third = build_third_party_records(
            self.activity, contract=self.third_contract
        )
        total = sum(Decimal(row["Total_Weight_kg"]) for row in self.activity)
        self.quality_config["expected_private_activity_total_kg"] = str(total)
        self.quality_path.write_text(json.dumps(self.quality_config), encoding="utf-8")
        self.standard_baseline = self.root / "standard.xlsx"
        self.activity_baseline = self.root / "activity.xlsx"
        self.third_baseline = self.root / "third.xlsx"
        self._write_table(
            self.standard_baseline,
            "2025_Standard",
            self.standard_contract_fields,
            self.checked,
        )
        activity_baseline_rows = deepcopy(self.activity)
        activity_baseline_rows[0]["Activity_Diff_g"] = "0.00000004"
        activity_baseline_rows[0]["Activity_Diff_Rate"] = "0.0000000016"
        self._write_table(
            self.activity_baseline,
            "2025_Activity_Data",
            self.activity_contract["fields"],
            activity_baseline_rows,
        )
        self._write_table(
            self.third_baseline,
            "2025_Third_Party_Input",
            self.third_contract["fields"],
            self.third,
        )
        self.run_dir = self._make_run("RUN-D5", self.records)
        self.result = self._run(self.run_dir)

    def tearDown(self) -> None:
        self.temp.cleanup()

    @property
    def standard_contract_fields(self) -> list[str]:
        return [item["name"] for item in self.standard_contract["fields"]]

    def _record(self, record_id: str, source_row: int, *, mapped: bool) -> dict[str, object]:
        return {
            "Record_ID": record_id,
            "Year": 2025,
            "Source_File": "synthetic.xlsx",
            "Source_Sheet": "Input",
            "Source_Row": source_row,
            "Business_Unit": "Synthetic Unit",
            "Activity_Category": "Cell",
            "Activity_Category_Code": "SYN",
            "Supplier_Abbreviation": "SUP",
            "Supplier_Name": "Synthetic Supplier",
            "Supplier_ID": "SUP-1",
            "Supplier_Status": "FORMAL_MASTER",
            "Mapping_Source": "SYNTHETIC_MASTER",
            "Customer_Raw_Value": "Customer" if mapped else "UNKNOWN",
            "Customer_Name": "Customer" if mapped else "UNKNOWN",
            "Customer_ID": "CUST-1" if mapped else "UNKNOWN",
            "Customer_Mapping_Status": "CUSTOMER_MAPPED" if mapped else "UNMAPPED",
            "Customer_Mapping_Source": "SYNTHETIC_MASTER" if mapped else "NO_MAPPING",
            "Project_Code": "P1" if mapped else "P2",
            "Cell_Model": "123456" if mapped else "654321",
            "Material_ID": "UNKNOWN",
            "Chemistry": "LCO" if mapped else "UNKNOWN",
            "Product_Description": f"Synthetic cell {source_row}",
            "PCS": "10" if mapped else "4",
            "Unit_Weight_g": "2.5" if mapped else "5",
            "Original_Activity_Value": "25" if mapped else "20",
            "Original_Activity_Unit": "g/year",
            "Data_Status": "COMPLETE" if mapped else "PARTIAL",
            "Mapping_Status": "SUPPLIER_MAPPED",
            "QC_Status": "PASS" if mapped else "WARNING",
            "Issue_Code": "NONE" if mapped else "CUSTOMER_UNMAPPED;CHEMISTRY_UNKNOWN",
        }

    def _make_run(self, name: str, records: list[dict[str, object]]) -> Path:
        run = self.root / name
        for stage in ("03_standardized", "04_qc", "05_activity", "06_third_party_input"):
            (run / stage).mkdir(parents=True, exist_ok=True)
        with (run / "03_standardized/day4_standard_31_fields.csv").open(
            "w", encoding="utf-8-sig", newline=""
        ) as handle:
            writer = csv.DictWriter(handle, fieldnames=self.standard_contract_fields)
            writer.writeheader()
            writer.writerows(records)
        (run / "03_standardized/day4_standardization_summary.json").write_text(
            json.dumps({"status": "PASS"}), encoding="utf-8"
        )
        return run

    @staticmethod
    def _write_table(path: Path, sheet_name: str, fields: list[str], rows: list[dict[str, object]]) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet.append(["Synthetic baseline"])
        sheet.append([])
        sheet.append([])
        sheet.append(fields)
        for row in rows:
            sheet.append([row.get(field, "") for field in fields])
        workbook.save(path)
        workbook.close()

    def _run(self, run_dir: Path, *, third_baseline: Path | None = None) -> dict[str, object]:
        return run_day5_upstream_rebuild(
            run_dir,
            profile_config_path=PRIVATE_PROFILE,
            standard_contract_path=STANDARD_CONTRACT,
            activity_contract_path=ACTIVITY_CONTRACT,
            third_party_contract_path=THIRD_CONTRACT,
            quality_config_path=self.quality_path,
            interface_open_items_path=INTERFACE_ITEMS,
            standard_baseline_path=self.standard_baseline,
            activity_baseline_path=self.activity_baseline,
            third_party_baseline_path=third_baseline or self.third_baseline,
        )

    @staticmethod
    def _read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            return list(reader.fieldnames or []), list(reader)

    def test_d5_01_strict_31_36_20_schemas(self) -> None:
        activity_fields, _ = self._read_csv(self.run_dir / "05_activity/day5_activity_36_fields.csv")
        third_fields, _ = self._read_csv(self.run_dir / "06_third_party_input/day5_third_party_20_fields.csv")
        self.assertEqual(len(self.standard_contract_fields), 31)
        self.assertEqual(activity_fields, self.activity_contract["fields"])
        self.assertEqual(third_fields, self.third_contract["fields"])

    def test_d5_02_quality_counts_include_all_three_statuses(self) -> None:
        self.assertEqual(self.result["quality_status_counts"], {"PASS": 1, "WARNING": 1, "ERROR": 0})

    def test_d5_03_customer_and_chemistry_warnings_are_ordered(self) -> None:
        self.assertEqual(self.checked[1]["Issue_Code"], "CUSTOMER_UNMAPPED;CHEMISTRY_UNKNOWN")

    def test_d5_04_complete_record_passes_without_issue(self) -> None:
        self.assertEqual((self.checked[0]["QC_Status"], self.checked[0]["Issue_Code"]), ("PASS", "NONE"))

    def test_d5_05_required_value_missing_is_error(self) -> None:
        bad = deepcopy(self.records[0])
        bad["Product_Description"] = ""
        checked, _, _ = run_quality_checks([bad], config=self.quality_config)
        self.assertEqual(checked[0]["QC_Status"], "ERROR")

    def test_d5_06_duplicate_record_id_is_error(self) -> None:
        duplicate = deepcopy(self.records[0])
        checked, _, _ = run_quality_checks([self.records[0], duplicate], config=self.quality_config)
        self.assertTrue(all(row["QC_Status"] == "ERROR" for row in checked))

    def test_d5_07_non_positive_numeric_values_are_errors(self) -> None:
        bad = deepcopy(self.records[0])
        bad.update({"PCS": "0", "Unit_Weight_g": "-1", "Original_Activity_Value": "0"})
        checked, _, _ = run_quality_checks([bad], config=self.quality_config)
        self.assertIn("PCS_INVALID", checked[0]["Issue_Code"])
        self.assertIn("UNIT_WEIGHT_INVALID", checked[0]["Issue_Code"])

    def test_d5_08_activity_unit_is_case_and_space_sensitive(self) -> None:
        bad = deepcopy(self.records[0])
        bad["Original_Activity_Unit"] = "g/year "
        checked, _, _ = run_quality_checks([bad], config=self.quality_config)
        self.assertIn("ACTIVITY_UNIT_INVALID_EXACT", checked[0]["Issue_Code"])

    def test_d5_09_quantity_weight_reconciliation_mismatch_is_error(self) -> None:
        bad = deepcopy(self.records[0])
        bad["Original_Activity_Value"] = "24"
        checked, _, _ = run_quality_checks([bad], config=self.quality_config)
        self.assertIn("ACTIVITY_RECONCILIATION_MISMATCH", checked[0]["Issue_Code"])

    def test_d5_10_decimal_activity_conversions_are_exact(self) -> None:
        row = self.activity[0]
        self.assertEqual((row["Total_Weight_g"], row["Total_Weight_kg"], row["Total_Weight_t"]), ("25", "0.025", "0.000025"))
        self.assertEqual((row["Activity_Diff_g"], row["Activity_Diff_Rate"]), ("0", "0"))

    def test_d5_11_changing_pcs_changes_activity_result(self) -> None:
        changed = deepcopy(self.checked[0])
        changed["PCS"] = "20"
        activity, _ = build_activity_records([changed], activity_fields=self.activity_contract["fields"])
        self.assertEqual(activity[0]["Total_Weight_g"], "50")

    def test_d5_12_error_records_are_blocked_from_activity(self) -> None:
        bad = deepcopy(self.checked[0])
        bad["QC_Status"] = "ERROR"
        activity, blocked = build_activity_records([bad], activity_fields=self.activity_contract["fields"])
        self.assertEqual((len(activity), len(blocked)), (0, 1))

    def test_d5_13_third_party_warning_fields_are_explicit(self) -> None:
        row = self.third[1]
        self.assertEqual(row["Missing_Fields"], "Chemistry")
        self.assertEqual(row["Data_Quality_Flag"], "WARNING")
        self.assertIn("SOURCE_ISSUE=CUSTOMER_UNMAPPED;CHEMISTRY_UNKNOWN", row["Remarks"])

    def test_d5_14_open_items_remain_layered(self) -> None:
        _, record_items = self._read_csv(self.run_dir / "05_activity/day5_record_open_items.csv")
        _, interface_items = self._read_csv(self.run_dir / "05_activity/day5_interface_open_items.csv")
        self.assertEqual((len(record_items), len(interface_items)), (1, 5))
        self.assertNotEqual(set(record_items[0]), set(interface_items[0]))

    def test_d5_15_g1a_passes_and_accepts_registered_float_tail_semantics(self) -> None:
        self.assertEqual(self.result["g1a_gate_status"], "G1A_UPSTREAM_REBUILD_RECONCILED")
        gate = json.loads((self.run_dir / "05_activity/day5_g1a_gate.json").read_text(encoding="utf-8"))
        self.assertGreaterEqual(gate["comparisons"]["activity_36"]["registered_float_tail_semantic_matches"], 2)

    def test_d5_16_g1a_blocks_a_real_business_field_difference(self) -> None:
        changed_baseline = self.root / "third_changed.xlsx"
        self._write_table(changed_baseline, "2025_Third_Party_Input", self.third_contract["fields"], self.third)
        workbook = load_workbook(changed_baseline)
        workbook["2025_Third_Party_Input"].cell(5, 3).value = "DIFFERENT_PRODUCT_TYPE"
        workbook.save(changed_baseline)
        workbook.close()
        run_dir = self._make_run("RUN-D5-MISMATCH", self.records)
        result = self._run(run_dir, third_baseline=changed_baseline)
        self.assertEqual(result["status"], "FAIL")
        self.assertEqual(result["g1a_gate_status"], "G1A_UPSTREAM_REBUILD_BLOCKED")


if __name__ == "__main__":
    unittest.main()
