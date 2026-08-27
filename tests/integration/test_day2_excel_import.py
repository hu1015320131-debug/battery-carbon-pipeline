from __future__ import annotations

import hashlib
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from carbon_excel_pipeline.errors import PipelineUserError
from carbon_excel_pipeline.io.excel_importer import (
    inspect_excel_to_run,
    validate_xlsx_source,
)


ROOT = Path(__file__).resolve().parents[2]
ALIAS_CONFIG = ROOT / "config/import/field_aliases.json"
HEADERS = [
    "采购分类",
    "物料描述",
    "求和项:数量",
    "单位",
    "单PCS净重（G/PCS）",
    "采购量（g/年）",
]


def file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest().upper()


def build_workbook(
    path: Path,
    *,
    header_row: int = 1,
    headers: list[str] | None = None,
    with_formula_and_merge: bool = False,
    sheet_names: tuple[str, ...] = ("SheetA",),
) -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = sheet_names[0]
    for name in sheet_names[1:]:
        workbook.create_sheet(name)
    for worksheet in workbook.worksheets:
        for row in range(1, header_row):
            worksheet.cell(row=row, column=1, value=f"说明{row}")
        for column, value in enumerate(headers or HEADERS, start=1):
            worksheet.cell(row=header_row, column=column, value=value)
        worksheet.append(["合成类别", "合成描述", 10, "PCS", 5.5, 55])
        worksheet.append(["合成类别", "合成描述2", 20, "PCS", 6, 120])
        if with_formula_and_merge:
            worksheet.cell(row=header_row + 3, column=7, value="=C2*E2")
            worksheet.merge_cells(
                start_row=header_row + 4,
                start_column=1,
                end_row=header_row + 4,
                end_column=2,
            )
            worksheet.cell(row=header_row + 4, column=1, value="合成合并单元格")
    workbook.save(path)
    workbook.close()


class Day2ExcelInputTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.source_dir = self.root / "source"
        self.run_root = self.root / "runs"
        self.source_dir.mkdir()

    def tearDown(self) -> None:
        self.temp.cleanup()

    def _valid_source(self, **kwargs) -> Path:
        path = self.source_dir / "sample.xlsx"
        build_workbook(path, **kwargs)
        return path

    def _inspect(self, source: Path, **kwargs):
        return inspect_excel_to_run(
            source,
            run_root=self.run_root,
            alias_config_path=ALIAS_CONFIG,
            **kwargs,
        )

    def test_d2_01_accepts_standard_xlsx(self) -> None:
        source = self._valid_source()
        result = validate_xlsx_source(source, max_size_bytes=50 * 1024 * 1024)
        self.assertEqual(result["file_name"], "sample.xlsx")
        self.assertEqual(len(result["sha256"]), 64)

    def test_d2_02_rejects_unsupported_extension(self) -> None:
        source = self.source_dir / "sample.xls"
        source.write_bytes(b"not-an-xlsx")
        with self.assertRaises(PipelineUserError) as caught:
            validate_xlsx_source(source, max_size_bytes=50 * 1024 * 1024)
        self.assertEqual(caught.exception.error_code, "UNSUPPORTED_FILE_EXTENSION")

    def test_d2_03_rejects_file_above_configured_limit(self) -> None:
        source = self._valid_source()
        with self.assertRaises(PipelineUserError) as caught:
            validate_xlsx_source(source, max_size_bytes=1)
        self.assertEqual(caught.exception.error_code, "FILE_TOO_LARGE")

    def test_d2_04_rejects_corrupt_xlsx(self) -> None:
        source = self.source_dir / "corrupt.xlsx"
        source.write_bytes(b"plain text posing as xlsx")
        with self.assertRaises(PipelineUserError) as caught:
            validate_xlsx_source(source, max_size_bytes=50 * 1024 * 1024)
        self.assertEqual(caught.exception.error_code, "FILE_CORRUPT_NOT_XLSX_PACKAGE")

    def test_d2_05_rejects_encrypted_or_legacy_container(self) -> None:
        source = self.source_dir / "protected.xlsx"
        source.write_bytes(bytes.fromhex("D0CF11E0A1B11AE1") + b"synthetic")
        with self.assertRaises(PipelineUserError) as caught:
            validate_xlsx_source(source, max_size_bytes=50 * 1024 * 1024)
        self.assertEqual(caught.exception.error_code, "FILE_ENCRYPTED_OR_LEGACY_CONTAINER")

    def test_d2_06_copies_input_without_changing_source(self) -> None:
        source = self._valid_source()
        before_hash = file_hash(source)
        before_stat = source.stat()
        result = self._inspect(source)
        run_dir = Path(result["run_directory"])
        copied = run_dir / "00_input_copy" / source.name
        self.assertTrue(copied.is_file())
        self.assertEqual(file_hash(copied), before_hash)
        self.assertEqual(file_hash(source), before_hash)
        self.assertEqual(source.stat().st_mtime_ns, before_stat.st_mtime_ns)

    def test_d2_07_enumerates_sheets_rows_and_columns(self) -> None:
        source = self._valid_source(sheet_names=("业务一", "业务二"))
        result = self._inspect(source)
        inventory_path = Path(result["run_directory"]) / "01_import/sheet_inventory.json"
        inventory = json.loads(inventory_path.read_text(encoding="utf-8"))
        self.assertEqual([item["sheet_name"] for item in inventory], ["业务一", "业务二"])
        self.assertTrue(all(item["physical_row_count"] == 3 for item in inventory))
        self.assertTrue(all(item["column_count"] == 6 for item in inventory))

    def test_d2_08_detects_header_within_first_ten_rows(self) -> None:
        source = self._valid_source(header_row=7)
        result = self._inspect(source)
        header_path = Path(result["run_directory"]) / "01_import/header_detection.json"
        header = json.loads(header_path.read_text(encoding="utf-8"))[0]
        self.assertTrue(header["detected"])
        self.assertEqual(header["header_row"], 7)

    def test_d2_09_maps_all_six_fields_with_aliases(self) -> None:
        aliases = [
            "采购类别",
            "产品描述",
            "求和项：数量",
            "采购单位",
            "单PCS净重(g/PCS)",
            "年度采购量(g/年)",
        ]
        source = self._valid_source(headers=aliases)
        result = self._inspect(source)
        mapping_path = Path(result["run_directory"]) / "01_import/field_mapping_preview.json"
        mapping = json.loads(mapping_path.read_text(encoding="utf-8"))[0]
        self.assertEqual(mapping["match_count"], 6)
        self.assertTrue(all(item["status"] == "AUTO_MATCHED" for item in mapping["mapping_preview"]))

    def test_d2_10_reports_formulas_and_merged_cells(self) -> None:
        source = self._valid_source(with_formula_and_merge=True)
        result = self._inspect(source)
        inventory_path = Path(result["run_directory"]) / "01_import/sheet_inventory.json"
        inventory = json.loads(inventory_path.read_text(encoding="utf-8"))[0]
        self.assertEqual(inventory["formula_count"], 1)
        self.assertEqual(inventory["merged_cell_count"], 1)

    def test_d2_11_preview_is_read_only_and_capped_at_twenty_rows(self) -> None:
        source = self._valid_source()
        before_hash = file_hash(source)
        result = self._inspect(source, preview_rows=20)
        preview_path = Path(result["run_directory"]) / "01_import/input_preview.json"
        preview = json.loads(preview_path.read_text(encoding="utf-8"))[0]
        self.assertLessEqual(len(preview["rows"]), 20)
        self.assertEqual(preview["rows"][0]["values"], HEADERS)
        self.assertEqual(file_hash(source), before_hash)

    def test_d2_12_blocks_unrecognizable_header_with_safe_message(self) -> None:
        source = self._valid_source(headers=["A", "B", "C", "D", "E", "F"])
        result = self._inspect(source)
        self.assertEqual(result["status"], "BLOCKED")
        mapping_path = Path(result["run_directory"]) / "01_import/field_mapping_preview.json"
        mapping = json.loads(mapping_path.read_text(encoding="utf-8"))[0]
        error = mapping["blocking_errors"][0]
        self.assertEqual(error["error_code"], "HEADER_REQUIRED_FIELDS_NOT_FOUND")
        self.assertIn("重试", error["fix_suggestion"])


if __name__ == "__main__":
    unittest.main()
