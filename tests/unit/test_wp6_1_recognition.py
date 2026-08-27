from __future__ import annotations

import json
import unittest
from pathlib import Path

from openpyxl import Workbook

from carbon_excel_pipeline.io.recognition import recognize_sheet, recognize_workbook
from carbon_excel_pipeline.io.recognition_models import RecognitionStatus
from carbon_excel_pipeline.io.semantic_registry import (
    SemanticFieldRegistry,
    normalize_header,
    parse_header_unit,
)


ROOT = Path(__file__).resolve().parents[2]
CONFIG = json.loads(
    (ROOT / "config/import/field_aliases.json").read_text(encoding="utf-8")
)


def sheet_with_rows(rows: list[list[object]]):
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    return workbook, sheet


class WP61RecognitionUnitTests(unittest.TestCase):
    def setUp(self) -> None:
        self.registry = SemanticFieldRegistry(CONFIG)

    def test_header_normalization_handles_spacing_newlines_and_unnamed(self) -> None:
        self.assertEqual(normalize_header(" 物料\n描述 \t"), "物料 描述")
        self.assertEqual(normalize_header("ＰＲＯＤＵＣＴ　ＤＥＳＣＲＩＰＴＩＯＮ"), "product description")
        self.assertEqual(normalize_header("Unnamed: 3"), "")
        self.assertEqual(
            self.registry.map_header("物料\n描述", column_index=1).semantic_field,
            "Product_Description",
        )

    def test_unit_parser_uses_controlled_variants(self) -> None:
        cases = {
            "年度购买原料量（T/年）": "t/year",
            "Activity (t/year)": "t/year",
            "LCA排放因子（KGCO2/KG）": "kgCO2e/kg",
            "EF (kgCO2e/kg)": "kgCO2e/kg",
            "GHG排放量（TCO2/年）": "tCO2e/year",
            "单PCS净重（g/PCS）": "g/PCS",
        }
        for raw, expected in cases.items():
            with self.subTest(raw=raw):
                self.assertEqual(parse_header_unit(raw), expected)

    def test_alias_mapping_supports_2024_and_2025_purchase_category(self) -> None:
        for raw in ("采购分类", "公司外购原料和辅料名称"):
            with self.subTest(raw=raw):
                mapped = self.registry.map_header(raw, column_index=1)
                self.assertEqual(mapped.semantic_field, "Purchase_Category")

    def test_unknown_fields_are_preserved(self) -> None:
        workbook, sheet = sheet_with_rows(
            [["采购分类", "物料描述", "内部备注"], ["电芯", "SYNA", "保留"]]
        )
        result = recognize_sheet(sheet, sheet_index=0, registry=self.registry)
        workbook.close()
        self.assertEqual(result.status, RecognitionStatus.RECOGNIZED_WITH_WARNING)
        self.assertIn("内部备注", result.unmapped_fields)

    def test_duplicate_semantic_mapping_is_ambiguous(self) -> None:
        workbook, sheet = sheet_with_rows(
            [["采购分类", "物料描述", "产品描述"], ["电芯", "SYNA-A", "SYNA-B"]]
        )
        result = recognize_sheet(sheet, sheet_index=0, registry=self.registry)
        workbook.close()
        self.assertEqual(result.status, RecognitionStatus.AMBIGUOUS)
        self.assertIn("Product_Description", result.duplicate_mappings)
        self.assertTrue(
            any(item.code == "DUPLICATE_SEMANTIC_MAPPING" for item in result.warnings)
        )

    def test_header_row_offsets_are_detected(self) -> None:
        for offset in (1, 2, 5):
            with self.subTest(offset=offset):
                rows = [[f"说明{index}"] for index in range(1, offset)]
                rows.extend(
                    [["采购分类", "物料描述"], ["电芯", "SYNA"]]
                )
                workbook, sheet = sheet_with_rows(rows)
                result = recognize_sheet(sheet, sheet_index=0, registry=self.registry)
                workbook.close()
                self.assertEqual(result.selected_header_row, offset)

    def test_multi_sheet_scan_keeps_unrecognized_sheet(self) -> None:
        workbook = Workbook()
        unknown = workbook.active
        unknown.title = "说明"
        unknown.append(["A", "B"])
        unknown.append([1, 2])
        known = workbook.create_sheet("数据")
        known.append(["采购分类", "物料描述"])
        known.append(["电芯", "SYNA"])
        result = recognize_workbook(
            workbook,
            config=CONFIG,
            workbook_name="synthetic.xlsx",
            input_fingerprint="0" * 64,
        )
        workbook.close()
        self.assertEqual(result.sheet_count, 2)
        self.assertEqual(result.best_candidate_sheet, "数据")
        self.assertEqual(result.sheets[0].status, RecognitionStatus.UNRECOGNIZED)

    def test_equal_header_candidates_are_ambiguous(self) -> None:
        workbook, sheet = sheet_with_rows(
            [
                ["采购分类", "物料描述"],
                ["采购分类", "物料描述"],
                ["电芯", "SYNA"],
            ]
        )
        result = recognize_sheet(sheet, sheet_index=0, registry=self.registry)
        workbook.close()
        self.assertEqual(result.status, RecognitionStatus.AMBIGUOUS)
        self.assertTrue(
            any(item.code == "AMBIGUOUS_HEADER_CANDIDATES" for item in result.warnings)
        )

    def test_only_product_description_is_unrecognized(self) -> None:
        workbook, sheet = sheet_with_rows([["物料描述"], ["SYNA"]])
        result = recognize_sheet(sheet, sheet_index=0, registry=self.registry)
        workbook.close()
        self.assertEqual(result.status, RecognitionStatus.UNRECOGNIZED)

    def test_missing_activity_unit_is_warning_not_guess(self) -> None:
        workbook, sheet = sheet_with_rows(
            [["年度购买原料量", "物料描述"], [1.2, "SYNA"]]
        )
        result = recognize_sheet(sheet, sheet_index=0, registry=self.registry)
        workbook.close()
        activity = next(
            item for item in result.mappings if item.semantic_field == "Reported_Activity_Value"
        )
        self.assertIsNone(activity.detected_unit)
        self.assertEqual(activity.warning_code, "UNIT_MISSING_OR_UNKNOWN")
        self.assertEqual(result.status, RecognitionStatus.RECOGNIZED_WITH_WARNING)


if __name__ == "__main__":
    unittest.main()
