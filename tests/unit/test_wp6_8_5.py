import json
import hashlib
from pathlib import Path

from openpyxl import Workbook

from carbon_excel_pipeline.io.header_detector import load_alias_config
from carbon_excel_pipeline.wp6_8_4.attribute_enrichment import enrich_canonical_with_attributes
from carbon_excel_pipeline.wp6_8_4.file_roles import classify_workbook_role
from carbon_excel_pipeline.wp6_8_4.input_set import ROLE_LEDGER, input_set_sha256
from carbon_excel_pipeline.wp6_8_5.cell_scope import apply_cell_scope_to_run
from carbon_excel_pipeline.wp6_8_5.current_run import persist_current_run, restore_current_run
from carbon_excel_pipeline.wp6_8_5.ledger_reference import extract_cell_ledger_evidence
from carbon_excel_pipeline.wp6_8_5.public_factor import apply_public_cell_factor


ROOT = Path(__file__).resolve().parents[2]
ALIAS = load_alias_config(ROOT / "config" / "import" / "field_aliases.json")


def _write_json(path: Path, payload) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")


def _recognized(sheet: str, rows: list[tuple[int, str]]) -> dict:
    return {
        "sheet_name": sheet,
        "workbook_name": "main.xlsx",
        "column_mappings": [{"semantic_field": "Purchase_Category"}],
        "record_count": len(rows),
        "records": [
            {
                "Source_Row": source_row,
                "values": {
                    "Purchase_Category": category,
                    "Product_Description": f"物料{source_row}",
                    "Quantity_PCS": 1,
                    "Unit_Weight": 1000,
                },
            }
            for source_row, category in rows
        ],
    }


def test_cell_boundary_is_applied_before_capability_input(tmp_path: Path):
    import_dir = tmp_path / "RUN-A" / "01_import"
    _write_json(import_dir / "recognized_records.json", _recognized("二部", [(2, "电芯.聚合物电芯"), (3, "包装")]))
    _write_json(
        import_dir / "recognized_records_by_sheet.json",
        {"sheets": [_recognized("一部", [(2, "电芯"), (3, "五金件")])]},
    )
    summary = apply_cell_scope_to_run(tmp_path / "RUN-A")
    assert summary["raw_record_count"] == 4
    assert summary["cell_record_count"] == 2
    assert summary["capability_denominator"] == 1
    primary = json.loads((import_dir / "recognized_records.json").read_text(encoding="utf-8"))
    assert len(primary["records"]) == 1
    assert primary["records"][0]["values"]["Business_Unit"] == "合成二部"
    assert (import_dir / "recognized_records_raw.json").is_file()


def _ledger_workbook(path: Path) -> Path:
    workbook = Workbook()
    activity = workbook.active
    activity.title = "表3活动数据管理表"
    activity.append(["编号", "排放源", "活动数据", "单位"])
    activity.append([1, "电芯", 100, "kg"])
    factor = workbook.create_sheet("表4排放系数管理表")
    factor.append(["编号", "活动/设施", "排放系数", "单位"])
    factor.append([1, "电芯", 1.250000, "kg CO2/kg"])
    emission = workbook.create_sheet("表5排放量计算表")
    emission.append(["基本数据"])
    emission.append([])
    emission.append(["编号", "活动/设施", *([None] * 33), "年总排放量", "年总排放量"])
    emission.append([None, None, *([None] * 33), "kg CO2当量", "t CO2当量"])
    emission.append([1, "电芯", *([None] * 33), 125.0, 0.125])
    workbook.save(path)
    return path


def test_ledger_role_and_cell_factor_evidence_use_sheet_semantics(tmp_path: Path):
    path = _ledger_workbook(tmp_path / "not-a-ledger-name.xlsx")
    classified = classify_workbook_role(path, ALIAS)
    evidence = extract_cell_ledger_evidence(path)
    assert classified["role"] == ROLE_LEDGER
    assert evidence["status"] == "PASS"
    assert evidence["EF_Value"] == "1.250000"
    assert evidence["Historical_Activity_kg"] == "100"
    assert evidence["Historical_Emission_tCO2e"] == "0.125"


def test_attribute_exact_match_is_year_and_business_unit_scoped():
    canonical = [
        {"Year": "2025", "Business_Unit": "合成二部", "Product_Description": "电芯A"},
        {"Year": "2025", "Business_Unit": "合成一部", "Product_Description": "电芯A"},
    ]
    attributes = [
        {
            "Attribute_Year": "2025",
            "Attribute_Business_Unit": "合成二部",
            "Product_Description": "电芯A",
            "Chemistry": "LFP",
            "Supplier": "供应商甲",
            "Attribute_Source_Row": row,
        }
        for row in (2, 3)
    ]
    output, summary = enrich_canonical_with_attributes(canonical, attributes)
    assert output[0]["Attribute_Match_Status"] == "MATCHED"
    assert output[0]["Chemistry"] == "LFP"
    assert output[0]["Attribute_Match_Key"] == "Product_Description=电芯A"
    assert output[1]["Attribute_Match_Status"] == "UNMATCHED"
    assert not output[1].get("Chemistry")
    assert summary == {"matched": 1, "unmatched": 1, "ambiguous": 0, "attribute_rows": 2, "message": "部分补充属性无法可靠匹配，请确认关联字段。"}


def test_public_factor_applies_to_both_units_but_never_non_cell():
    evidence = {"status": "PASS", "EF_Value": "1.250000", "Source_File": "ledger.xlsx", "Factor_Source_Sheet": "表4", "Factor_Source_Cell": "F57", "Source_SHA256": "ABC"}
    rows = [
        {"Business_Unit": "合成一部", "Purchase_Category": "电芯", "Activity_Data_kg": "2", "Blocking_Codes": "FACTOR_NOT_AVAILABLE|BOUNDARY_POLICY_NOT_AVAILABLE"},
        {"Business_Unit": "合成二部", "Purchase_Category": "电芯.聚合物电芯", "Activity_Data_kg": "3"},
        {"Business_Unit": "合成一部", "Purchase_Category": "包装", "Activity_Data_kg": "4"},
    ]
    output, summary = apply_public_cell_factor(rows, evidence)
    assert summary["applied_records"] == 2
    assert output[0]["EF_Value"] == output[1]["EF_Value"] == "1.250000"
    assert output[0]["Emission_kgCO2e"] == "2.500000"
    assert output[0]["Calculation_QC"] == "PASS"
    assert output[0]["Simulation_Flag"] == "TRUE"
    assert output[0]["Production_Eligible"] == "FALSE"
    assert "EF_Value" not in output[2]


def test_completed_current_run_restores_only_with_matching_input_fingerprints(tmp_path: Path):
    run = tmp_path / "RUN-A"
    copy = run / "00_input_copy" / "main.xlsx"
    copy.parent.mkdir(parents=True)
    copy.write_bytes(b"primary")
    files = [{"name": "main.xlsx", "sha256": hashlib.sha256(b"primary").hexdigest().upper(), "role": "主核算数据"}]
    _write_json(
        run / "e2e_run_summary.json",
        {
            "Run_ID": "RUN-A",
            "Status": "PASS",
            "Record_ID_Schema_Version": "RID_V2",
            "Input_Set_SHA256": input_set_sha256(files),
            "Input_Files": files,
        },
    )
    _write_json(run / "pipeline_stage_status.json", {"stages": {"COMPLETED": "PASS"}})
    canonical = run / "08_download" / "canonical_results.csv"
    canonical.parent.mkdir(parents=True)
    canonical.write_text("Record_ID\nA\n", encoding="utf-8")
    assert persist_current_run(run)
    restored = restore_current_run(tmp_path)
    assert restored and restored["Current_Run_ID"] == "RUN-A"
    assert restored["Record_ID_Schema_Version"] == "RID_V2"
    copy.write_bytes(b"changed")
    assert restore_current_run(tmp_path) is None


def test_legacy_current_run_is_not_restored_or_converted(tmp_path: Path):
    run = tmp_path / "RUN-OLD"
    copy = run / "00_input_copy" / "main.xlsx"
    copy.parent.mkdir(parents=True)
    copy.write_bytes(b"primary")
    digest = hashlib.sha256(b"primary").hexdigest().upper()
    files = [{"name": "main.xlsx", "sha256": digest, "role": "主核算数据"}]
    _write_json(
        run / "e2e_run_summary.json",
        {
            "Run_ID": "RUN-OLD",
            "Status": "PASS",
            "Input_Set_SHA256": input_set_sha256(files),
            "Input_Files": files,
        },
    )
    _write_json(run / "pipeline_stage_status.json", {"stages": {"COMPLETED": "PASS"}})
    canonical = run / "08_download" / "canonical_results.csv"
    canonical.parent.mkdir(parents=True)
    canonical.write_text("Record_ID\n2025-DY2-SYNA-DX000001\n", encoding="utf-8")
    _write_json(
        tmp_path / "current_run.json",
        {
            "schema_version": "WP6_8_5_CURRENT_RUN_POINTER_V1",
            "Current_Run_ID": "RUN-OLD",
            "Current_Run_Root": str(run),
            "Input_Set_SHA256": input_set_sha256(files),
            "Input_Files": files,
        },
    )
    restored = restore_current_run(tmp_path)
    assert restored and restored["Restore_Status"] == "LEGACY_RECORD_ID_SCHEMA"
    assert "重新运行" in restored["Message"]
